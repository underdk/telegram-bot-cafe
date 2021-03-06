import telebot
import mysql.connector
from telebot import types
import datetime
import logging
import uuid
import openpyxl
import time

db = mysql.connector.connect(
    host='localhost',
    user='root',
    passwd='undertalefanatic_45',
    port='3306',
    database='mydatabase'
)
#mysql://
# ba67d3559da4eb
# :
# 50936339
# @
# eu-cdbr-west-03.cleardb.net
# /heroku_745e09c97ab9640?reconnect=true
cursor = db.cursor(buffered=True)
#cursor.execute('UPDATE users SET phone_number = NULL WHERE user_id = 724952483')
#cursor.execute('CREATE DATABASE mydatabase')
#cursor.execute("ALTER TABLE productlist ADD COLUMN desc_ VARCHAR(255)")
#cursor.execute('CREATE TABLE test (user_id INT, image BLOB)')
#cursor.execute('CREATE TABLE users (first_name VARCHAR(255), last_name VARCHAR(255), user_id INT UNIQUE, phone_number VARCHAR(255), address VARCHAR(255), value INT, hour INT, min INT, role VARCHAR(255), key_admin VARCHAR(255))')
#cursor.execute("ALTER TABLE users ADD COLUMN (step VARCHAR (255))")
#cursor.execute('CREATE TABLE categorylist (id INT AUTO_INCREMENT PRIMARY KEY, name VARCHAR (255), seen VARCHAR(255))')
#cursor.execute('CREATE TABLE productlist (id INT AUTO_INCREMENT PRIMARY KEY, product VARCHAR (255))')
#cursor.execute('DROP TABLE users')
#cursor.execute('DROP TABLE cart')
#cursor.execute('CREATE TABLE usercount (id INT AUTO_INCREMENT PRIMARY KEY, first_name VARCHAR(255), last_name VARCHAR(255), user_id INT UNIQUE)')
#cursor.execute('INSERT INTO categorylist (name, seen) VALUES ("sales", "sales")')
db.commit()


bot = telebot.TeleBot('1269314724:AAG0rleabFDoZ-y-VC7-E65RxiaRl06HEK8')
dad_id = 826867214



start_text = "Добро пожаловать в онлайн-пиццерию Freddy Fazbear's Pizza! Если вы желаете что-нибудь заказать нажмите на кнопку меню. Ваши заказы будут сохраняться в корзине."
logo = 'https://imbt.ga/vfcM0yHFjI'


@bot.message_handler(commands=['start', 'home'])     #Начало
def start_message(message):
    try:
        start_buttons = types.InlineKeyboardMarkup(2)
        menu = types.InlineKeyboardButton(text='Меню \ud83d\udcd6', callback_data='menu')
        basket = types.InlineKeyboardButton(text='Корзина \ud83d\uded2', callback_data='cart')
        sales = types.InlineKeyboardButton('\ud83d\udd25Акции\ud83d\udd25', callback_data='sales')
        start_buttons.add(menu, basket, sales)
        cart = types.KeyboardButton(text='Корзина \ud83d\uded2')
        bot.send_message(message.chat.id, parse_mode='Markdown', text='[ ](https://imbt.ga/vfcM0yHFjI)' + start_text, reply_markup=start_buttons)
        user_id = message.chat.id
        first_name = message.chat.first_name
        last_name = message.chat.last_name

        sql = "INSERT INTO users (first_name, last_name, user_id, role, key_admin) \
                                              VALUES (%s, %s, %s, %s, %s)"
        val = (first_name, last_name, user_id, 'user', str(uuid.uuid4()))
        cursor.execute(sql, val)
        sql2 = 'INSERT INTO usercount (first_name, last_name, user_id) VALUES (%s, %s, %s)'
        val2 = (first_name, last_name, user_id)
        cursor.execute(sql2, val2)
        sql3 = "INSERT INTO calldata (user_id) \
                                                      VALUES (%s)"
        val3 = (user_id, )
        cursor.execute(sql3, val3)
        db.commit()
        print (first_name, last_name, 'user registered with user ID -', user_id)
    except Exception:
        print(first_name, last_name, user_id, 'has logged')


@bot.message_handler(content_types=['text', 'contact', 'photo'])
def order_process(message):
    if message.text and (message.text != 'Изменить адрес' and message.text != 'Отмена\u274c') and not message.text.startswith('/'):
        alltimemarkup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        alltimebutton = types.KeyboardButton('Отмена\u274c')
        alltimemarkup.add(alltimebutton)
        sel = 'SELECT message FROM calldata WHERE user_id = %s'
        selval = (message.chat.id,)
        cursor.execute(sel, selval)
        message_fetch = cursor.fetchone()
        sel_del = 'SELECT step FROM users WHERE user_id = %s'
        sel_delval = (message.chat.id, )
        cursor.execute(sel_del, sel_delval)
        step_fetch = cursor.fetchone()
        if step_fetch[0] == 'address':
            bot.edit_message_text(text='Способ доставки - курьер' + '\nАдрес заказчика - ' + message.text + '\n...', chat_id=message.chat.id, message_id=message_fetch[0])
            add_address = 'UPDATE users SET address = %s WHERE user_id = %s'
            add_val = (message.text, message.chat.id)
            cursor.execute(add_address, add_val)
            db.commit()
            phone_check = 'SELECT phone_number FROM users WHERE user_id = %s'
            phone_check2 = (message.chat.id,)
            cursor.execute(phone_check, phone_check2)
            phone_number = cursor.fetchone()
            if phone_number[0] == None:
                bot.delete_message(message.chat.id, message_fetch[0] + 2)
                phone_step = 'UPDATE users SET step = %s WHERE user_id = %s'
                phone_vals = ('phone_number', message.chat.id)
                cursor.execute(phone_step, phone_vals)
                db.commit()
                contact = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1, one_time_keyboard=True)
                contact_button = types.KeyboardButton(text='Отправить контакты', request_contact=True)
                alltimebutton = types.KeyboardButton('Отмена\u274c')
                contact.add(contact_button, alltimebutton)
                bot.send_message(text='Чтобы продолжить вы должны отправить нам ваши контактные данные для дальнейших услуг',
                                      chat_id=message.chat.id, reply_markup=contact)
            else:
                phone_step = 'UPDATE users SET step = %s WHERE user_id = %s'
                phone_vals = ('paymethod', message.chat.id)
                cursor.execute(phone_step, phone_vals)
                db.commit()
                pay_method = types.InlineKeyboardMarkup()
                pay_card = types.InlineKeyboardButton('Перевод картой', callback_data='card')
                pay_money = types.InlineKeyboardButton('Оплата наличными', callback_data='money')
                kaspigold = types.InlineKeyboardButton('Kaspi Gold', callback_data='kaspi')
                pay_method.add(kaspigold, pay_card, pay_money)
                bot.delete_message(message.chat.id, message_fetch[0] + 2)
                bot.send_message(message.chat.id, 'Выберите способ оплаты:', reply_markup=pay_method)
        elif step_fetch[0] == 'money':
            if not message.text.isdigit():
                bot.send_message(message.chat.id, 'Вы неправильно ввели сумму, повторите еще раз (Вводить только цифрами)')
                return
            else:
                select = 'SELECT price FROM cart WHERE cart_id = %s'
                select2 = (message.chat.id,)
                cursor.execute(select, select2)
                fetch_price = cursor.fetchall()
                price_list = []
                for i in range(len(fetch_price)):
                    sc = fetch_price[i]
                    price_list.append(sc[0])
                if int(message.text) < sum(price_list):
                    bot.send_message(message.chat.id,
                                              'Введенной вами суммы не хватит, чтобы оплатить заказ, повторите еще раз')
                    return
                else:
                    sql = 'SELECT * FROM orders WHERE process = %s'
                    val = ('finished', )
                    cursor.execute(sql, val)
                    num = cursor.rowcount
                    sql = 'UPDATE orders SET № = %s WHERE user_id = %s AND process = %s'
                    val = (num + 1, message.chat.id, 'ordering')
                    cursor.execute(sql, val)
                    db.commit()
                    sql = 'UPDATE orders SET moneychange = %s WHERE user_id = %s AND process = %s'
                    val = (int(message.text), message.chat.id, 'ordering')
                    cursor.execute(sql, val)
                    db.commit()
                    conf = 'UPDATE users SET step = %s WHERE user_id = %s'
                    conf_val = ('confirm', message.chat.id)
                    db.commit()
                    sql2 = 'SELECT product, amount, price FROM cart WHERE cart_id = %s'
                    val2 = (message.chat.id,)
                    cursor.execute(sql2, val2)
                    fetch = cursor.fetchall()
                    sel = 'SELECT message FROM calldata WHERE user_id = %s'
                    selval = (message.chat.id,)
                    cursor.execute(sel, selval)
                    message_fetch = cursor.fetchone()
                    sql_add = 'SELECT address FROM users WHERE user_id = %s'
                    sql_val = (message.chat.id,)
                    cursor.execute(sql_add, sql_val)
                    add_fetch = cursor.fetchone()
                    phone_check = 'SELECT phone_number FROM users WHERE user_id = %s'
                    phone_check2 = (message.chat.id,)
                    cursor.execute(phone_check, phone_check2)
                    phone_number = cursor.fetchone()
                    select = 'SELECT price FROM cart WHERE cart_id = %s'
                    select2 = (message.chat.id,)
                    cursor.execute(select, select2)
                    fetch_price = cursor.fetchall()
                    price_list = []
                    orders = ''
                    for i in range(len(fetch_price)):
                        sc = fetch_price[i]
                        price_list.append(sc[0])
                        orders += '\n' + str(i+1) + '. ' + (fetch[i])[0] + ' ' +str((fetch[i])[1]) + ' шт. - ' + str((fetch[i])[2]) + 'тг.'
                    bot.delete_message(message.chat.id, message_fetch[0])
                    confirmation = types.InlineKeyboardMarkup()
                    confirm = types.InlineKeyboardButton('Подтвердить\u2705', callback_data='confirm')
                    uhhhnogetback = types.InlineKeyboardButton('Отмена\u274c', callback_data='bacc')
                    confirmation.add(confirm, uhhhnogetback)
                    remove = types.ReplyKeyboardRemove()
                    bot.send_message(message.chat.id, 'Все верно?', reply_markup=remove)
                    bot.send_message(parse_mode='Markdown', text='*Заказ №' + str(num+1) + '*\n' + '\nСпособ доставки - курьер' + '\nАдрес заказчика - ' + add_fetch[0] + '\nСпособ оплаты - наличными' + '\nСумма оплаты - ' + message.text + 'тг.' + '\nСдача - ' + str(int(message.text) - sum(price_list)) + 'тг.' + '\nИмя - ' + message.chat.first_name + '\nНомер телефона - ' + str(phone_number[0]) + '\n*Ваш заказ:*' + orders + '\n*Общая стоимость - *' + '*' + str(sum(price_list)) + 'тг.*', chat_id=message.chat.id, reply_markup=confirmation)
        elif step_fetch[0] == 'hours':
            if not message.text.isdigit():
                bot.send_message(message.chat.id, 'Вы неправильно ввели часы, повторите еще раз (Вводить только цифрами от 0 до 23)')
                return
            else:
                if int(message.text) < 0 or int(message.text) > 23:
                    bot.send_message(message.chat.id, 'Вы неправильно ввели часы, повторите еще раз (Вводить только цифрами от 0 до 23)')
                else:
                    add_hr = 'UPDATE users SET hour = %s WHERE user_id = %s'
                    hrvam = (int(message.text), message.chat.id)
                    cursor.execute(add_hr, hrvam)
                    mins = 'UPDATE users SET step = %s WHERE user_id = %s'
                    valmins = ('minutes', message.chat.id)
                    cursor.execute(mins, valmins)
                    db.commit()
                    skip_markup = types.InlineKeyboardMarkup()
                    skip = types.InlineKeyboardButton('Далее \u23e9', callback_data='skip')
                    skip_markup.add(skip)
                    bot.send_message(message.chat.id, 'Укажите минуты (Нажмите далее, чтобы пропустить)', reply_markup=skip_markup)
        elif step_fetch[0] == 'minutes':
            if not message.text.isdigit():
                bot.send_message(message.chat.id, 'Вы неправильно ввели минуты, повторите еще раз (Вводить только цифрами от 0 до 59)')
                return
            else:
                if int(message.text) < 0 or int(message.text) > 59:
                    bot.send_message(message.chat.id, 'Вы неправильно ввели минуты, повторите еще раз (Вводить только цифрами от 0 до 59)')
                else:
                    sql = 'SELECT * FROM orders WHERE process = %s'
                    val = ('finished',)
                    cursor.execute(sql, val)
                    num = cursor.rowcount
                    sql = 'UPDATE orders SET № = %s WHERE user_id = %s AND process = %s'
                    val = (num + 1, message.chat.id, 'ordering')
                    cursor.execute(sql, val)
                    db.commit()
                    add_min = 'UPDATE users SET min = %s WHERE user_id = %s'
                    minval = (int(message.text), message.chat.id)
                    cursor.execute(add_min, minval)
                    db.commit()
                    conf = 'UPDATE users SET step = %s WHERE user_id = %s'
                    conf_val = ('confirm', message.chat.id)
                    db.commit()
                    sql2 = 'SELECT product, amount, price FROM cart WHERE cart_id = %s'
                    val2 = (message.chat.id,)
                    cursor.execute(sql2, val2)
                    fetch = cursor.fetchall()
                    sel = 'SELECT message FROM calldata WHERE user_id = %s'
                    selval = (message.chat.id,)
                    cursor.execute(sel, selval)
                    message_fetch = cursor.fetchone()
                    sql_add = 'SELECT address FROM users WHERE user_id = %s'
                    sql_val = (message.chat.id,)
                    cursor.execute(sql_add, sql_val)
                    add_fetch = cursor.fetchone()
                    phone_check = 'SELECT phone_number FROM users WHERE user_id = %s'
                    phone_check2 = (message.chat.id,)
                    cursor.execute(phone_check, phone_check2)
                    phone_number = cursor.fetchone()
                    select = 'SELECT price FROM cart WHERE cart_id = %s'
                    select2 = (message.chat.id,)
                    cursor.execute(select, select2)
                    fetch_price = cursor.fetchall()
                    sql = 'SELECT hour, min FROM users WHERE user_id = %s'
                    val = (message.chat.id,)
                    cursor.execute(sql, val)
                    time_fetch = cursor.fetchone()
                    sql = 'UPDATE orders SET time = %s WHERE user_id = %s AND process = %s'
                    val = (str(time_fetch[0]) + ':' + str(time_fetch[1]), message.chat.id, 'ordering')
                    cursor.execute(sql, val)
                    db.commit()
                    if phone_number[0] == None:
                        phone_step = 'UPDATE users SET step = %s WHERE user_id = %s'
                        phone_vals = ('noob', message.chat.id)
                        cursor.execute(phone_step, phone_vals)
                        db.commit()
                        contact = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1, one_time_keyboard=True)
                        contact_button = types.KeyboardButton(text='Отправить контакты', request_contact=True)
                        alltimebutton = types.KeyboardButton('Отмена\u274c')
                        contact.add(contact_button, alltimebutton)
                        bot.send_message(
                            text='Чтобы продолжить вы должны отправить нам ваши контактные данные для дальнейших услуг',
                            chat_id=message.chat.id, reply_markup=contact)
                    else:
                        price_list = []
                        orders = ''
                        for i in range(len(fetch_price)):
                            sc = fetch_price[i]
                            price_list.append(sc[0])
                            orders += '\n' + str(i + 1) + '. ' + (fetch[i])[0] + ' ' + str((fetch[i])[1]) + ' шт. - ' + str(
                                (fetch[i])[2]) + 'тг.'
                        bot.delete_message(message.chat.id, message_fetch[0])
                        confirmation = types.InlineKeyboardMarkup()
                        confirm = types.InlineKeyboardButton('Подтвердить\u2705', callback_data='confirm')
                        uhhhnogetback = types.InlineKeyboardButton('Отмена\u274c', callback_data='bacc')
                        confirmation.add(confirm, uhhhnogetback)
                        remove = types.ReplyKeyboardRemove()
                        bot.send_message(message.chat.id, 'Все верно?', reply_markup=remove)
                        if 0 <= time_fetch[1] <= 9:
                            bot.send_message(parse_mode='Markdown',
                                             text='*Заказ №' + str(num+1) + '*\n' + '\nСпособ доставки - самовывоз' + '\nУказанное время - ' + str(
                                                 time_fetch[0]) + ':' + '0' + str(time_fetch[
                                                                                      1]) + '\nИмя - ' + message.chat.first_name + '\nНомер телефона - ' + str(
                                                 phone_number[
                                                     0]) + '\n*Ваш заказ:*' + orders + '\n*Общая стоимость - *' + '*' + str(
                                                 sum(price_list)) + 'тг.*', chat_id=message.chat.id,
                                             reply_markup=confirmation)
                        else:
                            bot.send_message(parse_mode='Markdown',
                                             text='*Заказ №' + str(num+1) + '*\n' + '\nСпособ доставки - самовывоз' + '\nУказанное время - ' + str(
                                                 time_fetch[0]) + ':' + str(time_fetch[
                                                                                1]) + '\nИмя - ' + message.chat.first_name + '\nНомер телефона - ' + str(
                                                 phone_number[
                                                     0]) + '\n*Ваш заказ:*' + orders + '\n*Общая стоимость - *' + '*' + str(
                                                 sum(price_list)) + 'тг.*', chat_id=message.chat.id,
                                             reply_markup=confirmation)
        elif step_fetch[0] == 'naming':
            try:
                bot.delete_message(message.chat.id, message.message_id - 1)
                sql = 'INSERT INTO productlist (product, seen, process, user_id, disc) VALUES (%s, %s, %s, %s, %s)'
                val = (message.text, '<b>НЕТ</b>', 'making', message.chat.id, '<b>НЕТ</b>')
                cursor.execute(sql, val)
                db.commit()
                sql = 'SELECT desc_, price, type, link FROM productlist WHERE user_id = %s AND process = %s AND product = %s'
                val = (message.chat.id, 'making', message.text)
                cursor.execute(sql, val)
                product = cursor.fetchone()
                if product[0] == None:
                    sql = 'UPDATE productlist SET desc_ = %s WHERE user_id = %s AND product = %s'
                    val = ('не указано', message.chat.id, message.text)
                    cursor.execute(sql, val)
                    db.commit()
                if product[1] == None:
                    sql = 'UPDATE productlist SET price = %s WHERE user_id = %s AND product = %s'
                    val = ('не указано', message.chat.id, message.text)
                    cursor.execute(sql, val)
                    db.commit()
                if product[2] == None:
                    sql = 'UPDATE productlist SET type = %s WHERE user_id = %s AND product = %s'
                    val = ('не указано', message.chat.id, message.text)
                    cursor.execute(sql, val)
                    db.commit()
                if product[3] == None:
                    sql = 'UPDATE productlist SET link = %s WHERE user_id = %s AND product = %s'
                    val = ('не указано', message.chat.id, message.text)
                    cursor.execute(sql, val)
                    db.commit()
                sql = 'SELECT desc_, price, type, seen, link FROM productlist WHERE user_id = %s AND process = %s AND product = %s'
                val = (message.chat.id, 'making', message.text)
                cursor.execute(sql, val)
                product = cursor.fetchone()
                adding_prod = types.InlineKeyboardMarkup(2)
                preview = types.InlineKeyboardButton('Предпросмотр', callback_data='preview_prod')
                edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='edit_name')
                edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='edit_photo')
                edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='edit_desc')
                edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='edit_price')
                edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='edit_cate')
                show = types.InlineKeyboardButton('Показать\u2705', callback_data='show' + message.text)
                ready = types.InlineKeyboardButton('Сохранить\ud83d\udcbe', callback_data='ready_prod')
                delete = types.InlineKeyboardButton('Удалить продукт\u274c', callback_data='delete_prod')
                disc = types.InlineKeyboardButton('Акция \u25fb\ufe0f', callback_data='disc')
                adding_prod.add(ready, delete, edit_name, edit_photo, edit_desc, edit_cate, edit_price, show, disc)
                adding_prod.add(preview)
                bot.send_message(message.chat.id, 'Добавлен новый продукт: ' + message.text + '!\n<b>Информация о продукте</b>\n\nИмя: ' + message.text + '\nОписание: \n' + product[0] + '\nЦена: ' + str(product[1]) + '\nКатегория: ' + product[2] + '\nВидимость: ' + product[3] + '\nФото: ' + product[4], parse_mode='HTML', reply_markup=adding_prod)
            except Exception:
                markup_cancel = types.InlineKeyboardMarkup()
                cancel = types.InlineKeyboardButton('Отмена\u274c', callback_data='cancel_name')
                markup_cancel.add(cancel)
                bot.send_message(message.chat.id, 'Слишком длинное название, повторите еще раз.', reply_markup=markup_cancel)
        elif step_fetch[0] == 'edit_name':
            try:
                bot.delete_message(message.chat.id, message.message_id - 1)
                sql = 'UPDATE productlist SET product = %s WHERE user_id = %s AND process = %s'
                val = (message.text, message.chat.id, 'making')
                cursor.execute(sql, val)
                db.commit()
                sql = 'SELECT product, desc_, price, type, seen, link, disc FROM productlist WHERE user_id = %s AND process = %s'
                val = (message.chat.id, 'making')
                cursor.execute(sql, val)
                product = cursor.fetchone()
                adding_prod = types.InlineKeyboardMarkup(2)
                preview = types.InlineKeyboardButton('Предпросмотр', callback_data='preview_prod')
                edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='edit_name')
                edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='edit_photo')
                edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='edit_desc')
                edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='edit_price')
                edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='edit_cate')
                show = types.InlineKeyboardButton('Показать\u2705', callback_data='show' + product[0])
                hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='hide' + product[0])
                ready = types.InlineKeyboardButton('Сохранить\ud83d\udcbe', callback_data='ready_prod')
                delete = types.InlineKeyboardButton('Удалить продукт\u274c', callback_data='delete_prod')
                disc_no = types.InlineKeyboardButton('Акция \u2611\ufe0f', callback_data='disc_no')
                disc = types.InlineKeyboardButton('Акция \u25fb\ufe0f', callback_data='disc')
                adding_prod.add(ready, delete)
                if product[4] == '<b>ДА</b>':
                    adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, hide)
                else:
                    adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show)
                if product[6] == '<b>ДА</b>':
                    adding_prod.add(disc_no)
                else:
                    adding_prod.add(disc)
                adding_prod.add(preview)
                bot.send_message(message.chat.id,
                                 'Имя продукта отредактировано!\n<b>Информация о продукте</b>\n\nИмя: ' + product[0] + '\nОписание: \n' + product[
                                     1] + '\nЦена: ' + str(product[2]) + '\nКатегория: ' + product[3] + '\nВидимость: ' + product[4] + '\nФото: ' + product[5],
                                 parse_mode='HTML', reply_markup=adding_prod)
            except:
                markup_cancel = types.InlineKeyboardMarkup()
                cancel = types.InlineKeyboardButton('Отмена\u274c', callback_data='cancel_all')
                markup_cancel.add(cancel)
                bot.send_message(message.chat.id, 'Слишком длинное название, повторите еще раз.',
                                 reply_markup=markup_cancel)
        elif step_fetch[0] == 'edit_desc':
            bot.delete_message(message.chat.id, message.message_id - 1)
            sql = 'UPDATE productlist SET desc_ = %s WHERE user_id = %s AND process = %s'
            val = (message.text, message.chat.id, 'making')
            cursor.execute(sql, val)
            db.commit()
            sql = 'SELECT product, desc_, price, type, seen, link, disc FROM productlist WHERE user_id = %s AND process = %s'
            val = (message.chat.id, 'making')
            cursor.execute(sql, val)
            product = cursor.fetchone()
            adding_prod = types.InlineKeyboardMarkup(2)
            preview = types.InlineKeyboardButton('Предпросмотр', callback_data='preview_prod')
            edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='edit_name')
            edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='edit_photo')
            edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='edit_desc')
            edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='edit_price')
            edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='edit_cate')
            show = types.InlineKeyboardButton('Показать\u2705', callback_data='show' + product[0])
            hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='hide' + product[0])
            ready = types.InlineKeyboardButton('Сохранить\ud83d\udcbe', callback_data='ready_prod')
            delete = types.InlineKeyboardButton('Удалить продукт\u274c', callback_data='delete_prod')
            disc_no = types.InlineKeyboardButton('Акция \u2611\ufe0f', callback_data='disc_no')
            disc = types.InlineKeyboardButton('Акция \u25fb\ufe0f', callback_data='disc')
            adding_prod.add(ready, delete)
            if product[4] == '<b>ДА</b>':
                adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, hide)
            else:
                adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show)
            if product[6] == '<b>ДА</b>':
                adding_prod.add(disc_no)
            else:
                adding_prod.add(disc)
            adding_prod.add(preview)
            bot.send_message(message.chat.id,
                             'Описание продукта отредактировано\n<b>Информация о продукте</b>\n\nИмя: ' + product[0] + '\nОписание: \n' + product[
                                 1] + '\nЦена: ' + str(product[2]) + '\nКатегория: ' + product[3] + '\nВидимость: ' +
                             product[4] + '\nФото: ' + product[5],
                             parse_mode='HTML', reply_markup=adding_prod)
        elif step_fetch[0] == 'edit_price':
            if message.text is not None:
                if message.text.isdigit():
                    bot.delete_message(message.chat.id, message.message_id - 1)
                    sql = 'UPDATE productlist SET price = %s WHERE user_id = %s AND process = %s'
                    val = (message.text, message.chat.id, 'making')
                    cursor.execute(sql, val)
                    db.commit()
                    sql = 'SELECT product, desc_, price, type, seen, link, disc FROM productlist WHERE user_id = %s AND process = %s'
                    val = (message.chat.id, 'making')
                    cursor.execute(sql, val)
                    product = cursor.fetchone()
                    adding_prod = types.InlineKeyboardMarkup(2)
                    preview = types.InlineKeyboardButton('Предпросмотр', callback_data='preview_prod')
                    edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='edit_name')
                    edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='edit_photo')
                    edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='edit_desc')
                    edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='edit_price')
                    edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='edit_cate')
                    show = types.InlineKeyboardButton('Показать\u2705', callback_data='show' + product[0])
                    hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='hide' + product[0])
                    ready = types.InlineKeyboardButton('Сохранить\ud83d\udcbe', callback_data='ready_prod')
                    delete = types.InlineKeyboardButton('Удалить продукт\u274c', callback_data='delete_prod')
                    disc_no = types.InlineKeyboardButton('Акция \u2611\ufe0f', callback_data='disc_no')
                    disc = types.InlineKeyboardButton('Акция \u25fb\ufe0f', callback_data='disc')
                    adding_prod.add(ready, delete)
                    if product[4] == '<b>ДА</b>':
                        adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, hide)
                    else:
                        adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show)
                    if product[6] == '<b>ДА</b>':
                        adding_prod.add(disc_no)
                    else:
                        adding_prod.add(disc)
                    adding_prod.add(preview)
                    bot.send_message(message.chat.id,
                                     'Цена продукта отредактирована!\n<b>Информация о продукте</b>\n\nИмя: ' + product[0] + '\nОписание: \n' + product[
                                         1] + '\nЦена: ' + str(product[2]) + '\nКатегория: ' + product[
                                         3] + '\nВидимость: ' + product[4] + '\nФото: ' + product[5],
                                     parse_mode='HTML', reply_markup=adding_prod)
                else:
                    markup_cancel = types.InlineKeyboardMarkup()
                    cancel = types.InlineKeyboardButton('Отмена\u274c', callback_data='cancel_all')
                    markup_cancel.add(cancel)
                    bot.send_message(message.chat.id, 'Цена указана неправильно! Вводить только цифрами.', reply_markup=markup_cancel)
        elif step_fetch[0] == 'edit_photo':
            bot.delete_message(message.chat.id, message.message_id - 1)
            sql = 'UPDATE productlist SET link = %s WHERE user_id = %s AND process = %s'
            val = (message.text, message.chat.id, 'making')
            cursor.execute(sql, val)
            db.commit()
            sql = 'SELECT product, desc_, price, type, seen, link, disc FROM productlist WHERE user_id = %s AND process = %s'
            val = (message.chat.id, 'making')
            cursor.execute(sql, val)
            product = cursor.fetchone()
            adding_prod = types.InlineKeyboardMarkup(2)
            preview = types.InlineKeyboardButton('Предпросмотр', callback_data='preview_prod')
            edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='edit_name')
            edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='edit_photo')
            edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='edit_desc')
            edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='edit_price')
            edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='edit_cate')
            show = types.InlineKeyboardButton('Показать\u2705', callback_data='show' + product[0])
            hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='hide' + product[0])
            ready = types.InlineKeyboardButton('Сохранить\ud83d\udcbe', callback_data='ready_prod')
            delete = types.InlineKeyboardButton('Удалить продукт\u274c', callback_data='delete_prod')
            disc_no = types.InlineKeyboardButton('Акция \u2611\ufe0f', callback_data='disc_no')
            disc = types.InlineKeyboardButton('Акция \u25fb\ufe0f', callback_data='disc')
            adding_prod.add(ready, delete)
            if product[4] == '<b>ДА</b>':
                adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, hide)
            else:
                adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show)
            if product[6] == '<b>ДА</b>':
                adding_prod.add(disc_no)
            else:
                adding_prod.add(disc)
            adding_prod.add(preview)
            bot.send_message(message.chat.id,
                                 'Фото продукта отредактировано\n<b>Информация о продукте</b>\n\nИмя: ' + product[0] + '\nОписание: \n' + product[
                                     1] + '\nЦена: ' + str(product[2]) + '\nКатегория: ' + product[
                                     3] + '\nВидимость: ' + product[4] + '\nФото: ' + product[5],
                                 parse_mode='HTML', reply_markup=adding_prod)
        elif step_fetch[0] == 'eedit_name':
            try:
                sql = 'UPDATE users SET step = %s WHERE user_id = %s'
                val = (None, call.message.chat.id)
                cursor.execute(sql, val)
                db.commit()
                sql = 'UPDATE calldata SET start_msg = %s WHERE user_id = %s'
                val = (message.message_id + 1, message.chat.id)
                cursor.execute(sql, val)
                db.commit()
                bot.delete_message(message.chat.id, message.message_id - 1)
                sql = 'UPDATE productlist SET product = %s WHERE user_id = %s AND process = %s'
                val = (message.text, message.chat.id, 'editing')
                cursor.execute(sql, val)
                db.commit()
                sql = 'SELECT product, price, desc_, type, seen, link FROM productlist WHERE user_id = %s AND process = %s'
                val = (message.chat.id, 'editing')
                cursor.execute(sql, val)
                productt = cursor.fetchone()
                sql = 'SELECT product, price, desc_, link, seen, type FROM productlist WHERE type = %s AND process NOT IN (%s)'
                val = (productt[3], 'making')
                cursor.execute(sql, val)
                product = cursor.fetchall()
                message_id = message.message_id + 2
                bot.send_message(text='Категория - ' + productt[3], chat_id=message.chat.id)
                for i in range(len(product)):
                    food_type = types.InlineKeyboardMarkup(2)
                    button = types.InlineKeyboardButton('Редактировать\ud83d\udd27', callback_data='edit_prod' + str(i))
                    food_type.add(button)
                    sql = 'INSERT INTO messages (user_id, message) VALUES (%s, %s)'
                    val = (message.chat.id, message_id + i)
                    cursor.execute(sql, val)
                    db.commit()
                    if product[i][0] == productt[0]:
                        sql = 'UPDATE calldata SET message = %s WHERE user_id = %s'
                        val = (message_id + i, message.chat.id)
                        cursor.execute(sql, val)
                        adding_prod = types.InlineKeyboardMarkup(2)
                        preview = types.InlineKeyboardButton('Предпросмотр', callback_data='epreview_prod' + product[i][0])
                        edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='eedit_name' + product[i][0])
                        edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='eedit_photo' + product[i][0])
                        edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='eedit_desc' + product[i][0])
                        edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='eedit_price' + product[i][0])
                        edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='eedit_cate' + product[i][0])
                        show = types.InlineKeyboardButton('Показать\u2705', callback_data='eshow' + product[i][0])
                        hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='ehide' + product[i][0])
                        delete = types.InlineKeyboardButton('Удалить продукт\u274c',
                                                            callback_data='edelete_prod' + product[i][0])
                        adding_prod.add(delete, preview)
                        if product[i][4] == '<b>ДА</b>':
                            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, hide)
                        else:
                            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show)
                        if i == len(product) - 1:
                            cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
                            adding_prod.row(cancel)
                        bot.send_message(chat_id=message.chat.id,
                                         text='Имя продукта отредактировано!\n<b>Информация о продукте</b>\n\nИмя: ' + product[
                                             i][0] + '\nОписание: \n' + product[
                                                  i][2] + '\nЦена: ' + str(product[i][1]) + '\nКатегория: ' + product[
                                                  i][5] + '\nВидимость: ' +
                                              product[i][4] + '\nФото: ' + product[i][3], parse_mode='HTML',
                                         reply_markup=adding_prod)
                        sql = 'UPDATE calldata SET message = %s WHERE user_id = %s'
                        val = (message_id + i, message.chat.id)
                        cursor.execute(sql, val)
                        db.commit()
                        continue
                    if i == len(product) - 1:
                        cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
                        food_type.add(cancel)
                    if product[i][3] != 'не указано':
                        bot.send_message(
                            text='<a href="' + product[i][3] + '"> </a>' + product[i][0] + ' - ' + str(
                                product[i][1]) + 'тг.\n' + product[i][2] + '\nВидно: ' + product[i][4],
                            chat_id=message.chat.id, parse_mode='HTML', reply_markup=food_type)
                    else:
                        bot.send_message(
                            text=product[i][0] + ' - ' + str(product[i][1]) + 'тг.\n' + product[i][
                                2] + '\nОТСУТСТВУЕТ ФОТОГРАФИЯ\nВидно: ' + product[i][4],
                            chat_id=message.chat.id, parse_mode='HTML', reply_markup=food_type)
            except:
                sql = 'SELECT product, price, desc_, type, seen, link FROM productlist WHERE user_id = %s AND process = %s'
                val = (message.chat.id, 'editing')
                cursor.execute(sql, val)
                productt = cursor.fetchone()
                food_type = types.InlineKeyboardMarkup(2)
                cancel = types.InlineKeyboardButton('Отмена\u274c', callback_data='stop_edit' + product[0])
                food_type.add(cancel)
                bot.send_message(text='Введите новое имя к продукту ' + product[0] + ':', chat_id=call.message.chat.id,
                                 reply_markup=food_type)
        elif step_fetch[0] == 'eedit_photo':
            sql = 'UPDATE users SET step = %s WHERE user_id = %s'
            val = (None, call.message.chat.id)
            cursor.execute(sql, val)
            db.commit()
            sql = 'UPDATE calldata SET start_msg = %s WHERE user_id = %s'
            val = (message.message_id + 1, message.chat.id)
            cursor.execute(sql, val)
            db.commit()
            bot.delete_message(message.chat.id, message.message_id - 1)
            sql = 'UPDATE productlist SET link = %s WHERE user_id = %s AND process = %s'
            val = (message.text, message.chat.id, 'editing')
            cursor.execute(sql, val)
            db.commit()
            sql = 'SELECT product, price, desc_, type, seen, link FROM productlist WHERE user_id = %s AND process = %s'
            val = (message.chat.id, 'editing')
            cursor.execute(sql, val)
            productt = cursor.fetchone()
            sql = 'SELECT product, price, desc_, link, seen, type FROM productlist WHERE type = %s AND process NOT IN (%s)'
            val = (productt[3], 'making')
            cursor.execute(sql, val)
            product = cursor.fetchall()
            message_id = message.message_id + 2
            bot.send_message(text='Категория - ' + productt[3], chat_id=message.chat.id)
            for i in range(len(product)):
                food_type = types.InlineKeyboardMarkup(2)
                button = types.InlineKeyboardButton('Редактировать\ud83d\udd27', callback_data='edit_prod' + str(i))
                food_type.add(button)
                sql = 'INSERT INTO messages (user_id, message) VALUES (%s, %s)'
                val = (message.chat.id, message_id + i)
                cursor.execute(sql, val)
                db.commit()
                if product[i][0] == productt[0]:
                    sql = 'UPDATE calldata SET message = %s WHERE user_id = %s'
                    val = (message_id + i, message.chat.id)
                    cursor.execute(sql, val)
                    adding_prod = types.InlineKeyboardMarkup(2)
                    preview = types.InlineKeyboardButton('Предпросмотр', callback_data='epreview_prod' + product[i][0])
                    edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='eedit_name' + product[i][0])
                    edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='eedit_photo' + product[i][0])
                    edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='eedit_desc' + product[i][0])
                    edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='eedit_price' + product[i][0])
                    edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='eedit_cate' + product[i][0])
                    show = types.InlineKeyboardButton('Показать\u2705', callback_data='eshow' + product[i][0])
                    hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='ehide' + product[i][0])
                    delete = types.InlineKeyboardButton('Удалить продукт\u274c',
                                                        callback_data='edelete_prod' + product[i][0])
                    adding_prod.add(delete, preview)
                    if product[i][4] == '<b>ДА</b>':
                        adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, hide)
                    else:
                        adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show)
                    if i == len(product) - 1:
                        cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
                        adding_prod.row(cancel)
                    bot.send_message(chat_id=message.chat.id,
                                     text='Фото продукта отредактировано!\n<b>Информация о продукте</b>\n\nИмя: ' +
                                          product[
                                              i][0] + '\nОписание: \n' + product[
                                              i][2] + '\nЦена: ' + str(product[i][1]) + '\nКатегория: ' + product[
                                              i][5] + '\nВидимость: ' +
                                          product[i][4] + '\nФото: ' + product[i][3], parse_mode='HTML',
                                     reply_markup=adding_prod)
                    sql = 'UPDATE calldata SET message = %s WHERE user_id = %s'
                    val = (message_id + i, message.chat.id)
                    cursor.execute(sql, val)
                    db.commit()
                    continue
                if i == len(product) - 1:
                    cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
                    food_type.add(cancel)
                if product[i][3] != 'не указано':
                    bot.send_message(
                        text='<a href="' + product[i][3] + '"> </a>' + product[i][0] + ' - ' + str(
                            product[i][1]) + 'тг.\n' + product[i][2] + '\nВидно: ' + product[i][4],
                        chat_id=message.chat.id, parse_mode='HTML', reply_markup=food_type)
                else:
                    bot.send_message(
                        text=product[i][0] + ' - ' + str(product[i][1]) + 'тг.\n' + product[i][
                            2] + '\nОТСУТСТВУЕТ ФОТОГРАФИЯ\nВидно: ' + product[i][4],
                        chat_id=message.chat.id, parse_mode='HTML', reply_markup=food_type)
        elif step_fetch[0] == 'eedit_desc':
            sql = 'UPDATE users SET step = %s WHERE user_id = %s'
            val = (None, call.message.chat.id)
            cursor.execute(sql, val)
            db.commit()
            sql = 'UPDATE calldata SET start_msg = %s WHERE user_id = %s'
            val = (message.message_id + 1, message.chat.id)
            cursor.execute(sql, val)
            db.commit()
            bot.delete_message(message.chat.id, message.message_id - 1)
            sql = 'UPDATE productlist SET desc_ = %s WHERE user_id = %s AND process = %s'
            val = (message.text, message.chat.id, 'editing')
            cursor.execute(sql, val)
            db.commit()
            sql = 'SELECT product, price, desc_, type, seen, link FROM productlist WHERE user_id = %s AND process = %s'
            val = (message.chat.id, 'editing')
            cursor.execute(sql, val)
            productt = cursor.fetchone()
            sql = 'SELECT product, price, desc_, link, seen, type FROM productlist WHERE type = %s AND process NOT IN (%s)'
            val = (productt[3], 'making')
            cursor.execute(sql, val)
            product = cursor.fetchall()
            message_id = message.message_id + 2
            bot.send_message(text='Категория - ' + productt[3], chat_id=message.chat.id)
            for i in range(len(product)):
                food_type = types.InlineKeyboardMarkup(2)
                button = types.InlineKeyboardButton('Редактировать\ud83d\udd27', callback_data='edit_prod' + str(i))
                food_type.add(button)
                sql = 'INSERT INTO messages (user_id, message) VALUES (%s, %s)'
                val = (message.chat.id, message_id + i)
                cursor.execute(sql, val)
                db.commit()
                if product[i][0] == productt[0]:
                    sql = 'UPDATE calldata SET message = %s WHERE user_id = %s'
                    val = (message_id + i, message.chat.id)
                    cursor.execute(sql, val)
                    adding_prod = types.InlineKeyboardMarkup(2)
                    preview = types.InlineKeyboardButton('Предпросмотр', callback_data='epreview_prod' + product[i][0])
                    edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='eedit_name' + product[i][0])
                    edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='eedit_photo' + product[i][0])
                    edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='eedit_desc' + product[i][0])
                    edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='eedit_price' + product[i][0])
                    edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='eedit_cate' + product[i][0])
                    show = types.InlineKeyboardButton('Показать\u2705', callback_data='eshow' + product[i][0])
                    hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='ehide' + product[i][0])
                    delete = types.InlineKeyboardButton('Удалить продукт\u274c',
                                                        callback_data='edelete_prod' + product[i][0])
                    adding_prod.add(delete, preview)
                    if product[i][4] == '<b>ДА</b>':
                        adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, hide)
                    else:
                        adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show)
                    if i == len(product) - 1:
                        cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
                        adding_prod.row(cancel)
                    bot.send_message(chat_id=message.chat.id,
                                     text='Описание продукта отредактировано!\n<b>Информация о продукте</b>\n\nИмя: ' +
                                          product[
                                              i][0] + '\nОписание: \n' + product[
                                              i][2] + '\nЦена: ' + str(product[i][1]) + '\nКатегория: ' + product[
                                              i][5] + '\nВидимость: ' +
                                          product[i][4] + '\nФото: ' + product[i][3], parse_mode='HTML',
                                     reply_markup=adding_prod)
                    sql = 'UPDATE calldata SET message = %s WHERE user_id = %s'
                    val = (message_id + i, message.chat.id)
                    cursor.execute(sql, val)
                    db.commit()
                    continue
                if i == len(product) - 1:
                    cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
                    food_type.add(cancel)
                if product[i][3] != 'не указано':
                    bot.send_message(
                        text='<a href="' + product[i][3] + '"> </a>' + product[i][0] + ' - ' + str(
                            product[i][1]) + 'тг.\n' + product[i][2] + '\nВидно: ' + product[i][4],
                        chat_id=message.chat.id, parse_mode='HTML', reply_markup=food_type)
                else:
                    bot.send_message(
                        text=product[i][0] + ' - ' + str(product[i][1]) + 'тг.\n' + product[i][
                            2] + '\nОТСУТСТВУЕТ ФОТОГРАФИЯ\nВидно: ' + product[i][4],
                        chat_id=message.chat.id, parse_mode='HTML', reply_markup=food_type)
        elif step_fetch[0] == 'eedit_price':
            if message.text.isdigit():
                try:
                    sql = 'UPDATE users SET step = %s WHERE user_id = %s'
                    val = (None, call.message.chat.id)
                    cursor.execute(sql, val)
                    db.commit()
                    sql = 'UPDATE calldata SET start_msg = %s WHERE user_id = %s'
                    val = (message.message_id + 1, message.chat.id)
                    cursor.execute(sql, val)
                    db.commit()
                    bot.delete_message(message.chat.id, message.message_id - 1)
                    sql = 'UPDATE productlist SET price = %s WHERE user_id = %s AND process = %s'
                    val = (message.text, message.chat.id, 'editing')
                    cursor.execute(sql, val)
                    db.commit()
                    sql = 'SELECT product, price, desc_, type, seen, link FROM productlist WHERE user_id = %s AND process = %s'
                    val = (message.chat.id, 'editing')
                    cursor.execute(sql, val)
                    productt = cursor.fetchone()
                    sql = 'SELECT product, price, desc_, link, seen, type FROM productlist WHERE type = %s AND process NOT IN (%s)'
                    val = (productt[3], 'making')
                    cursor.execute(sql, val)
                    product = cursor.fetchall()
                    message_id = message.message_id + 2
                    bot.send_message(text='Категория - ' + productt[3], chat_id=message.chat.id)
                    for i in range(len(product)):
                        food_type = types.InlineKeyboardMarkup(2)
                        button = types.InlineKeyboardButton('Редактировать\ud83d\udd27', callback_data='edit_prod' + str(i))
                        food_type.add(button)
                        sql = 'INSERT INTO messages (user_id, message) VALUES (%s, %s)'
                        val = (message.chat.id, message_id + i)
                        cursor.execute(sql, val)
                        db.commit()
                        if product[i][0] == productt[0]:
                            sql = 'UPDATE calldata SET message = %s WHERE user_id = %s'
                            val = (message_id + i, message.chat.id)
                            cursor.execute(sql, val)
                            adding_prod = types.InlineKeyboardMarkup(2)
                            preview = types.InlineKeyboardButton('Предпросмотр',
                                                                 callback_data='epreview_prod' + product[i][0])
                            edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='eedit_name' + product[i][0])
                            edit_photo = types.InlineKeyboardButton('Ред. фото',
                                                                    callback_data='eedit_photo' + product[i][0])
                            edit_desc = types.InlineKeyboardButton('Ред. описание',
                                                                   callback_data='eedit_desc' + product[i][0])
                            edit_price = types.InlineKeyboardButton('Ред. цену',
                                                                    callback_data='eedit_price' + product[i][0])
                            edit_cate = types.InlineKeyboardButton('Ред. категорию',
                                                                   callback_data='eedit_cate' + product[i][0])
                            show = types.InlineKeyboardButton('Показать\u2705', callback_data='eshow' + product[i][0])
                            hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='ehide' + product[i][0])
                            delete = types.InlineKeyboardButton('Удалить продукт\u274c',
                                                                callback_data='edelete_prod' + product[i][0])
                            adding_prod.add(delete, preview)
                            if product[i][4] == '<b>ДА</b>':
                                adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, hide)
                            else:
                                adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show)
                            if i == len(product) - 1:
                                cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
                                adding_prod.row(cancel)
                            bot.send_message(chat_id=message.chat.id,
                                             text='Описание продукта отредактировано!\n<b>Информация о продукте</b>\n\nИмя: ' +
                                                  product[
                                                      i][0] + '\nОписание: \n' + product[
                                                      i][2] + '\nЦена: ' + str(product[i][1]) + '\nКатегория: ' + product[
                                                      i][5] + '\nВидимость: ' +
                                                  product[i][4] + '\nФото: ' + product[i][3], parse_mode='HTML',
                                             reply_markup=adding_prod)
                            sql = 'UPDATE calldata SET message = %s WHERE user_id = %s'
                            val = (message_id + i, message.chat.id)
                            cursor.execute(sql, val)
                            db.commit()
                            continue
                        if i == len(product) - 1:
                            cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
                            food_type.add(cancel)
                        if product[i][3] != 'не указано':
                            bot.send_message(
                                text='<a href="' + product[i][3] + '"> </a>' + product[i][0] + ' - ' + str(
                                    product[i][1]) + 'тг.\n' + product[i][2] + '\nВидно: ' + product[i][4],
                                chat_id=message.chat.id, parse_mode='HTML', reply_markup=food_type)
                        else:
                            bot.send_message(
                                text=product[i][0] + ' - ' + str(product[i][1]) + 'тг.\n' + product[i][
                                    2] + '\nОТСУТСТВУЕТ ФОТОГРАФИЯ\nВидно: ' + product[i][4],
                                chat_id=message.chat.id, parse_mode='HTML', reply_markup=food_type)
                except:
                    bot.send_message(text='Слишком большое значение, повторите еще раз:', chat_id=message.chat.id,
                                     reply_markup=mk)
                    sql = 'UPDATE calldata SET start_msg = %s WHERE user_id = %s'
                    val = (message.message_id, message.chat.id)
                    cursor.execute(sql, val)
                    db.commit()
            else:
                sql = 'SELECT product, price, desc_, type, seen, link FROM productlist WHERE user_id = %s AND process = %s'
                val = (message.chat.id, 'editing')
                cursor.execute(sql, val)
                product = cursor.fetchone()
                bot.delete_message(message.chat.id, message.message_id - 1)
                cancel = types.InlineKeyboardButton('Отмена\u274c', callback_data='stop_edit' + product[0])
                food_type = types.InlineKeyboardMarkup()
                food_type.add(cancel)
                bot.send_message(text='Неправильно указана цена, вводить только цифрами', chat_id=message.chat.id,
                                 reply_markup=food_type)
                sql = 'UPDATE calldata SET start_msg = %s WHERE user_id = %s'
                val = (message.message_id, message.chat.id)
                cursor.execute(sql, val)
                db.commit()
        elif step_fetch[0] == 'cate_naming':
            try:
                sql = 'SELECT start_msg FROM calldata WHERE user_id = %s'
                val = (message.chat.id,)
                cursor.execute(sql, val)
                msg = cursor.fetchone()
                bot.delete_message(message.chat.id, msg[0])
                sql = 'INSERT INTO categorylist (name, seen) VALUES (%s, %s)'
                val = (message.text, '<b>НЕТ</b>')
                cursor.execute(sql, val)
                db.commit()
                hide_or_not = types.InlineKeyboardMarkup()
                yes = types.InlineKeyboardButton('Да\u2705', callback_data='yes_hide_cate' + message.text)
                no = types.InlineKeyboardButton('Нет\u274c', callback_data='no_donthide_cate' + message.text)
                hide_or_not.add(yes, no)
                bot.send_message(message.chat.id, 'Скрыть категорию ' + message.text + '?', reply_markup=hide_or_not)
            except:
                sql = 'DELETE FROM categorylist WHERE name = %s'
                val = (message.text, )
                cursor.execute(sql, val)
                db.commit()
                mk = types.InlineKeyboardMarkup()
                cancel = types.InlineKeyboardButton('Отмена\u274c', callback_data='cate_list')
                mk.add(cancel)
                bot.send_message(text='Слишком длинное имя, повторите еще раз:', chat_id=message.chat.id,
                                 reply_markup=mk)
                sql = 'UPDATE calldata SET start_msg = %s WHERE user_id = %s'
                val = (message.message_id, message.chat.id)
                cursor.execute(sql, val)
                db.commit()
        if step_fetch[0] is not None:
            try:
                if step_fetch[0].startswith('edit_name_cate'):
                    sql = 'SELECT start_msg FROM calldata WHERE user_id = %s'
                    val = (message.chat.id, )
                    cursor.execute(sql, val)
                    msg = cursor.fetchone()[0]
                    bot.delete_message(message.chat.id, msg)
                    sql = 'UPDATE users SET step = %s WHERE user_id = %s'
                    val = (None, message.chat.id)
                    cursor.execute(sql, val)
                    sql = 'UPDATE productlist SET type = %s WHERE type = %s'
                    val = (message.text, step_fetch[0][14:])
                    cursor.execute(sql, val)
                    sql = 'UPDATE categorylist SET name = %s WHERE name = %s'
                    val = (message.text, step_fetch[0][14:])
                    cursor.execute(sql, val)
                    db.commit()
                    sql = 'SELECT name, seen FROM categorylist'
                    cursor.execute(sql)
                    category = cursor.fetchall()
                    cate_list = types.InlineKeyboardMarkup(4)
                    if cursor.rowcount > 0:
                        for i in range(len(category)):
                            button = types.InlineKeyboardButton(category[i][0], callback_data=str(i) + '//')
                            name = types.InlineKeyboardButton('Ред.Имя', callback_data='go_cate' + category[i][0])
                            show = types.InlineKeyboardButton('Показать\u2705',
                                                              callback_data='gshow_cate' + category[i][0])
                            hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='ghide_cate' + category[i][0])
                            delete = types.InlineKeyboardButton('Удалить\ud83d\uddd1',
                                                                callback_data='gdeletecate' + category[i][0])
                            if category[i][1] == '<b>НЕТ</b>':
                                cate_list.add(button, name, show, delete)
                            else:
                                cate_list.add(button, name, hide, delete)
                        back = types.InlineKeyboardButton('Назад\u2b05', callback_data='cate')
                        cate_list.row(back)
                        bot.send_message(text='<b>Список всех категорий</b>',
                                              chat_id=message.chat.id,
                                              reply_markup=cate_list, parse_mode='HTML')
                    else:
                        back = types.InlineKeyboardButton('Назад\u2b05', callback_data='cate')
                        cate_list.row(back)
                        bot.send_message(
                            text='Категорий не найдено. Попробуйте создать парочку нажав на кнопку Добавить\u2795!',
                            chat_id=message.chat.id, message_id=message.message_id, reply_markup=cate_list, parse_mode='HTML')
            except:
                mk = types.InlineKeyboardMarkup()
                cancel = types.InlineKeyboardButton('Отмена\u274c', callback_data='cate_list')
                mk.add(cancel)
                bot.send_message(text='Слишком длинное имя, повторите еще раз:', chat_id=message.chat.id, reply_markup=mk)
                sql = 'UPDATE calldata SET start_msg = %s WHERE user_id = %s'
                val = (message.message_id, message.chat.id)
                cursor.execute(sql, val)
                db.commit()
    if message.contact:
        sel_del = 'SELECT step FROM users WHERE user_id = %s'
        sel_delval = (message.chat.id,)
        cursor.execute(sel_del, sel_delval)
        step_fetch = cursor.fetchone()
        if step_fetch[0] == 'phone_number':
            sql = 'UPDATE users SET phone_number = %s WHERE user_id = %s'
            val = (message.contact.phone_number, message.chat.id)
            cursor.execute(sql, val)
            db.commit()
            pay_method = types.InlineKeyboardMarkup()
            pay_card = types.InlineKeyboardButton('Перевод картой', callback_data='card')
            pay_money = types.InlineKeyboardButton('Оплата наличными', callback_data='money')
            kaspigold = types.InlineKeyboardButton('Kaspi Gold', callback_data='kaspi')
            pay_method.add(kaspigold, pay_card, pay_money)
            bot.send_message(text='Выберите способ оплаты:', chat_id=message.chat.id,
                                  reply_markup=pay_method)
        elif step_fetch[0] == 'noob':
            sql = 'SELECT * FROM orders WHERE process = %s'
            val = ('finished',)
            cursor.execute(sql, val)
            num = cursor.rowcount
            sql = 'UPDATE orders SET № = %s WHERE user_id = %s AND process = %s'
            val = (num + 1, message.chat.id, 'ordering')
            cursor.execute(sql, val)
            db.commit()
            sql = 'UPDATE users SET phone_number = %s WHERE user_id = %s'
            val = (message.contact.phone_number, message.chat.id)
            cursor.execute(sql, val)
            db.commit()
            conf = 'UPDATE users SET step = %s WHERE user_id = %s'
            conf_val = ('confirm', message.chat.id)
            db.commit()
            sql2 = 'SELECT product, amount, price FROM cart WHERE cart_id = %s'
            val2 = (message.chat.id,)
            cursor.execute(sql2, val2)
            fetch = cursor.fetchall()
            sel = 'SELECT message FROM calldata WHERE user_id = %s'
            selval = (message.chat.id,)
            cursor.execute(sel, selval)
            message_fetch = cursor.fetchone()
            sql_add = 'SELECT address FROM users WHERE user_id = %s'
            sql_val = (message.chat.id,)
            cursor.execute(sql_add, sql_val)
            add_fetch = cursor.fetchone()
            phone_check = 'SELECT phone_number FROM users WHERE user_id = %s'
            phone_check2 = (message.chat.id,)
            cursor.execute(phone_check, phone_check2)
            phone_number = cursor.fetchone()
            select = 'SELECT price FROM cart WHERE cart_id = %s'
            select2 = (message.chat.id,)
            cursor.execute(select, select2)
            fetch_price = cursor.fetchall()
            sql = 'SELECT hour, min FROM users WHERE user_id = %s'
            val = (message.chat.id,)
            cursor.execute(sql, val)
            time_fetch = cursor.fetchone()
            sql = 'UPDATE orders SET time = %s WHERE user_id = %s AND process = %s'
            val = (str(time_fetch[0]) + ':' + str(time_fetch[1]), message.chat.id, 'ordering')
            cursor.execute(sql, val)
            db.commit()
            price_list = []
            orders = ''
            for i in range(len(fetch_price)):
                sc = fetch_price[i]
                price_list.append(sc[0])
                orders += '\n' + str(i + 1) + '. ' + (fetch[i])[0] + ' ' + str((fetch[i])[1]) + ' шт. - ' + str(
                    (fetch[i])[2]) + 'тг.'
            bot.delete_message(message.chat.id, message_fetch[0])
            confirmation = types.InlineKeyboardMarkup()
            confirm = types.InlineKeyboardButton('Подтвердить\u2705', callback_data='confirm')
            uhhhnogetback = types.InlineKeyboardButton('Отмена\u274c', callback_data='bacc')
            confirmation.add(confirm, uhhhnogetback)
            remove = types.ReplyKeyboardRemove()
            bot.send_message(message.chat.id, 'Все верно?', reply_markup=remove)
            if 0 <= time_fetch[1] <= 9:
                bot.send_message(parse_mode='Markdown',
                                 text='Заказ №' + str(num+1) + '*\n' + '\nСпособ доставки - самовывоз' + '\nУказанное время - ' + str(
                                     time_fetch[0]) + ':' + '0' + str(time_fetch[
                                                                          1]) + '\nИмя - ' + message.chat.first_name + '\nНомер телефона - ' + str(
                                     phone_number[0]) + '\n*Ваш заказ:*' + orders + '\n*Общая стоимость - *' + '*' + str(
                                     sum(price_list)) + 'тг.*', chat_id=message.chat.id,
                                 reply_markup=confirmation)
            else:
                bot.send_message(parse_mode='Markdown',
                                 text='Способ доставки - самовывоз' + '\nУказанное время - ' + str(
                                     time_fetch[0]) + ':' + str(time_fetch[
                                                                    1]) + '\nИмя - ' + message.chat.first_name + '\nНомер телефона - ' + str(
                                     phone_number[0]) + '\n*Ваш заказ:*' + orders + '\n*Общая стоимость - *' + '*' + str(
                                     sum(price_list)) + 'тг.*', chat_id=message.chat.id,
                                 reply_markup=confirmation)
        elif step_fetch[0] == 'fast':
            sql = 'SELECT * FROM orders WHERE process = %s'
            val = ('finished',)
            cursor.execute(sql, val)
            num = cursor.rowcount
            sql = 'UPDATE orders SET № = %s WHERE user_id = %s AND process = %s'
            val = (num + 1, message.chat.id, 'ordering')
            cursor.execute(sql, val)
            db.commit()
            sql = 'UPDATE users SET phone_number = %s WHERE user_id = %s'
            val = (message.contact.phone_number, message.chat.id)
            cursor.execute(sql, val)
            db.commit()
            conf = 'UPDATE users SET step = %s WHERE user_id = %s'
            conf_val = ('confirm', message.chat.id)
            db.commit()
            sql2 = 'SELECT product, amount, price FROM cart WHERE cart_id = %s'
            val2 = (message.chat.id,)
            cursor.execute(sql2, val2)
            fetch = cursor.fetchall()
            sel = 'SELECT message FROM calldata WHERE user_id = %s'
            selval = (message.chat.id,)
            cursor.execute(sel, selval)
            message_fetch = cursor.fetchone()
            phone_check = 'SELECT phone_number FROM users WHERE user_id = %s'
            phone_check2 = (message.chat.id,)
            cursor.execute(phone_check, phone_check2)
            phone_number = cursor.fetchone()
            select = 'SELECT price FROM cart WHERE cart_id = %s'
            select2 = (message.chat.id,)
            cursor.execute(select, select2)
            fetch_price = cursor.fetchall()
            sql = 'UPDATE orders SET time = %s WHERE user_id = %s AND process = %s'
            val = ('Как можно скорее', message.chat.id, 'ordering')
            cursor.execute(sql, val)
            db.commit()
            price_list = []
            orders = ''
            for i in range(len(fetch_price)):
                sc = fetch_price[i]
                price_list.append(sc[0])
                orders += '\n' + str(i + 1) + '. ' + (fetch[i])[0] + ' ' + str((fetch[i])[1]) + ' шт. - ' + str(
                    (fetch[i])[2]) + 'тг.'
            bot.delete_message(message.chat.id, message_fetch[0])
            confirmation = types.InlineKeyboardMarkup()
            confirm = types.InlineKeyboardButton('Подтвердить\u2705', callback_data='confirm')
            uhhhnogetback = types.InlineKeyboardButton('Отмена\u274c', callback_data='bacc')
            confirmation.add(confirm, uhhhnogetback)
            remove = types.ReplyKeyboardRemove()
            bot.send_message(message.chat.id, 'Все верно?', reply_markup=remove)
            bot.send_message(parse_mode='Markdown',
                                 text='Заказ №' + str(num+1) + '*\n' + '\nСпособ доставки - самовывоз' + '\nУказанное время - как можно скорее' + '\nИмя - ' + message.chat.first_name + '\nНомер телефона - ' + str(
                                     phone_number[0]) + '\n*Ваш заказ:*' + orders + '\n*Общая стоимость - *' + '*' + str(
                                     sum(price_list)) + 'тг.*', chat_id=message.chat.id,
                                 reply_markup=confirmation)
        elif step_fetch[0] == 'kaspi':
            sql = 'SELECT * FROM orders WHERE process = %s'
            val = ('finished',)
            cursor.execute(sql, val)
            num = cursor.rowcount
            sql = 'UPDATE orders SET № = %s WHERE user_id = %s AND process = %s'
            val = (num + 1, call.message.chat.id, 'ordering')
            cursor.execute(sql, val)
            db.commit()
            contact = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1, one_time_keyboard=True)
            contact_button = types.KeyboardButton(text='Отправить контакты', request_contact=True)
            alltimebutton = types.KeyboardButton('Отмена\u274c')
            contact.add(contact_button, alltimebutton)
            conf = 'UPDATE users SET step = %s WHERE user_id = %s'
            conf_val = ('confirm', call.message.chat.id)
            db.commit()
            sql2 = 'SELECT product, amount, price FROM cart WHERE cart_id = %s'
            val2 = (call.message.chat.id,)
            cursor.execute(sql2, val2)
            fetch = cursor.fetchall()
            select = 'SELECT price FROM cart WHERE cart_id = %s'
            select2 = (call.message.chat.id,)
            cursor.execute(select, select2)
            fetch_price = cursor.fetchall()
            sel = 'SELECT message FROM calldata WHERE user_id = %s'
            selval = (call.message.chat.id,)
            cursor.execute(sel, selval)
            message_fetch = cursor.fetchone()
            phone_check = 'SELECT phone_number FROM users WHERE user_id = %s'
            phone_check2 = (call.message.chat.id,)
            cursor.execute(phone_check, phone_check2)
            phone_number = cursor.fetchone()
            sql = 'UPDATE users SET min = %s WHERE user_id = %s'
            val = (0, call.message.chat.id)
            cursor.execute(sql, val)
            db.commit()
            price_list = []
            orders = ''
            for i in range(len(fetch_price)):
                sc = fetch_price[i]
                price_list.append(sc[0])
                orders += '\n' + str(i + 1) + '. ' + (fetch[i])[0] + ' ' + str((fetch[i])[1]) + ' шт. - ' + str(
                        (fetch[i])[2]) + 'тг.'
            bot.delete_message(call.message.chat.id, message_fetch[0])
            confirmation = types.InlineKeyboardMarkup()
            confirm = types.InlineKeyboardButton('Подтвердить\u2705', callback_data='confirm_k')
            uhhhnogetback = types.InlineKeyboardButton('Отмена\u274c', callback_data='bacc')
            confirmation.add(confirm, uhhhnogetback)
            remove = types.ReplyKeyboardRemove()
            bot.send_message(call.message.chat.id, 'Все верно?', reply_markup=remove)
            bot.send_message(parse_mode='Markdown',
                                 text='*Заказ №' + str(
                                     num + 1) + '*\n' + '\nСпособ доставки - курьер' + '\nСпособ оплаты - каспи голд' + '\nИмя - ' + call.message.chat.first_name + '\nНомер телефона - ' + str(
                                     phone_number[
                                         0]) + '\n*Ваш заказ:*' + orders + '\n*Общая стоимость - *' + '*' + str(
                                     sum(
                                         price_list)) + 'тг.*\nПеревод на каспи голд производится после подтверждения заказа',
                                 chat_id=call.message.chat.id,
                                 reply_markup=confirmation)
    if message.text == 'Изменить адрес':
        sel_del = 'SELECT step FROM users WHERE user_id = %s'
        sel_delval = (message.chat.id,)
        cursor.execute(sel_del, sel_delval)
        step_fetch = cursor.fetchone()
        if step_fetch[0] == 'address':
            remove = types.ReplyKeyboardRemove()
            bot.send_message(text='Введите новый адрес:', chat_id=message.chat.id, reply_markup=remove)
    if message.text == 'Отмена\u274c':
        sel_del = 'SELECT step FROM users WHERE user_id = %s'
        sel_delval = (message.chat.id,)
        cursor.execute(sel_del, sel_delval)
        step_fetch = cursor.fetchone()
        if step_fetch[0] != None:
            sql = 'UPDATE users SET step = %s WHERE user_id = %s'
            val= (None, message.chat.id)
            cursor.execute(sql, val)
            sql = 'DELETE FROM orders WHERE user_id = %s AND process = %s'
            val = (message.chat.id, 'ordering')
            cursor.execute(sql, val)
            db.commit()
            start_buttons = types.InlineKeyboardMarkup()
            menu = types.InlineKeyboardButton(text='Меню \ud83d\udcd6', callback_data='menu')
            basket = types.InlineKeyboardButton(text='Корзина \ud83d\uded2', callback_data='cart')
            start_buttons.add(menu, basket)
            additional_kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
            cart = types.KeyboardButton(text='Корзина \ud83d\uded2')
            bot.send_message(message.chat.id, parse_mode='Markdown',
                             text='[ ](https://imbt.ga/vfcM0yHFjI)' + start_text, reply_markup=start_buttons)
            sql = 'UPDATE users SET step = %s WHERE user_id = %s'
            val = (None, message.chat.id)
            cursor.execute(sql, val)
            db.commit()
    if message.text is not None:
        if message.text.startswith('/'):
            sql = 'SELECT key_admin FROM users WHERE user_id = %s'
            val = (message.chat.id,)
            cursor.execute(sql, val)
            key = cursor.fetchone()
            if message.text == '/getadmin ' + key[0]:
                bot.send_message(message.chat.id,
                                 'Поздравляем, вы теперь админ! Чтобы открыть админ панель, введите /admin')
                sql = 'UPDATE users SET role = %s WHERE user_id = %s'
                val = ('admin', message.chat.id)
                cursor.execute(sql, val)
                db.commit()
            elif message.text == '/admin':
                sql = 'SELECT role FROM users WHERE user_id = %s'
                val = (message.chat.id,)
                cursor.execute(sql, val)
                role = cursor.fetchone()
                if role[0] == 'admin':
                    admin_start = types.InlineKeyboardMarkup(row_width=1)
                    reports = types.InlineKeyboardButton('Отчеты', callback_data='reports')
                    categories = types.InlineKeyboardButton('Меню', callback_data='edit')
                    admin_start.add(reports, categories)
                    bot.send_message(message.chat.id, 'Админ панель', reply_markup=admin_start)
                else:
                    bot.send_message(message.chat.id, 'Вы не являетесь админом!')
    if message.photo:
        sel_del = 'SELECT step FROM users WHERE user_id = %s'
        sel_delval = (message.chat.id,)
        cursor.execute(sel_del, sel_delval)
        step_fetch = cursor.fetchone()
        if step_fetch[0] == 'kaspi_ss':
            bot.forward_message(dad_id, message.chat.id, message.message_id)
            main_page = types.InlineKeyboardMarkup()
            main_pagebutton = types.InlineKeyboardButton('Главная\ud83c\udfe0', callback_data='home')
            main_page.add(main_pagebutton)
            bot.send_message(message.chat.id, 'Ваш заказ успешно оформлен!', reply_markup=main_page)


@bot.callback_query_handler(func=lambda call: True)      #Меню
def menu(call):
    if call.data == 'menu':
        food_type = types.InlineKeyboardMarkup(2)
        sql = 'SELECT name FROM categorylist WHERE seen = %s'
        val = ('<b>ДА</b>', )
        cursor.execute(sql, val)
        category = cursor.fetchall()
        for i in range(len(category)):
            button = types.InlineKeyboardButton(category[i][0], callback_data=str(i))
            food_type.add(button)
        back = types.InlineKeyboardButton('Назад \u2b05', callback_data='back')
        cursor.execute(sql, val)
        food_type.add(back)
        bot.edit_message_text(text='Выберите желаемую категорию, пожалуйста:', chat_id=call.message.chat.id, message_id=call.message.message_id, reply_markup=food_type)
    if call.data == 'sales':
        sql = 'SELECT message FROM messages WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        msg = cursor.fetchall()[0]
        for i in msg:
            bot.delete_message(msg[i])
        sql = 'DELETE FROM messages WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        db.commit()
        sql2 = 'SELECT product, price, link, desc_ FROM productlist WHERE disc = %s AND seen = %s AND process = %s'
        val2 = ('<b>ДА</b>', '<b>ДА</b>', 'saved')
        fetchfood = cursor.fetchall()
        if cursor.rowcount == 0:
            order_markup = types.InlineKeyboardMarkup()
            back_menu = types.InlineKeyboardButton(text='В меню \u2b05', callback_data='backmenu')
            order_markup.add(back_menu)
            bot.edit_message_text(
                'Извините, акций сейчас нет.\nЗагляните сюда в другой раз!',
                call.message.chat.id, call.message.message_id, reply_markup=order_markup)
        else:
            sql = 'UPDATE calldata SET start_msg = %s WHERE user_id = %s'
            val = (call.message.message_id, call.message.chat.id)
            cursor.execute(sql, val)
            sql = 'UPDATE calldata SET message = %s WHERE user_id = %s'
            val = (None, call.message.chat.id)
            cursor.execute(sql, val)
            db.commit()
            sql = 'UPDATE calldata SET calldata = %s WHERE user_id = %s'
            val = (None, call.message.chat.id)
            cursor.execute(sql, val)
            db.commit()
            sql = 'UPDATE calldata SET calldata2 = %s WHERE user_id = %s'
            val = (None, call.message.chat.id)
            cursor.execute(sql, val)
            db.commit()
            sql = 'UPDATE calldata SET calldata2 = %s WHERE user_id = %s'
            val = (ind, call.message.chat.id)
            cursor.execute(sql, val)
            db.commit()
            bot.edit_message_text(text='Раздел - ' + category[ind][0], chat_id=call.message.chat.id,
                                  message_id=call.message.message_id)
            for p in range(len(fetchfood)):
                order_markup = types.InlineKeyboardMarkup(1)
                button = types.InlineKeyboardButton(text='В корзину',
                                                    callback_data=str((p + cate)))
                order_markup.add(button)
                if p == len(fetchfood) - 1:
                    back_menu = types.InlineKeyboardButton(text='В меню \u2b05', callback_data='backmenu')
                    order_markup.add(back_menu)
                bot.send_message(call.message.chat.id,
                                 '[ ](' + fetchfood[p][2] + ')*' + (fetchfood[p])[0] + '* - ' + str(
                                     (fetchfood[p])[1]) + 'тг.\n\n' + fetchfood[p][3], parse_mode='Markdown',
                                 reply_markup=order_markup)
                sql = 'INSERT INTO messages (user_id, message) VALUES (%s, %s)'
                val = (call.message.chat.id, call.message.message_id + 1 + p)
                cursor.execute(sql, val)
                db.commit()
    if call.data.isdigit():
        sql = 'SELECT name FROM categorylist WHERE seen = %s'
        val = ('<b>ДА</b>', )
        cursor.execute(sql, val)
        cate = cursor.rowcount
        category = cursor.fetchall()
        if int(call.data) < cate:
            sql = 'DELETE FROM messages WHERE user_id = %s'
            val = (call.message.chat.id,)
            cursor.execute(sql, val)
            db.commit()
            ind = int(call.data)
            sql = 'SELECT product, price, link, desc_ FROM productlist WHERE type = %s AND seen = %s AND process = %s'
            val = (category[ind][0], '<b>ДА</b>', 'saved')
            cursor.execute(sql, val)
            fetchfood = cursor.fetchall()
            if cursor.rowcount == 0:
                order_markup = types.InlineKeyboardMarkup()
                back_menu = types.InlineKeyboardButton(text='В меню \u2b05', callback_data='backmenu')
                order_markup.add(back_menu)
                bot.edit_message_text('Извините, в категории ' + category[ind][0] + ' сейчас пусто.\nЗагляните сюда в другой раз!', call.message.chat.id, call.message.message_id, reply_markup=order_markup)
            else:
                sql = 'UPDATE calldata SET start_msg = %s WHERE user_id = %s'
                val = (call.message.message_id, call.message.chat.id)
                cursor.execute(sql, val)
                sql = 'UPDATE calldata SET message = %s WHERE user_id = %s'
                val = (None, call.message.chat.id)
                cursor.execute(sql, val)
                db.commit()
                sql = 'UPDATE calldata SET calldata = %s WHERE user_id = %s'
                val = (None, call.message.chat.id)
                cursor.execute(sql, val)
                db.commit()
                sql = 'UPDATE calldata SET calldata2 = %s WHERE user_id = %s'
                val = (None, call.message.chat.id)
                cursor.execute(sql, val)
                db.commit()
                sql = 'UPDATE calldata SET calldata2 = %s WHERE user_id = %s'
                val = (ind, call.message.chat.id)
                cursor.execute(sql, val)
                db.commit()
                cursor.execute('DELETE FROM messages')
                bot.edit_message_text(text='Раздел - ' + category[ind][0], chat_id=call.message.chat.id,
                                            message_id=call.message.message_id)
                for p in range(len(fetchfood)):
                    order_markup = types.InlineKeyboardMarkup(1)
                    button = types.InlineKeyboardButton(text='В корзину',
                                                        callback_data=str((p + cate)))
                    order_markup.add(button)
                    if p == len(fetchfood) - 1:
                        back_menu = types.InlineKeyboardButton(text='В меню \u2b05', callback_data='backmenu')
                        order_markup.add(back_menu)
                    bot.send_message(call.message.chat.id, '[ ](' + fetchfood[p][2] + ')*' + (fetchfood[p])[0] + '* - ' + str((fetchfood[p])[1]) + 'тг.\n\n' + fetchfood[p][3], parse_mode='Markdown', reply_markup=order_markup)
                    sql = 'INSERT INTO messages (user_id, message) VALUES (%s, %s)'
                    val = (call.message.chat.id, call.message.message_id + 1 + p)
                    cursor.execute(sql, val)
                    db.commit()
        if 1000 > int(call.data) >= cate:
            call_check1 = 'SELECT calldata FROM calldata WHERE user_id = %s'
            call_check2 = (call.message.chat.id,)
            cursor.execute(call_check1, call_check2)
            call_fetch = cursor.fetchone()
            calldata = call_fetch[0]
            calldata1 = 'UPDATE calldata SET calldata = %s WHERE user_id = %s'
            calldata2 = (int(call.data) - cate, call.message.chat.id)
            cursor.execute(calldata1, calldata2)
            db.commit()
            call_check1 = 'SELECT calldata FROM calldata WHERE user_id = %s'
            call_check2 = (call.message.chat.id,)
            cursor.execute(call_check1, call_check2)
            call_fetch = cursor.fetchone()
            calldataa = call_fetch[0]
            sql = 'SELECT name FROM categorylist WHERE seen = %s'
            val = ('<b>ДА</b>',)
            cursor.execute(sql, val)
            category = cursor.fetchall()
            sql1 = 'SELECT product FROM cart WHERE cart_id = %s'
            val1 = (call.message.chat.id,)
            cursor.execute(sql1, val1)
            fetchcheck = cursor.fetchall()
            sql = 'SELECT calldata2 FROM calldata WHERE user_id = %s'
            val = (call.message.chat.id, )
            cursor.execute(sql, val)
            category_call = cursor.fetchone()
            sql2 = 'SELECT product, price, link FROM productlist WHERE type = %s'
            val2 = (category[category_call[0]][0],)
            cursor.execute(sql2, val2)
            fetchfood = cursor.fetchall()
            sql = 'SELECT message FROM calldata WHERE user_id = %s'
            val = (call.message.chat.id,)
            cursor.execute(sql, val)
            message = cursor.fetchone()
            sql = 'UPDATE calldata SET message = %s WHERE user_id = %s'
            val = (call.message.message_id, call.message.chat.id)
            cursor.execute(sql, val)
            db.commit()
            if fetchcheck == []:
                sql = 'SELECT message FROM messages WHERE user_id = %s'
                val = (call.message.chat.id,)
                cursor.execute(sql, val)
                messages = cursor.fetchall()
                call_check1 = 'SELECT calldata FROM calldata WHERE user_id = %s'
                call_check2 = (call.message.chat.id, )
                cursor.execute(call_check1, call_check2)
                call_fetch = cursor.fetchone()
                calldataa = call_fetch[0]
                sql2 = 'SELECT product, price, link FROM productlist WHERE type = %s'
                val2 = (category[category_call[0]][0],)
                cursor.execute(sql2, val2)
                fetchfood = cursor.fetchall()
                sql = 'UPDATE users SET value = %s WHERE user_id = %s'
                val = (1, call.message.chat.id)
                cursor.execute(sql, val)
                db.commit()
                sql = 'SELECT value FROM users WHERE user_id = %s'
                val = (call.message.chat.id,)
                cursor.execute(sql, val)
                value = cursor.fetchone()
                count = types.InlineKeyboardMarkup(row_width=4)
                cancel = types.InlineKeyboardButton(text='\u274c', callback_data='cancel')
                down = types.InlineKeyboardButton(text='\ud83d\udd3b', callback_data='down')
                number = types.InlineKeyboardButton(text=str(value[0]) + ' шт.', callback_data='/')
                up = types.InlineKeyboardButton(text='\ud83d\udd3a', callback_data='up')
                order = types.InlineKeyboardButton(text='Подвердить выбор \u2705', callback_data=str(calldataa + 10000))
                count.add(cancel, number, up)
                count.add(order)
                if message[0] != None:
                    val = (call.message.chat.id,)
                    cursor.execute(sql, val)
                    messages = cursor.fetchall()
                    order_markup = types.InlineKeyboardMarkup(1)
                    button = types.InlineKeyboardButton(text='В корзину',
                                                        callback_data=str((calldata + cate)))
                    order_markup.add(button)
                    if calldataa == len(messages) - 1:
                        back_menu = types.InlineKeyboardButton(text='В меню \u2b05', callback_data='backmenu')
                        count.row(back_menu)
                    if calldata == len(messages) - 1:
                        back_menu = types.InlineKeyboardButton(text='В меню \u2b05', callback_data='backmenu')
                        order_markup.row(back_menu)
                    bot.edit_message_reply_markup(call.message.chat.id, message[0], reply_markup=order_markup)
                else:
                    if calldataa == len(messages) - 1:
                        back_menu = types.InlineKeyboardButton(text='В меню \u2b05', callback_data='backmenu')
                        count.row(back_menu)
                bot.edit_message_reply_markup(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                      reply_markup=count)
            else:
                    sql = 'UPDATE users SET value = %s WHERE user_id = %s'
                    val = (None, call.message.chat.id)
                    cursor.execute(sql, val)
                    db.commit()
                    call_check1 = 'SELECT calldata FROM calldata WHERE user_id = %s'
                    call_check2 = (call.message.chat.id,)
                    cursor.execute(call_check1, call_check2)
                    call_fetch = cursor.fetchone()
                    calldataa = call_fetch[0]
                    for o in range(len(fetchcheck)):
                        if (fetchcheck[o])[0] == (fetchfood[calldataa])[0]:
                            sql = 'UPDATE calldata SET message = %s WHERE user_id = %s'
                            val = (None, call.message.chat.id)
                            cursor.execute(sql, val)
                            db.commit()
                            sql = 'SELECT message FROM messages WHERE user_id = %s'
                            val = (call.message.chat.id,)
                            cursor.execute(sql, val)
                            messages = cursor.fetchall()
                            sll = 'UPDATE users SET value = %s WHERE user_id = %s'
                            vll = (10000, call.message.chat.id)
                            cursor.execute(sll, vll)
                            db.commit()
                            choice = types.InlineKeyboardMarkup(1)
                            menu = types.InlineKeyboardButton('Уже выбрано', callback_data='.')
                            basket = types.InlineKeyboardButton('Корзина \ud83d\uded2', callback_data='cart1')
                            choice.add(menu, basket)
                            if message[0] != None:
                                sql = 'SELECT message FROM messages WHERE user_id = %s'
                                val = (call.message.chat.id,)
                                cursor.execute(sql, val)
                                messages = cursor.fetchall()
                                order_markup = types.InlineKeyboardMarkup(1)
                                button = types.InlineKeyboardButton(text='В корзину',
                                                                    callback_data=str((calldata + cate)))
                                order_markup.add(button)
                                if calldataa == len(messages) - 1:
                                    back_menu = types.InlineKeyboardButton(text='В меню \u2b05',
                                                                           callback_data='backmenu')
                                    choice.row(back_menu)
                                if calldata == len(messages) - 1:
                                    back_menu = types.InlineKeyboardButton(text='В меню \u2b05',
                                                                           callback_data='backmenu')
                                    order_markup.row(back_menu)
                                bot.edit_message_reply_markup(call.message.chat.id, message[0],
                                                              reply_markup=order_markup)
                                bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=choice)
                            else:
                                if calldataa == len(messages) - 1:
                                    back_menu = types.InlineKeyboardButton(text='В меню \u2b05',
                                                                           callback_data='backmenu')
                                    choice.row(back_menu)
                                bot.edit_message_reply_markup(chat_id=call.message.chat.id,
                                                              message_id=call.message.message_id, reply_markup=choice)
                    sel = 'SELECT value FROM users WHERE user_id = %s'
                    selval = (call.message.chat.id,)
                    cursor.execute(sel, selval)
                    value = cursor.fetchone()
                    if value[0] != 10000:
                        sql = 'SELECT message FROM messages WHERE user_id = %s'
                        val = (call.message.chat.id,)
                        cursor.execute(sql, val)
                        messages = cursor.fetchall()
                        call_check1 = 'SELECT calldata FROM calldata WHERE user_id = %s'
                        call_check2 = (call.message.chat.id,)
                        cursor.execute(call_check1, call_check2)
                        call_fetch = cursor.fetchone()
                        calldataa = call_fetch[0]
                        sql2 = 'SELECT product, price, link FROM productlist WHERE type = %s'
                        val2 = (category[category_call[0]][0],)
                        cursor.execute(sql2, val2)
                        fetchfood = cursor.fetchall()
                        sql = 'UPDATE users SET value = %s WHERE user_id = %s'
                        val = (1, call.message.chat.id)
                        cursor.execute(sql, val)
                        db.commit()
                        sql = 'SELECT value FROM users WHERE user_id = %s'
                        val = (call.message.chat.id,)
                        cursor.execute(sql, val)
                        value = cursor.fetchone()
                        count = types.InlineKeyboardMarkup(row_width=4)
                        cancel = types.InlineKeyboardButton(text='\u274c', callback_data='cancel')
                        down = types.InlineKeyboardButton(text='\ud83d\udd3b', callback_data='down')
                        number = types.InlineKeyboardButton(text=str(value[0]) + ' шт.', callback_data='/')
                        up = types.InlineKeyboardButton(text='\ud83d\udd3a', callback_data='up')
                        order = types.InlineKeyboardButton(text='Подтвердить выбор \u2705', callback_data=str(calldataa + 10000))
                        count.add(cancel, number, up)
                        count.add(order)
                        if message[0] != None:
                            sql = 'SELECT message FROM messages WHERE user_id = %s'
                            val = (call.message.chat.id,)
                            cursor.execute(sql, val)
                            messages = cursor.fetchall()
                            order_markup = types.InlineKeyboardMarkup(1)
                            button = types.InlineKeyboardButton(text='В корзину',
                                                                callback_data=str((calldata + cate)))
                            order_markup.add(button)
                            if calldataa == len(messages) - 1:
                                back_menu = types.InlineKeyboardButton(text='В меню \u2b05', callback_data='backmenu')
                                count.row(back_menu)
                            if calldata == len(messages) - 1:
                                back_menu = types.InlineKeyboardButton(text='В меню \u2b05', callback_data='backmenu')
                                order_markup.row(back_menu)
                            bot.edit_message_reply_markup(call.message.chat.id, message[0], reply_markup=order_markup)
                        else:
                            if calldataa == len(messages) - 1:
                                back_menu = types.InlineKeyboardButton(text='В меню \u2b05', callback_data='backmenu')
                                count.row(back_menu)
                        bot.edit_message_reply_markup(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                                      reply_markup=count)


    if call.data == 'back':
        start_buttons = types.InlineKeyboardMarkup()
        menu = types.InlineKeyboardButton(text='Меню \ud83d\udcd6', callback_data='menu')
        basket = types.InlineKeyboardButton(text='Корзина \ud83d\uded2', callback_data='cart')
        start_buttons.add(menu, basket)
        bot.edit_message_text(text='[ ](https://imbt.ga/vfcM0yHFjI)' + start_text, chat_id=call.message.chat.id, message_id=call.message.message_id, parse_mode='Markdown', reply_markup=start_buttons)
    if call.data == 'backcart':
        start_buttons = types.InlineKeyboardMarkup()
        menu = types.InlineKeyboardButton(text='Меню \ud83d\udcd6', callback_data='menu')
        basket = types.InlineKeyboardButton(text='Корзина \ud83d\uded2', callback_data='cart')
        start_buttons.add(menu, basket)
        bot.edit_message_text(text='[ ](https://imbt.ga/vfcM0yHFjI)' + start_text, chat_id=call.message.chat.id,
                              message_id=call.message.message_id, parse_mode='Markdown', reply_markup=start_buttons)
    if call.data == 'cancel':
        sql = 'SELECT name FROM categorylist WHERE seen = %s'
        val = ('<b>ДА</b>',)
        cursor.execute(sql, val)
        cate = cursor.rowcount
        sql = 'SELECT message FROM messages WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        messages = cursor.fetchall()
        call_check1 = 'SELECT calldata FROM calldata WHERE user_id = %s'
        call_check2 = (call.message.chat.id,)
        cursor.execute(call_check1, call_check2)
        call_fetch = cursor.fetchone()
        calldata = call_fetch[0]
        sql = 'UPDATE calldata SET message = %s WHERE user_id = %s'
        val = (None, call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        button = types.InlineKeyboardButton('В корзину', callback_data=str(calldata + cate))
        markup = types.InlineKeyboardMarkup(1)
        markup.add(button)
        if calldata == len(messages) - 1:
            button = types.InlineKeyboardButton('Назад \u2b05', callback_data='backmenu')
            markup.row(button)
        bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=markup)

    if call.data == 'backmenu':
        sql = 'SELECT name FROM categorylist WHERE seen = %s'
        val = ('<b>ДА</b>',)
        cursor.execute(sql, val)
        category = cursor.fetchall()
        sel = 'SELECT message FROM messages WHERE user_id = %s'
        selval = (call.message.chat.id,)
        cursor.execute(sel, selval)
        messages = cursor.fetchall()
        food_type = types.InlineKeyboardMarkup(row_width=2)
        if cursor.rowcount == 0:
            for i in range(len(category)):
                button = types.InlineKeyboardButton(category[i][0], callback_data=str(i))
                food_type.add(button)
            back = types.InlineKeyboardButton('Назад \u2b05', callback_data='back')
            food_type.row(back)
            bot.edit_message_text('Выберите желаемую категорию, пожалуйста:', call.message.chat.id, call.message.message_id, reply_markup=food_type)
        else:
            sql = 'SELECT start_msg FROM calldata WHERE user_id = %s'
            val = (call.message.chat.id, )
            cursor.execute(sql, val)
            msg = cursor.fetchone()[0]
            bot.delete_message(call.message.chat.id, msg)
            for d in range(len(messages)):
                bot.delete_message(call.message.chat.id, messages[d][0])
            for i in range(len(category)):
                button = types.InlineKeyboardButton(category[i][0], callback_data=str(i))
                food_type.add(button)
            discos = types.InlineKeyboardButton('\ud83d\udd25Акции\ud83d\udd25', callback_data='diisc')
            back = types.InlineKeyboardButton('Назад \u2b05', callback_data='back')
            food_type.add(discos)
            food_type.add(back)
            bot.send_message(text='Выберите желаемую категорию, пожалуйста:', chat_id=call.message.chat.id,
                                  reply_markup=food_type)
        sql = 'DELETE FROM messages WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        db.commit()


    #ПАПИН ЖЕЛАЕМЫЙ ИНТЕРФЕЙС
    elif call.data == 'up':
        call_check1 = 'SELECT calldata FROM calldata WHERE user_id = %s'
        call_check2 = (call.message.chat.id,)
        cursor.execute(call_check1, call_check2)
        call_fetch = cursor.fetchone()
        calldata = call_fetch[0]
        sql = 'SELECT value FROM users WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        value = cursor.fetchone()
        sql = 'UPDATE users SET value = %s WHERE user_id = %s'
        val = (value[0] + 1, call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT value FROM users WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        value = cursor.fetchone()
        count = types.InlineKeyboardMarkup(row_width=4)
        cancel = types.InlineKeyboardButton(text='\u274c', callback_data='cancel')
        down = types.InlineKeyboardButton(text='\ud83d\udd3b', callback_data='down')
        number = types.InlineKeyboardButton(text=str(value[0]) + ' шт.', callback_data='/')
        up = types.InlineKeyboardButton(text='\ud83d\udd3a', callback_data='up')
        order = types.InlineKeyboardButton(text='Подтвердить выбор \u2705', callback_data=str(calldata + 10000))
        count.add(cancel, down, number, up)
        count.add(order)
        sql = 'SELECT message FROM messages WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        messages = cursor.fetchall()
        if calldata == len(messages) - 1:
            back_menu = types.InlineKeyboardButton(text='В меню \u2b05', callback_data='backmenu')
            count.row(back_menu)
        bot.edit_message_reply_markup(chat_id=call.message.chat.id, message_id=call.message.message_id, reply_markup=count)
    elif call.data == 'down':
        call_check1 = 'SELECT calldata FROM calldata WHERE user_id = %s'
        call_check2 = (call.message.chat.id,)
        cursor.execute(call_check1, call_check2)
        call_fetch = cursor.fetchone()
        calldata = call_fetch[0]
        sql = 'SELECT value FROM users WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        value = cursor.fetchone()
        sql = 'UPDATE users SET value = %s WHERE user_id = %s'
        val = (value[0] - 1, call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT value FROM users WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        value = cursor.fetchone()
        count = types.InlineKeyboardMarkup(row_width=4)
        cancel = types.InlineKeyboardButton(text='\u274c', callback_data='cancel')
        down = types.InlineKeyboardButton(text='\ud83d\udd3b', callback_data='down')
        number = types.InlineKeyboardButton(text=str(value[0]) + ' шт.', callback_data='/')
        up = types.InlineKeyboardButton(text='\ud83d\udd3a', callback_data='up')
        order = types.InlineKeyboardButton(text='Подтвердить выбор \u2705', callback_data=str(calldata + 10000))
        count.add(cancel, down, number, up)
        count.add(order)
        sql = 'SELECT message FROM messages WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        messages = cursor.fetchall()
        if calldata == len(messages) - 1:
            back_menu = types.InlineKeyboardButton(text='В меню \u2b05', callback_data='backmenu')
            count.row(back_menu)
        bot.edit_message_reply_markup(chat_id=call.message.chat.id, message_id=call.message.message_id, reply_markup=count)
        if value[0] == 1:
            count = types.InlineKeyboardMarkup(row_width=4)
            cancel = types.InlineKeyboardButton(text='\u274c', callback_data='cancel')
            down = types.InlineKeyboardButton(text='\ud83d\udd3b', callback_data='down')
            number = types.InlineKeyboardButton(text=str(value[0]) + ' шт.', callback_data='/')
            up = types.InlineKeyboardButton(text='\ud83d\udd3a', callback_data='up')
            order = types.InlineKeyboardButton(text='Подтвердить выбор \u2705', callback_data=str(calldata + 10000))
            count.add(cancel, number, up)
            count.add(order)
            if calldata == len(messages) - 1:
                back_menu = types.InlineKeyboardButton(text='В меню \u2b05', callback_data='backmenu')
                count.row(back_menu)
            bot.edit_message_reply_markup(chat_id=call.message.chat.id, message_id=call.message.message_id, reply_markup=count)


    if call.data.isdigit():
        if int(call.data) >= 10000:
            sel = 'SELECT message FROM messages WHERE user_id = %s'
            selval = (call.message.chat.id,)
            cursor.execute(sel, selval)
            messages = cursor.fetchall()
            sql = 'SELECT calldata2 FROM calldata WHERE user_id = %s'
            val = (call.message.chat.id,)
            cursor.execute(sql, val)
            category_call = cursor.fetchone()
            sql = 'SELECT value FROM users WHERE user_id = %s'
            val = (call.message.chat.id,)
            cursor.execute(sql, val)
            value = cursor.fetchone()
            call_check1 = 'SELECT calldata FROM calldata WHERE user_id = %s'
            call_check2 = (call.message.chat.id,)
            cursor.execute(call_check1, call_check2)
            call_fetch = cursor.fetchone()
            calldata = (call_fetch)[0]
            sql = 'SELECT name FROM categorylist WHERE seen = %s'
            val = ('<b>ДА</b>',)
            cursor.execute(sql, val)
            category = cursor.fetchall()
            sql2 = 'SELECT product, price, link, desc_ FROM productlist WHERE type = %s'
            val2 = (category[category_call[0]][0],)
            cursor.execute(sql2, val2)
            fetchfood = cursor.fetchall()
            choice = types.InlineKeyboardMarkup()
            basket = types.InlineKeyboardButton('Корзина \ud83d\uded2', callback_data='cart1')
            choice.add(basket)
            if calldata == len(messages) - 1:
                back_menu = types.InlineKeyboardButton(text='В меню \u2b05', callback_data='backmenu')
                choice.row(back_menu)
            sql = 'INSERT INTO cart (link, product, amount, price, cart_id) \
                       VALUES (%s, %s, %s, %s, %s)'
            price = (value[0]*int(fetchfood[calldata][1]))
            user_id = call.message.chat.id
            val = ((fetchfood[calldata])[2], (fetchfood[calldata])[0], value[0], price, user_id)
            cursor.execute(sql, val)
            db.commit()
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text='[ ](' + fetchfood[calldata][2] + ')*Корзина пополнена!*\n*' + (fetchfood[calldata])[0] + '* - ' + str((fetchfood[calldata])[1]) + 'тг.\n\n' + fetchfood[calldata][3], parse_mode='Markdown', reply_markup=choice)
            bot.answer_callback_query(call.id, 'Корзина пополнена - ' + (fetchfood[calldata])[0])
            sql = 'UPDATE calldata SET message = %s WHERE user_id = %s'
            val = (None, call.message.chat.id)
            cursor.execute(sql, val)
            db.commit()
    if call.data == 'cart1':
        sql = 'SELECT name FROM categorylist WHERE seen = %s'
        val = ('<b>ДА</b>',)
        cursor.execute(sql, val)
        category = cursor.fetchall()
        sel = 'SELECT message FROM messages WHERE user_id = %s'
        selval = (call.message.chat.id,)
        cursor.execute(sel, selval)
        messages = cursor.fetchall()
        for d in range(len(messages)):
            bot.delete_message(call.message.chat.id, messages[d][0])
        sql = 'SELECT start_msg FROM calldata WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        msg = cursor.fetchone()[0]
        bot.delete_message(call.message.chat.id, msg)
        sql2 = 'SELECT link, product, amount, price FROM cart WHERE cart_id = %s'
        val2 = (call.message.chat.id,)
        cursor.execute(sql2, val2)
        fetch = cursor.fetchall()
        if fetch == []:  # Если в корзине ничего нет...
            from_cart = types.InlineKeyboardMarkup()
            back = types.InlineKeyboardButton(text='Назад \u2b05', callback_data='back')
            from_cart.add(back)
            bot.send_message(text='В вашей корзине ничего нет :(', chat_id=call.message.chat.id,
                                   reply_markup=from_cart)
        elif any(fetch) == True:  # Если в корзине что то есть...
            sqlms = 'UPDATE users SET value = %s WHERE user_id = %s'
            valms = (0, call.message.chat.id)
            cursor.execute(sqlms, valms)
            db.commit()
            sql = 'SELECT value FROM users WHERE user_id = %s'
            val = (call.message.chat.id,)
            cursor.execute(sql, val)
            slide_fetch = cursor.fetchone()
            select = 'SELECT price FROM cart WHERE cart_id = %s'
            select2 = (call.message.chat.id,)
            cursor.execute(select, select2)
            fetch_price = cursor.fetchall()
            price_list = []
            vals = slide_fetch[0] + 1
            val_ = slide_fetch[0]
            for i in range(len(fetch_price)):  # Подсчет конечной суммы и оформление заказа
                sc = fetch_price[i]
                price_list.append(sc[0])
            sql = 'SELECT * FROM productlist'
            cursor.execute(sql)
            prod = cursor.rowcount
            order_change = types.InlineKeyboardMarkup(row_width=3)
            down = types.InlineKeyboardButton(text='\ud83d\udd3b', callback_data='downc')
            number = types.InlineKeyboardButton(text=str((fetch[val_])[2]) + ' шт.', callback_data='/')
            up = types.InlineKeyboardButton(text='\ud83d\udd3a', callback_data='upc')
            pay = types.InlineKeyboardButton(text='Оформить заказ на ' + str(sum(price_list)) + 'тг.',
                                             callback_data='pay')
            remove_option = types.InlineKeyboardButton(text='Удалить \ud83d\uddd1', callback_data=str(val_ + 1000))
            back_cartver = types.InlineKeyboardButton(text='Назад \u2b05', callback_data='backcart')
            counter = types.InlineKeyboardButton(str(vals) + ' из ' + str(len(fetch)), callback_data='/')
            right = types.InlineKeyboardButton('\u25b6', callback_data='right')
            if (fetch[val_])[2] > 1:
                order_change.add(down, number, up)
            else:
                order_change.add(number, up)
            if len(fetch) > 1:
                order_change.row(counter, right)
                order_change.row(back_cartver, remove_option)
                order_change.row(pay)
            else:
                order_change.row(counter)
                order_change.row(back_cartver, remove_option)
                order_change.row(pay)
            bot.send_message(parse_mode='Markdown', chat_id=call.message.chat.id,
                                  text='[ ](' + str((fetch[val_])[0]) + ')' +
                                                                           str((fetch[val_])[1]) + ' - ' + str(
                    (fetch[val_])[2]) + ' шт.' + '\nСтоймость - ' + str(
                    (fetch[val_])[3]) + 'тг.', reply_markup=order_change)
    if call.data == 'cart':      #Если была выбрана корзина...
        sql2 = 'SELECT link, product, amount, price FROM cart WHERE cart_id = %s'
        val2 = (call.message.chat.id, )
        cursor.execute(sql2, val2)
        fetch = cursor.fetchall()
        if fetch == []:      #Если в корзине ничего нет...
            from_cart = types.InlineKeyboardMarkup()
            back = types.InlineKeyboardButton(text='Назад \u2b05', callback_data='back')
            from_cart.add(back)
            bot.edit_message_text(text='В вашей корзине ничего нет :(', chat_id=call.message.chat.id, message_id=call.message.message_id, reply_markup=from_cart)
        elif any(fetch) == True:    # Если в корзине что то есть...
            sqlms = 'UPDATE users SET value = %s WHERE user_id = %s'
            valms = (0, call.message.chat.id)
            cursor.execute(sqlms, valms)
            db.commit()
            sql = 'SELECT value FROM users WHERE user_id = %s'
            val = (call.message.chat.id, )
            cursor.execute(sql, val)
            slide_fetch = cursor.fetchone()
            select = 'SELECT price FROM cart WHERE cart_id = %s'
            select2 = (call.message.chat.id, )
            cursor.execute(select, select2)
            fetch_price = cursor.fetchall()
            price_list = []
            vals = slide_fetch[0] + 1
            val_ = slide_fetch[0]
            for i in range(len(fetch_price)):   #Подсчет конечной суммы и оформление заказа
                sc = fetch_price[i]
                price_list.append(sc[0])
            sql = 'SELECT * FROM productlist'
            cursor.execute(sql)
            prod = cursor.rowcount
            order_change = types.InlineKeyboardMarkup(row_width=3)
            down = types.InlineKeyboardButton(text='\ud83d\udd3b', callback_data='downc')
            number = types.InlineKeyboardButton(text=str((fetch[val_])[2]) + ' шт.', callback_data='/')
            up = types.InlineKeyboardButton(text='\ud83d\udd3a', callback_data='upc')
            pay = types.InlineKeyboardButton(text='Оформить заказ на ' + str(sum(price_list)) + 'тг.', callback_data='pay')
            remove_option = types.InlineKeyboardButton(text='Удалить \ud83d\uddd1', callback_data=str(val_ + 1000))
            back_cartver = types.InlineKeyboardButton(text='Назад \u2b05', callback_data='backcart')
            counter = types.InlineKeyboardButton(str(vals) + ' из ' + str(len(fetch)), callback_data='/')
            right = types.InlineKeyboardButton('\u25b6', callback_data='right')
            if (fetch[val_])[2] > 1:
                order_change.add(down, number, up)
            else:
                order_change.add(number, up)
            if len(fetch) > 1:
                order_change.row(counter, right)
                order_change.row(back_cartver, remove_option)
                order_change.row(pay)
            else:
                order_change.row(counter)
                order_change.row(back_cartver, remove_option)
                order_change.row(pay)
            bot.edit_message_text(parse_mode='Markdown', chat_id=call.message.chat.id, message_id=call.message.message_id, text='[ ](' + str((fetch[val_])[0]) + ')' +
                           str((fetch[val_])[1]) + ' - ' + str((fetch[val_])[2]) + ' шт.' + '\nСтоймость - ' + str(
                               (fetch[val_])[3]) + 'тг.', reply_markup=order_change)
    if call.data == 'right':
        sql2 = 'SELECT link, product, amount, price FROM cart WHERE cart_id = %s'
        val2 = (call.message.chat.id,)
        cursor.execute(sql2, val2)
        fetch = cursor.fetchall()
        sql = 'SELECT value FROM users WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        slide_fetch = cursor.fetchone()
        sqlms = 'UPDATE users SET value = %s WHERE user_id = %s'
        valms = (slide_fetch[0] + 1, call.message.chat.id)
        cursor.execute(sqlms, valms)
        db.commit()
        sql = 'SELECT value FROM users WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        slide_fetch = cursor.fetchone()
        select = 'SELECT price FROM cart WHERE cart_id = %s'
        select2 = (call.message.chat.id,)
        cursor.execute(select, select2)
        fetch_price = cursor.fetchall()
        price_list = []
        vals = slide_fetch[0] + 1
        val_ = slide_fetch[0]
        for i in range(len(fetch_price)):  # Подсчет конечной суммы и оформление заказа
            sc = fetch_price[i]
            price_list.append(sc[0])
        sql = 'SELECT * FROM productlist'
        cursor.execute(sql)
        prod = cursor.rowcount
        order_change = types.InlineKeyboardMarkup(row_width=3)
        down = types.InlineKeyboardButton(text='\ud83d\udd3b', callback_data='downc')
        number = types.InlineKeyboardButton(text=str((fetch[val_])[2]) + ' шт.', callback_data='/')
        up = types.InlineKeyboardButton(text='\ud83d\udd3a', callback_data='upc')
        pay = types.InlineKeyboardButton(text='Оформить заказ на ' + str(sum(price_list)) + 'тг.', callback_data='pay')
        remove_option = types.InlineKeyboardButton(text='Удалить \ud83d\uddd1', callback_data=str(val_ + 1000))
        back_cartver = types.InlineKeyboardButton(text='Назад \u2b05', callback_data='backcart')
        counter = types.InlineKeyboardButton(str(vals) + ' из ' + str(len(fetch)), callback_data='/')
        right = types.InlineKeyboardButton('\u25b6', callback_data='right')
        left = types.InlineKeyboardButton('\u25c0', callback_data='left')
        print(fetch[val_][2])
        if (fetch[val_])[2] > 1:
            order_change.add(down, number, up)
        else:
            order_change.add(number, up)
        if vals == len(fetch):
            order_change.row(left, counter)
            order_change.row(back_cartver, remove_option)
            order_change.row(pay)
        elif 1 < vals < len(fetch):
            order_change.row(left, counter, right)
            order_change.row(back_cartver, remove_option)
            order_change.row(pay)
        bot.edit_message_text(parse_mode='Markdown', chat_id=call.message.chat.id, message_id=call.message.message_id,
                              text='[ ](' + str((fetch[val_])[0]) + ')' +
                                   str((fetch[val_])[1]) + ' - ' + str(
                                  (fetch[val_])[2]) + ' шт.' + '\nСтоймость - ' + str(
                                  (fetch[val_])[3]) + 'тг.', reply_markup=order_change)
    if call.data == 'left':
        sql2 = 'SELECT link, product, amount, price FROM cart WHERE cart_id = %s'
        val2 = (call.message.chat.id,)
        cursor.execute(sql2, val2)
        fetch = cursor.fetchall()
        sql = 'SELECT value FROM users WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        slide_fetch = cursor.fetchone()
        sqlms = 'UPDATE users SET value = %s WHERE user_id = %s'
        valms = (slide_fetch[0] - 1, call.message.chat.id)
        cursor.execute(sqlms, valms)
        db.commit()
        sql = 'SELECT value FROM users WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        slide_fetch = cursor.fetchone()
        select = 'SELECT price FROM cart WHERE cart_id = %s'
        select2 = (call.message.chat.id,)
        cursor.execute(select, select2)
        fetch_price = cursor.fetchall()
        price_list = []
        vals = slide_fetch[0] + 1
        val_ = slide_fetch[0]
        for i in range(len(fetch_price)):  # Подсчет конечной суммы и оформление заказа
            sc = fetch_price[i]
            price_list.append(sc[0])
        sql = 'SELECT * FROM productlist'
        cursor.execute(sql)
        prod = cursor.rowcount
        order_change = types.InlineKeyboardMarkup(row_width=3)
        down = types.InlineKeyboardButton(text='\ud83d\udd3b', callback_data='downc')
        number = types.InlineKeyboardButton(text=str((fetch[val_])[2]) + ' шт.', callback_data='/')
        up = types.InlineKeyboardButton(text='\ud83d\udd3a', callback_data='upc')
        pay = types.InlineKeyboardButton(text='Оформить заказ на ' + str(sum(price_list)) + 'тг.', callback_data='pay')
        remove_option = types.InlineKeyboardButton(text='Удалить \ud83d\uddd1', callback_data=str(val_ + 1000))
        back_cartver = types.InlineKeyboardButton(text='Назад \u2b05', callback_data='backcart')
        counter = types.InlineKeyboardButton(str(vals) + ' из ' + str(len(fetch)), callback_data='/')
        right = types.InlineKeyboardButton('\u25b6', callback_data='right')
        left = types.InlineKeyboardButton('\u25c0', callback_data='left')
        if (fetch[val_])[2] > 1:
            order_change.add(down, number, up)
        else:
            order_change.add(number, up)
        if vals == 1:
            order_change.row(counter, right)
            order_change.row(back_cartver, remove_option)
            order_change.row(pay)
        elif 1 < vals < len(fetch):
            order_change.row(left, counter, right)
            order_change.row(back_cartver, remove_option)
            order_change.row(pay)
        bot.edit_message_text(parse_mode='Markdown', chat_id=call.message.chat.id, message_id=call.message.message_id,
                              text='[ ](' + str((fetch[val_])[0]) + ')' +
                                   str((fetch[val_])[1]) + ' - ' + str(
                                  (fetch[val_])[2]) + ' шт.' + '\nСтоймость - ' + str(
                                  (fetch[val_])[3]) + 'тг.', reply_markup=order_change)

    if call.data.isdigit():     #Если была нажата кнопка убрать...
        if 10000 > int(call.data) >= 1000 :
            calldata = int(call.data) - 1000
            sql2 = 'SELECT link, product, amount, price FROM cart WHERE cart_id = %s'
            val2 = (call.message.chat.id,)
            cursor.execute(sql2, val2)
            fetch = cursor.fetchall()
            sql = 'DELETE FROM cart WHERE cart_id = %s AND product = %s'
            val = (call.message.chat.id, (fetch[calldata])[1])
            cursor.execute(sql, val)
            db.commit()
            sql2 = 'SELECT link, product, amount, price FROM cart WHERE cart_id = %s'
            val2 = (call.message.chat.id,)
            cursor.execute(sql2, val2)
            fetch = cursor.fetchall()
            if any(fetch) == False:
                from_cart = types.InlineKeyboardMarkup()
                back = types.InlineKeyboardButton(text='Назад \u2b05', callback_data='back')
                from_cart.add(back)
                bot.edit_message_text(text='В вашей корзине ничего нет :(', chat_id=call.message.chat.id,
                                      message_id=call.message.message_id, reply_markup=from_cart)
                return
            sql = 'SELECT value FROM users WHERE user_id = %s'
            val = (call.message.chat.id,)
            cursor.execute(sql, val)
            slide_fetch = cursor.fetchone()
            select = 'SELECT price FROM cart WHERE cart_id = %s'
            select2 = (call.message.chat.id,)
            cursor.execute(select, select2)
            fetch_price = cursor.fetchall()
            price_list = []
            for i in range(len(fetch_price)):  # Подсчет конечной суммы и оформление заказа
                sc = fetch_price[i]
                price_list.append(sc[0])
            vals = slide_fetch[0] + 1
            val_ = slide_fetch[0]
            if vals - 1 != 0:
                sqll = 'UPDATE users SET value = %s WHERE user_id = %s'
                vall = (val_ - 1, call.message.chat.id)
                cursor.execute(sqll, vall)
                db.commit()
            if len(fetch) == 1:
                sqll = 'UPDATE users SET value = %s WHERE user_id = %s'
                vall = (0, call.message.chat.id)
                cursor.execute(sqll, vall)
                db.commit()
            sql = 'SELECT value FROM users WHERE user_id = %s'
            val = (call.message.chat.id,)
            cursor.execute(sql, val)
            slide_fetch = cursor.fetchone()
            vals = slide_fetch[0] + 1
            val_ = slide_fetch[0]
            order_change = types.InlineKeyboardMarkup(row_width=3)
            down = types.InlineKeyboardButton(text='\ud83d\udd3b', callback_data='downc')
            number = types.InlineKeyboardButton(text=str((fetch[val_])[2]) + ' шт.', callback_data='/')
            up = types.InlineKeyboardButton(text='\ud83d\udd3a', callback_data='upc')
            pay = types.InlineKeyboardButton(text='Оформить заказ на ' + str(sum(price_list)) + 'тг.', callback_data='pay')
            remove_option = types.InlineKeyboardButton(text='Удалить \ud83d\uddd1', callback_data=str(val_ + 1000))
            back_cartver = types.InlineKeyboardButton(text='Назад \u2b05', callback_data='backcart')
            counter = types.InlineKeyboardButton(str(vals) + ' из ' + str(len(fetch)),
                                                 callback_data='/')
            right = types.InlineKeyboardButton('\u27a1', callback_data='right')
            left = types.InlineKeyboardButton('\u25c0', callback_data='left')
            if (fetch[val_])[2] > 1:
                order_change.add(down, number, up)
            elif (fetch[val_])[2] == 1:
                order_change.add(number, up)
            if len(fetch) == 1:
                order_change.row(counter)
                order_change.add(back_cartver, remove_option)
                order_change.row(pay)
                bot.edit_message_text(parse_mode='Markdown', chat_id=call.message.chat.id, message_id=call.message.message_id, text='[ ](' + str((fetch[val_])[0]) + ')' +
                               str((fetch[val_])[1]) + ' - ' + str((fetch[val_])[2]) + ' шт.' + '\nСтоймость - ' + str(
                                   (fetch[val_])[3]) + 'тг.', reply_markup=order_change)
            elif len(fetch) > 1:
                if vals == len(fetch):
                    order_change.add(left, counter)
                    order_change.row(back_cartver, remove_option)
                    order_change.row(pay)
                    bot.edit_message_text(parse_mode='Markdown', chat_id=call.message.chat.id, message_id=call.message.message_id,
                                          text='[ ](' + str((fetch[val_])[0]) + ')' +
                                               str((fetch[val_])[1]) + ' - ' + str(
                                              (fetch[val_])[2]) + ' шт.' + '\nСтоймость - ' + str(
                                              (fetch[val_])[3]) + 'тг.', reply_markup=order_change)
                elif vals == 1:
                    order_change.add(counter, right)
                    order_change.row(back_cartver, remove_option)
                    order_change.row(pay)
                    bot.edit_message_text(parse_mode='Markdown', chat_id=call.message.chat.id, message_id=call.message.message_id,
                                          text='[ ](' + str((fetch[val_])[0]) + ')' +
                                               str((fetch[val_])[1]) + ' - ' + str(
                                              (fetch[val_])[2]) + ' шт.' + '\nСтоймость - ' + str(
                                              (fetch[val_])[3]) + 'тг.', reply_markup=order_change)
                elif 1 < vals < len(fetch):
                    order_change.add(left, counter, right)
                    order_change.row(back_cartver, remove_option)
                    order_change.row(pay)
                    bot.edit_message_text(parse_mode='Markdown', chat_id=call.message.chat.id, message_id=call.message.message_id,
                                          text='[ ](' + str((fetch[val_])[0]) + ')' +
                                               str((fetch[val_])[1]) + ' - ' + str(
                                              (fetch[val_])[2]) + ' шт.' + '\nСтоймость - ' + str(
                                              (fetch[val_])[3]) + 'тг.', reply_markup=order_change)
    if call.data == 'upc':
        order_change = types.InlineKeyboardMarkup(row_width=3)
        sql = 'SELECT value FROM users WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        slide_fetch = cursor.fetchone()
        val_ = slide_fetch[0]
        vals = slide_fetch[0] + 1
        sql2 = 'SELECT link, product, amount, price FROM cart WHERE cart_id = %s'
        val2 = (call.message.chat.id,)
        cursor.execute(sql2, val2)
        fetch = cursor.fetchall()
        div = (fetch[val_])[3]/(fetch[val_])[2]
        sql3 = 'UPDATE cart SET amount = %s WHERE cart_id = %s AND product = %s'
        val3 = ((fetch[slide_fetch[0]])[2] + 1, call.message.chat.id, fetch[val_][1])
        cursor.execute(sql3, val3)
        db.commit()
        sql2 = 'SELECT link, product, amount, price FROM cart WHERE cart_id = %s'
        val2 = (call.message.chat.id,)
        cursor.execute(sql2, val2)
        fetch = cursor.fetchall()
        sql4 = 'UPDATE cart SET price = %s WHERE cart_id = %s AND product = %s'
        val4 = (div*(fetch[val_])[2], call.message.chat.id, fetch[val_][1])
        cursor.execute(sql4, val4)
        db.commit()
        sql2 = 'SELECT link, product, amount, price FROM cart WHERE cart_id = %s'
        val2 = (call.message.chat.id,)
        cursor.execute(sql2, val2)
        fetch = cursor.fetchall()
        select = 'SELECT price FROM cart WHERE cart_id = %s'
        select2 = (call.message.chat.id,)
        cursor.execute(select, select2)
        fetch_price = cursor.fetchall()
        price_list = []
        for i in range(len(fetch_price)):  # Подсчет конечной суммы и оформление заказа
            sc = fetch_price[i]
            price_list.append(sc[0])
        sql = 'SELECT * FROM productlist'
        cursor.execute(sql)
        prod = cursor.rowcount
        down = types.InlineKeyboardButton(text='\ud83d\udd3b', callback_data='downc')
        number = types.InlineKeyboardButton(text=str((fetch[val_])[2]) + ' шт.', callback_data='/')
        up = types.InlineKeyboardButton(text='\ud83d\udd3a', callback_data='upc')
        counter = types.InlineKeyboardButton(str(vals) + ' из ' + str(len(fetch)),
                                             callback_data='/')
        right = types.InlineKeyboardButton('\u27a1', callback_data='right')
        left = types.InlineKeyboardButton('\u25c0', callback_data='left')
        remove_option = types.InlineKeyboardButton(text='Удалить \ud83d\uddd1', callback_data=str(val_ + 1000))
        back_cartver = types.InlineKeyboardButton(text='Назад \u2b05', callback_data='backcart')
        pay = types.InlineKeyboardButton(text='Оформить заказ на ' + str(sum(price_list)) + 'тг.', callback_data='pay')
        order_change.add(down, number, up)
        if vals == 1:
            if len(fetch) == 1:
                order_change.row(counter)
                order_change.row(back_cartver, remove_option)
                order_change.row(pay)
            else:
                order_change.row(counter, right)
                order_change.row(back_cartver, remove_option)
                order_change.row(pay)
        elif 1 < vals < len(fetch):
            order_change.row(left, counter, right)
            order_change.row(back_cartver, remove_option)
            order_change.row(pay)
        elif vals == len(fetch):
            order_change.row(left, counter)
            order_change.row(back_cartver, remove_option)
            order_change.row(pay)
        bot.edit_message_text(parse_mode='Markdown', chat_id=call.message.chat.id, message_id=call.message.message_id,
                              text='[ ](' + str((fetch[val_])[0]) + ')' +
                                   str((fetch[val_])[1]) + ' - ' + str(
                                  (fetch[val_])[2]) + ' шт.' + '\nСтоймость - ' + str(
                                  (fetch[val_])[3]) + 'тг.', reply_markup=order_change)
    if call.data == 'downc':
        order_change = types.InlineKeyboardMarkup(row_width=3)
        sql = 'SELECT value FROM users WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        slide_fetch = cursor.fetchone()
        val_ = slide_fetch[0]
        vals = slide_fetch[0] + 1
        sql2 = 'SELECT link, product, amount, price FROM cart WHERE cart_id = %s'
        val2 = (call.message.chat.id,)
        cursor.execute(sql2, val2)
        fetch = cursor.fetchall()
        div = (fetch[val_])[3] / (fetch[val_])[2]
        sql3 = 'UPDATE cart SET amount = %s WHERE cart_id = %s AND product = %s'
        val3 = ((fetch[slide_fetch[0]])[2] - 1, call.message.chat.id, fetch[val_][1])
        cursor.execute(sql3, val3)
        db.commit()
        sql2 = 'SELECT link, product, amount, price FROM cart WHERE cart_id = %s'
        val2 = (call.message.chat.id,)
        cursor.execute(sql2, val2)
        fetch = cursor.fetchall()
        sql4 = 'UPDATE cart SET price = %s WHERE cart_id = %s AND product = %s'
        val4 = (div * (fetch[val_])[2], call.message.chat.id, fetch[val_][1])
        cursor.execute(sql4, val4)
        db.commit()
        sql2 = 'SELECT link, product, amount, price FROM cart WHERE cart_id = %s'
        val2 = (call.message.chat.id,)
        cursor.execute(sql2, val2)
        fetch = cursor.fetchall()
        select = 'SELECT price FROM cart WHERE cart_id = %s'
        select2 = (call.message.chat.id,)
        cursor.execute(select, select2)
        fetch_price = cursor.fetchall()
        price_list = []
        for i in range(len(fetch_price)):  # Подсчет конечной суммы и оформление заказа
            sc = fetch_price[i]
            price_list.append(sc[0])
        sql = 'SELECT * FROM productlist'
        cursor.execute(sql)
        prod = cursor.rowcount
        down = types.InlineKeyboardButton(text='\ud83d\udd3b', callback_data='downc')
        number = types.InlineKeyboardButton(text=str((fetch[val_])[2]) + ' шт.', callback_data='/')
        up = types.InlineKeyboardButton(text='\ud83d\udd3a', callback_data='upc')
        counter = types.InlineKeyboardButton(str(vals) + ' из ' + str(len(fetch)),
                                             callback_data='/')
        right = types.InlineKeyboardButton('\u27a1', callback_data='right')
        left = types.InlineKeyboardButton('\u25c0', callback_data='left')
        remove_option = types.InlineKeyboardButton(text='Удалить \ud83d\uddd1', callback_data=str(val_ + 1000))
        back_cartver = types.InlineKeyboardButton(text='Назад \u2b05', callback_data='backcart')
        pay = types.InlineKeyboardButton(text='Оформить заказ на ' + str(sum(price_list)) + 'тг.', callback_data='pay')
        if (fetch[val_])[2] == 1:
            order_change.add(number, up)
        else:
            order_change.add(down, number, up)
        if vals == 1:
            if len(fetch) == 1:
                order_change.row(counter)
                order_change.row(back_cartver, remove_option)
                order_change.row(pay)
            else:
                order_change.row(counter, right)
                order_change.row(back_cartver, remove_option)
                order_change.row(pay)
        elif 1 < vals < len(fetch):
            order_change.row(left, counter, right)
            order_change.row(back_cartver, remove_option)
            order_change.row(pay)
        elif vals == len(fetch):
            order_change.row(left, counter)
            order_change.row(back_cartver, remove_option)
            order_change.row(pay)
        bot.edit_message_text(parse_mode='Markdown', chat_id=call.message.chat.id, message_id=call.message.message_id,
                              text='[ ](' + str((fetch[val_])[0]) + ')' +
                                   str((fetch[val_])[1]) + ' - ' + str(
                                  (fetch[val_])[2]) + ' шт.' + '\nСтоймость - ' + str(
                                  (fetch[val_])[3]) + 'тг.', reply_markup=order_change)

    if call.data == 'pay':
        sql = 'INSERT INTO orders (process, user_id, first_name, last_name) VALUES (%s, %s, %s, %s)'
        val = ('ordering', call.message.chat.id, call.message.chat.first_name, call.message.chat.last_name)
        cursor.execute(sql, val)
        db.commit()
        sql = 'UPDATE users SET step = %s WHERE user_id = %s'
        val = ('start', call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        delivery_method = types.InlineKeyboardMarkup()
        delivery = types.InlineKeyboardButton('Курьером', callback_data='delivery')
        get_ur_ass_noob = types.InlineKeyboardButton('Самовывозом', callback_data='noob')
        delivery_method.add(delivery, get_ur_ass_noob)
        alltimemarkup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        alltimebutton = types.KeyboardButton('Отмена\u274c')
        alltimemarkup.add(alltimebutton)
        msg_update = 'UPDATE calldata SET message = %s WHERE user_id = %s'
        msg_vals = (call.message.message_id + 2, call.message.chat.id)
        cursor.execute(msg_update, msg_vals)
        db.commit()
        bot.delete_message(call.message.chat.id, call.message.message_id)
        bot.send_message(reply_markup=alltimemarkup, text='Идет оформление заказа', chat_id=call.message.chat.id)
        bot.send_message(call.message.chat.id, '...')
        bot.send_message(call.message.chat.id, 'Какой способ доставки вам будет удобен?', reply_markup=delivery_method)
    if call.data == 'delivery':
        sql = 'UPDATE orders SET delivery_method = %s WHERE user_id = %s AND process = %s'
        val = ('Курьер', call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        sel = 'SELECT message FROM calldata WHERE user_id = %s'
        selval = (call.message.chat.id,)
        cursor.execute(sel, selval)
        message_fetch = cursor.fetchone()
        add_sel = 'SELECT address FROM users WHERE user_id = %s'
        add_val = (call.message.chat.id,)
        cursor.execute(add_sel, add_val)
        add_check = cursor.fetchone()
        if add_check[0] == None:
            sql = 'UPDATE users SET step = %s WHERE user_id = %s'
            val = ('address', call.message.chat.id)
            cursor.execute(sql, val)
            bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.message_id)
            bot.edit_message_text('Способ доставки - курьер' + '\n...', message_id=message_fetch[0], chat_id=call.message.chat.id)
            bot.send_message(text='Пожалуйста, введите свой адрес', chat_id=call.message.chat.id)
        else:
            sql = 'UPDATE users SET step = %s WHERE user_id = %s'
            val = ('address', call.message.chat.id)
            cursor.execute(sql, val)
            db.commit()
            address_markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1, one_time_keyboard=True)
            address = types.KeyboardButton(add_check[0])
            new_add = types.KeyboardButton('Изменить адрес')
            alltimebutton = types.KeyboardButton('Отмена\u274c')
            address_markup.add(address, new_add, alltimebutton)
            bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.message_id)
            bot.edit_message_text(text='Способ доставки - курьер' + '\n...', message_id=message_fetch[0], chat_id=call.message.chat.id)
            bot.send_message(text='Пожалуйста, введите свой адрес', chat_id=call.message.chat.id,
                             reply_markup=address_markup)
    if call.data == 'money':
        sql = 'UPDATE orders SET pay_method = %s WHERE user_id = %s AND process = %s'
        val = ('Наличными', call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        alltimemarkup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        alltimebutton = types.KeyboardButton('Отмена\u274c')
        alltimemarkup.add(alltimebutton)
        bot.edit_message_text('Выберите способ оплаты: наличными', call.message.chat.id, call.message.message_id, reply_markup=None)
        select = 'SELECT price FROM cart WHERE cart_id = %s'
        select2 = (call.message.chat.id,)
        cursor.execute(select, select2)
        fetch_price = cursor.fetchall()
        price_list = []
        for i in range(len(fetch_price)):
            sc = fetch_price[i]
            price_list.append(sc[0])
        sel = 'SELECT message FROM calldata WHERE user_id = %s'
        selval = (call.message.chat.id,)
        cursor.execute(sel, selval)
        message_fetch = cursor.fetchone()
        sql_add = 'SELECT address FROM users WHERE user_id = %s'
        sql_val = (call.message.chat.id, )
        cursor.execute(sql_add, sql_val)
        add_fetch = cursor.fetchone()
        sql = 'UPDATE users SET step = %s WHERE user_id = %s'
        val = ('money', call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        bot.edit_message_text('Способ доставки - курьер' + '\nАдрес заказчика - ' + add_fetch[0] + '\nСпособ оплаты - наличными' + '\n...', call.message.chat.id, message_fetch[0])
        bot.send_message(call.message.chat.id, 'С какой суммой вам потребуется сдача?' + '\n*Общая стоимость заказа - ' + str(sum(price_list)) + 'тг.*', reply_markup=alltimemarkup, parse_mode='Markdown')
    if call.data == 'confirm':
        sql = 'SELECT price FROM cart WHERE cart_id = %s'
        val = (call.message.chat.id, )
        cursor.execute(sql, val)
        fetch_price = cursor.fetchall()
        price_list = []
        for i in range(len(fetch_price)):  # Подсчет конечной суммы и оформление заказа
            sc = fetch_price[i]
            price_list.append(sc[0])
        sql = 'UPDATE orders SET price = %s WHERE user_id = %s AND process = %s'
        val = (sum(price_list), call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT address FROM users WHERE user_id = %s'
        val = (call.message.chat.id, )
        cursor.execute(sql, val)
        address = cursor.fetchone()
        sql = 'SELECT phone_number FROM users WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        phone = cursor.fetchone()
        sql = 'UPDATE orders SET address = %s WHERE user_id = %s AND process = %s'
        val = (address[0], call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        sql = 'UPDATE orders SET phone_number = %s WHERE user_id = %s AND process = %s'
        val = (phone[0], call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        sql = 'UPDATE orders SET order_date = %s WHERE user_id = %s AND process = %s'
        val = (str(datetime.datetime.today().date()) + ', ' + str(datetime.datetime.today().time()), call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        sql = 'UPDATE orders SET full_desc = %s WHERE user_id = %s AND process = %s'
        val = (call.message.text, call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        sql = 'UPDATE orders SET process = %s WHERE user_id = %s AND process = %s'
        val = ('finished', call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        sql = 'UPDATE users SET step = %s WHERE user_id = %s'
        val = (None, call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        sql = 'DELETE FROM cart WHERE cart_id = %s'
        val = (call.message.chat.id, )
        cursor.execute(sql,val)
        db.commit()
        bot.forward_message(dad_id, call.message.chat.id, call.message.message_id)
        bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
        main_page = types.InlineKeyboardMarkup()
        main_pagebutton = types.InlineKeyboardButton('Главная\ud83c\udfe0', callback_data='home')
        main_page.add(main_pagebutton)
        bot.send_message(call.message.chat.id, 'Ваш заказ успешно оформлен!', reply_markup=main_page)
    if call.data == 'home':
        bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
        start_buttons = types.InlineKeyboardMarkup()
        menu = types.InlineKeyboardButton(text='Меню \ud83d\udcd6', callback_data='menu')
        basket = types.InlineKeyboardButton(text='Корзина \ud83d\uded2', callback_data='cart')
        start_buttons.add(menu, basket)
        additional_kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
        cart = types.KeyboardButton(text='Корзина \ud83d\uded2')
        bot.send_message(chat_id=call.message.chat.id, parse_mode='Markdown', text='[ ](https://imbt.ga/vfcM0yHFjI)' + start_text,
                         reply_markup=start_buttons)
    if call.data == 'bacc':
        sql = 'DELETE FROM orders WHERE user_id = %s AND process = %s'
        val = (call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        bot.delete_message(call.message.chat.id, call.message.message_id - 1)
        bot.delete_message(call.message.chat.id, call.message.message_id)
        start_buttons = types.InlineKeyboardMarkup()
        menu = types.InlineKeyboardButton(text='Меню \ud83d\udcd6', callback_data='menu')
        basket = types.InlineKeyboardButton(text='Корзина \ud83d\uded2', callback_data='cart')
        start_buttons.add(menu, basket)
        additional_kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
        cart = types.KeyboardButton(text='Корзина \ud83d\uded2')
        bot.send_message(chat_id=call.message.chat.id, parse_mode='Markdown',
                              text='[ ](https://imbt.ga/vfcM0yHFjI)' + start_text,
                              reply_markup=start_buttons)
    if call.data == 'noob':
        sql = 'UPDATE orders SET delivery_method = %s WHERE user_id = %s AND process = %s'
        val = ('Самовывоз', call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        sql = 'UPDATE users SET step = %s WHERE user_id = %s'
        val = ('noob', call.message.chat.id)
        cursor.execute(sql, val)
        sel = 'SELECT message FROM calldata WHERE user_id = %s'
        selval = (call.message.chat.id,)
        cursor.execute(sel, selval)
        message_fetch = cursor.fetchone()
        bot.edit_message_text('Способ доставки - самовывоз' + '\n...', call.message.chat.id, message_fetch[0])
        bot.delete_message(call.message.chat.id, call.message.message_id)
        time = types.InlineKeyboardMarkup()
        gottagofast = types.InlineKeyboardButton('Как можно скорее', callback_data='fastt')
        noneedtobefast = types.InlineKeyboardButton('Выбрать время', callback_data='tpick')
        time.add(gottagofast, noneedtobefast)
        bot.send_message(call.message.chat.id, 'Когда вам будет удобно забрать свой заказ?', reply_markup=time)
    if call.data == 'tpick':
        sql = 'UPDATE users SET step = %s WHERE user_id = %s'
        val = ('hours', call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        time_markup = types.InlineKeyboardMarkup()
        bot.edit_message_text('Укажите часы:', call.message.chat.id, call.message.message_id)
    if call.data == 'fastt':
        bot.delete_message(call.message.chat.id, call.message.message_id)
        contact = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1, one_time_keyboard=True)
        contact_button = types.KeyboardButton(text='Отправить контакты', request_contact=True)
        alltimebutton = types.KeyboardButton('Отмена\u274c')
        contact.add(contact_button, alltimebutton)
        sql = 'SELECT * FROM orders WHERE process = %s'
        val = ('finished',)
        cursor.execute(sql, val)
        num = cursor.rowcount
        sql = 'UPDATE orders SET № = %s WHERE user_id = %s AND process = %s'
        val = (num + 1, call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        sql = 'UPDATE orders SET time = %s WHERE user_id = %s AND process = %s'
        val = ('Как можно скорее', call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        conf = 'UPDATE users SET step = %s WHERE user_id = %s'
        conf_val = ('confirm', call.message.chat.id)
        db.commit()
        sql2 = 'SELECT product, amount, price FROM cart WHERE cart_id = %s'
        val2 = (call.message.chat.id,)
        cursor.execute(sql2, val2)
        fetch = cursor.fetchall()
        select = 'SELECT price FROM cart WHERE cart_id = %s'
        select2 = (call.message.chat.id,)
        cursor.execute(select, select2)
        fetch_price = cursor.fetchall()
        sel = 'SELECT message FROM calldata WHERE user_id = %s'
        selval = (call.message.chat.id,)
        cursor.execute(sel, selval)
        message_fetch = cursor.fetchone()
        phone_check = 'SELECT phone_number FROM users WHERE user_id = %s'
        phone_check2 = (call.message.chat.id,)
        cursor.execute(phone_check, phone_check2)
        phone_number = cursor.fetchone()
        if phone_number[0] == None:
            bot.send_message(
                text='Чтобы продолжить вы должны отправить нам ваши контактные данные для дальнейших услуг',
                chat_id=call.message.chat.id, reply_markup=contact)
            bot.edit_message_text(
                'Способ доставки - самовывоз' + '\nУказанное время - как можно скорее' + '\n...', call.message.chat.id, message_fetch[0])
            bot.delete_message(call.message.chat.id, call.message.message_id)
            sql = 'UPDATE users SET step = %s WHERE user_id = %s'
            val = ('fast', call.message.chat.id)
            cursor.execute(sql, val)
            db.commit()
        else:
            price_list = []
            orders = ''
            for i in range(len(fetch_price)):
                sc = fetch_price[i]
                price_list.append(sc[0])
                orders += '\n' + str(i + 1) + '. ' + (fetch[i])[0] + ' ' + str((fetch[i])[1]) + ' шт. - ' + str(
                    (fetch[i])[2]) + 'тг.'
            confirmation = types.InlineKeyboardMarkup()
            confirm = types.InlineKeyboardButton('Подтвердить\u2705', callback_data='confirm')
            uhhhnogetback = types.InlineKeyboardButton('Отмена\u274c', callback_data='bacc')
            confirmation.add(confirm, uhhhnogetback)
            remove = types.ReplyKeyboardRemove()
            bot.send_message(call.message.chat.id, 'Все верно?', reply_markup=remove)
            bot.send_message(parse_mode='Markdown',
                             text='*Заказ №' + str(num+1) + '*\n' + '\nСпособ доставки - самовывоз' + '\nУказанное время - как можно скорее' + '\nИмя - ' + call.message.chat.first_name + '\nНомер телефона - ' + str(
                                 phone_number[0]) + '\n*Ваш заказ:*' + orders + '\n*Общая стоимость - *' + '*' + str(
                                 sum(price_list)) + 'тг.*', chat_id=call.message.chat.id,
                             reply_markup=confirmation)
            bot.delete_message(call.message.chat.id, message_fetch[0])

    if call.data == 'skip':
        sql = 'SELECT * FROM orders WHERE process = %s'
        val = ('finished',)
        cursor.execute(sql, val)
        num = cursor.rowcount
        sql = 'UPDATE orders SET № = %s WHERE user_id = %s AND process = %s'
        val = (num + 1, call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        contact = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1, one_time_keyboard=True)
        contact_button = types.KeyboardButton(text='Отправить контакты', request_contact=True)
        alltimebutton = types.KeyboardButton('Отмена\u274c')
        contact.add(contact_button, alltimebutton)
        conf = 'UPDATE users SET step = %s WHERE user_id = %s'
        conf_val = ('confirm', call.message.chat.id)
        db.commit()
        sql2 = 'SELECT product, amount, price FROM cart WHERE cart_id = %s'
        val2 = (call.message.chat.id,)
        cursor.execute(sql2, val2)
        fetch = cursor.fetchall()
        select = 'SELECT price FROM cart WHERE cart_id = %s'
        select2 = (call.message.chat.id,)
        cursor.execute(select, select2)
        fetch_price = cursor.fetchall()
        sel = 'SELECT message FROM calldata WHERE user_id = %s'
        selval = (call.message.chat.id,)
        cursor.execute(sel, selval)
        message_fetch = cursor.fetchone()
        phone_check = 'SELECT phone_number FROM users WHERE user_id = %s'
        phone_check2 = (call.message.chat.id,)
        cursor.execute(phone_check, phone_check2)
        phone_number = cursor.fetchone()
        sql = 'UPDATE users SET min = %s WHERE user_id = %s'
        val = (0, call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT hour, min FROM users WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        time_fetch = cursor.fetchone()
        sql = 'UPDATE orders SET time = %s WHERE user_id = %s AND process = %s'
        val = (str(time_fetch[0]) + ':00', call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        if phone_number[0] == None:
            bot.send_message(
                text='Чтобы продолжить вы должны отправить нам ваши контактные данные для дальнейших услуг',
                chat_id=call.message.chat.id, reply_markup=contact)
            bot.edit_message_text(
                'Способ доставки - самовывоз' + '\nУказанное время - ' + '\n...', call.message.chat.id,
                message_fetch[0])
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            sql = 'UPDATE users SET step = %s WHERE user_id = %s'
            val = ('noob', call.message.chat.id)
            cursor.execute(sql, val)
            db.commit()
        else:
            price_list = []
            orders = ''
            for i in range(len(fetch_price)):
                sc = fetch_price[i]
                price_list.append(sc[0])
                orders += '\n' + str(i + 1) + '. ' + (fetch[i])[0] + ' ' + str((fetch[i])[1]) + ' шт. - ' + str(
                    (fetch[i])[2]) + 'тг.'
            bot.delete_message(call.message.chat.id, message_fetch[0])
            confirmation = types.InlineKeyboardMarkup()
            confirm = types.InlineKeyboardButton('Подтвердить\u2705', callback_data='confirm')
            uhhhnogetback = types.InlineKeyboardButton('Отмена\u274c', callback_data='bacc')
            confirmation.add(confirm, uhhhnogetback)
            remove = types.ReplyKeyboardRemove()
            bot.send_message(call.message.chat.id, 'Все верно?', reply_markup=remove)
            bot.send_message(parse_mode='Markdown',
                                 text='*Заказ №' + str(num+1) + '*\n' + '\nСпособ доставки - самовывоз' + '\nУказанное время - ' + str(
                                     time_fetch[0]) + ':' + '0' + str(time_fetch[
                                                                          1]) + '\nИмя - ' + call.message.chat.first_name + '\nНомер телефона - ' + str(
                                     phone_number[0]) + '\n*Ваш заказ:*' + orders + '\n*Общая стоимость - *' + '*' + str(
                                     sum(price_list)) + 'тг.*', chat_id=call.message.chat.id,
                                 reply_markup=confirmation)
    if call.data == 'kaspi':
        bot.delete_message(call.message.chat.id, call.message.message_id)
        sql = 'UPDATE orders SET pay_method = %s WHERE user_id = %s AND process = %s'
        val = ('Каспи голд', call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT * FROM orders WHERE process = %s'
        val = ('finished',)
        cursor.execute(sql, val)
        num = cursor.rowcount
        sql = 'UPDATE orders SET № = %s WHERE user_id = %s AND process = %s'
        val = (num + 1, call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        contact = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1, one_time_keyboard=True)
        contact_button = types.KeyboardButton(text='Отправить контакты', request_contact=True)
        alltimebutton = types.KeyboardButton('Отмена\u274c')
        contact.add(contact_button, alltimebutton)
        conf = 'UPDATE users SET step = %s WHERE user_id = %s'
        conf_val = ('confirm', call.message.chat.id)
        db.commit()
        sql2 = 'SELECT product, amount, price FROM cart WHERE cart_id = %s'
        val2 = (call.message.chat.id,)
        cursor.execute(sql2, val2)
        fetch = cursor.fetchall()
        select = 'SELECT price FROM cart WHERE cart_id = %s'
        select2 = (call.message.chat.id,)
        cursor.execute(select, select2)
        fetch_price = cursor.fetchall()
        sel = 'SELECT message FROM calldata WHERE user_id = %s'
        selval = (call.message.chat.id,)
        cursor.execute(sel, selval)
        message_fetch = cursor.fetchone()
        phone_check = 'SELECT phone_number FROM users WHERE user_id = %s'
        phone_check2 = (call.message.chat.id,)
        cursor.execute(phone_check, phone_check2)
        phone_number = cursor.fetchone()
        sql = 'UPDATE users SET min = %s WHERE user_id = %s'
        val = (0, call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT hour, min FROM users WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        time_fetch = cursor.fetchone()
        sql = 'UPDATE orders SET time = %s WHERE user_id = %s AND process = %s'
        val = (str(time_fetch[0]) + ':00', call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        if phone_number[0] == None:
            bot.send_message(
                text='Чтобы продолжить вы должны отправить нам ваши контактные данные для дальнейших услуг',
                chat_id=call.message.chat.id, reply_markup=contact)
            bot.edit_message_text(
                'Способ доставки - курьер' + '\nСпособ доставки - каспи голд' + '\nУказанное время - ' + '\n...', call.message.chat.id,
                message_fetch[0])
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            sql = 'UPDATE users SET step = %s WHERE user_id = %s'
            val = ('kaspi', call.message.chat.id)
            cursor.execute(sql, val)
            db.commit()
        else:
            price_list = []
            orders = ''
            for i in range(len(fetch_price)):
                sc = fetch_price[i]
                price_list.append(sc[0])
                orders += '\n' + str(i + 1) + '. ' + (fetch[i])[0] + ' ' + str((fetch[i])[1]) + ' шт. - ' + str(
                    (fetch[i])[2]) + 'тг.'
            bot.delete_message(call.message.chat.id, message_fetch[0])
            confirmation = types.InlineKeyboardMarkup()
            confirm = types.InlineKeyboardButton('Подтвердить\u2705', callback_data='confirm_k')
            uhhhnogetback = types.InlineKeyboardButton('Отмена\u274c', callback_data='bacc')
            confirmation.add(confirm, uhhhnogetback)
            remove = types.ReplyKeyboardRemove()
            bot.send_message(call.message.chat.id, 'Все верно?', reply_markup=remove)
            bot.send_message(parse_mode='Markdown',
                             text='*Заказ №' + str(
                                 num + 1) + '*\n' + '\nСпособ доставки - курьер' + '\nСпособ доставки - каспи голд' + '\nИмя - ' + call.message.chat.first_name + '\nНомер телефона - ' + str(
                                 phone_number[0]) + '\n*Ваш заказ:*' + orders + '\n*Общая стоимость - *' + '*' + str(
                                 sum(price_list)) + 'тг.*\nПеревод на каспи голд производится после подтверждения заказа', chat_id=call.message.chat.id,
                             reply_markup=confirmation)
    if call.data == 'confirm_k':
        sql = 'SELECT price FROM cart WHERE cart_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        fetch_price = cursor.fetchall()
        price_list = []
        for i in range(len(fetch_price)):  # Подсчет конечной суммы и оформление заказа
            sc = fetch_price[i]
            price_list.append(sc[0])
        sql = 'UPDATE orders SET price = %s WHERE user_id = %s AND process = %s'
        val = (sum(price_list), call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT address FROM users WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        address = cursor.fetchone()
        sql = 'SELECT phone_number FROM users WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        phone = cursor.fetchone()
        sql = 'UPDATE orders SET address = %s WHERE user_id = %s AND process = %s'
        val = (address[0], call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        sql = 'UPDATE orders SET phone_number = %s WHERE user_id = %s AND process = %s'
        val = (phone[0], call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        sql = 'UPDATE orders SET order_date = %s WHERE user_id = %s AND process = %s'
        val = (
        str(datetime.datetime.today().date()) + ', ' + str(datetime.datetime.today().time()), call.message.chat.id,
        'ordering')
        cursor.execute(sql, val)
        db.commit()
        sql = 'UPDATE orders SET full_desc = %s WHERE user_id = %s AND process = %s'
        val = (call.message.text, call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        sql = 'UPDATE orders SET process = %s WHERE user_id = %s AND process = %s'
        val = ('finished', call.message.chat.id, 'ordering')
        cursor.execute(sql, val)
        db.commit()
        sql = 'UPDATE users SET step = %s WHERE user_id = %s'
        val = ('kaspi_ss', call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        sql = 'DELETE FROM cart WHERE cart_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        db.commit()
        bot.forward_message(dad_id, call.message.chat.id, call.message.message_id)
        bot.send_message(dad_id, 'Ожидайте скриншот')
        bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
        main_page = types.InlineKeyboardMarkup()
        main_pagebutton = types.InlineKeyboardButton('Главная\ud83c\udfe0', callback_data='home')
        main_page.add(main_pagebutton)
        bot.send_message(call.message.chat.id, 'Отправьте сумму на номер <номер телефона> или на карту <карта> в размере ' + str(sum(price_list)) + 'тг.\nПосле перевода прикрепите скриншот.')
    if call.data == 'reports':
        reports = types.InlineKeyboardMarkup(1)
        users = types.InlineKeyboardButton('Отчеты по пользователям', callback_data='users')
        order_rep = types.InlineKeyboardButton('Отчеты по заказам', callback_data='order_rep')
        back = types.InlineKeyboardButton('Назад \u2b05', callback_data='backadmin')
        reports.add(users, order_rep, back)
        bot.edit_message_text('Раздел - отчеты', call.message.chat.id, call.message.message_id, reply_markup=reports)
    if call.data == 'users':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Users'
        sql = 'SELECT first_name, last_name, user_id, phone_number, address, role FROM users'
        cursor.execute(sql)
        users_report = cursor.fetchall()
        ws['A1'] = 'Имя'
        ws['B1'] = 'Фамилия'
        ws['C1'] = 'ID'
        ws['D1'] = 'Номер тел.'
        ws['E1'] = 'Адрес'
        ws['F1'] = 'Роль'
        ws['G1'] = 'Сделано заказов'
        for i in range(2, cursor.rowcount + 2):
            sql = 'SELECT * FROM orders WHERE user_id = %s AND process = %s'
            val = (users_report[i - 2][2], 'finished')
            cursor.execute(sql, val)
            order_count = cursor.rowcount
            ws.cell(i, 1, users_report[i - 2][0])
            ws.cell(i, 2, users_report[i - 2][1])
            ws.cell(i, 3, users_report[i - 2][2])
            ws.cell(i, 4, users_report[i - 2][3])
            ws.cell(i, 5, users_report[i - 2][4])
            ws.cell(i, 6, users_report[i - 2][5])
            ws.cell(i, 7, order_count)
        wb.save('users' + str(datetime.datetime.today().date()) + '.xlsx')
        f = open('users' + str(datetime.datetime.today().date()) + '.xlsx', 'rb')
        bot.delete_message(call.message.chat.id, call.message.message_id)
        bot.send_document(call.message.chat.id, f)
        reports = types.InlineKeyboardMarkup(1)
        users = types.InlineKeyboardButton('Отчеты по пользователям', callback_data='users')
        order_rep = types.InlineKeyboardButton('Отчеты по заказам', callback_data='order_rep')
        back = types.InlineKeyboardButton('Назад \u2b05', callback_data='backadmin')
        reports.add(users, order_rep, back)
        bot.send_message(text='Раздел - отчеты', chat_id=call.message.chat.id, reply_markup=reports)


    if call.data == 'order_rep':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Orders'
        sql = 'SELECT №, first_name, last_name, phone_number, address, delivery_method, pay_method, moneychange, price, time, order_date FROM orders WHERE process = %s'
        val = ('finished', )
        cursor.execute(sql, val)
        users_report = cursor.fetchall()
        ws['A1'] = '№'
        ws['B1'] = 'Имя'
        ws['C1'] = 'Фамилия'
        ws['D1'] = 'Номер тел.'
        ws['E1'] = 'Адрес доставки'
        ws['F1'] = 'Способ доставки'
        ws['G1'] = 'Способ оплаты'
        ws['H1'] = 'Сумма оплаты'
        ws['I1'] = 'Стоимость заказа'
        ws['J1'] = 'Указанное время'
        ws['K1'] = 'Дата и время заказа'
        for i in range(2, cursor.rowcount + 2):
            ws.cell(i, 1, users_report[i - 2][0])
            ws.cell(i, 2, users_report[i - 2][1])
            ws.cell(i, 3, users_report[i - 2][2])
            ws.cell(i, 4, users_report[i - 2][3])
            ws.cell(i, 5, users_report[i - 2][4])
            ws.cell(i, 6, users_report[i - 2][5])
            ws.cell(i, 7, users_report[i - 2][6])
            ws.cell(i, 8, users_report[i - 2][7])
            ws.cell(i, 9, users_report[i - 2][8])
            ws.cell(i, 10, users_report[i - 2][9])
            ws.cell(i, 11, users_report[i - 2][10])
        wb.save('orders' + str(datetime.datetime.today().date()) + '.xlsx')
        f = open('orders' + str(datetime.datetime.today().date()) + '.xlsx', 'rb')
        bot.send_document(call.message.chat.id, f)
        bot.delete_message(call.message.chat.id, call.message.message_id)
        reports = types.InlineKeyboardMarkup(1)
        users = types.InlineKeyboardButton('Отчеты по пользователям', callback_data='users')
        order_rep = types.InlineKeyboardButton('Отчеты по заказам', callback_data='order_rep')
        back = types.InlineKeyboardButton('Назад \u2b05', callback_data='backadmin')
        reports.add(users, order_rep, back)
        bot.send_message(text='Раздел - отчеты', chat_id=call.message.chat.id,
                         reply_markup=reports)
    if call.data == 'edit':
        prod_cate = types.InlineKeyboardMarkup(1)
        prods = types.InlineKeyboardButton('Продукты', callback_data='prod')
        cates = types.InlineKeyboardButton('Категории', callback_data='cate')
        back = types.InlineKeyboardButton('Назад \u2b05', callback_data='backadmin')
        prod_cate.add(cates, prods, back)
        bot.edit_message_text('Раздел - меню', call.message.chat.id, call.message.message_id, reply_markup=prod_cate)
    if call.data == 'backadmin':
        admin_start = types.InlineKeyboardMarkup(row_width=1)
        reports = types.InlineKeyboardButton('Отчеты', callback_data='reports')
        categories = types.InlineKeyboardButton('Меню', callback_data='edit')
        admin_start.add(reports, categories)
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text='Админ панель', reply_markup=admin_start)
    if call.data == 'backadnu':
        prod_cate = types.InlineKeyboardMarkup(1)
        prods = types.InlineKeyboardButton('Продукты', callback_data='prod')
        cates = types.InlineKeyboardButton('Категории', callback_data='cate')
        back = types.InlineKeyboardButton('Назад \u2b05', callback_data='backadmin')
        prod_cate.add(cates, prods, back)
        bot.edit_message_text('Раздел - меню', call.message.chat.id, call.message.message_id, reply_markup=prod_cate)
    if call.data == 'prod':
        markup_action = types.InlineKeyboardMarkup(1)
        add = types.InlineKeyboardButton('Добавить продукт\u2795', callback_data='add_prod')
        among_cates = types.InlineKeyboardButton('Редактировать продукт', callback_data='among_cates')
        back = types.InlineKeyboardButton('Назад \u2b05', callback_data='backadnu')
        markup_action.add(add, among_cates, back)
        bot.edit_message_text('<b>Продукты</b>\nВыберите действие:', call.message.chat.id, call.message.message_id, reply_markup=markup_action, parse_mode='HTML')
    if call.data == 'add_prod':
        sql = 'SELECT product, desc_, price, type, seen, link, disc FROM productlist WHERE user_id = %s AND process = %s'
        val = (call.message.chat.id, 'making')
        cursor.execute(sql, val)
        product = cursor.fetchone()
        if cursor.rowcount > 0:
            if product[0] == None:
                sql = 'UPDATE productlist SET desc_ = %s WHERE user_id = %s AND process = %s'
                val = ('не указано', call.message.chat.id, 'making')
                cursor.execute(sql, val)
                db.commit()
            if product[1] == None:
                sql = 'UPDATE productlist SET price = %s WHERE user_id = %s AND process = %s'
                val = ('не указано', call.message.chat.id, 'making')
                cursor.execute(sql, val)
                db.commit()
            if product[2] == None:
                sql = 'UPDATE productlist SET type = %s WHERE user_id = %s AND process = %s'
                val = ('не указано', call.message.chat.id, 'making')
                cursor.execute(sql, val)
                db.commit()
            if product[5] == None:
                sql = 'UPDATE productlist SET link = %s WHERE user_id = %s AND process = %s'
                val = ('не указано', call.message.chat.id, 'making')
                cursor.execute(sql, val)
                db.commit()
            unsaved = types.InlineKeyboardMarkup()
            continue_work = types.InlineKeyboardButton('Продолжить', callback_data='continue_work')
            nope = types.InlineKeyboardButton('Нет, удалить данные', callback_data='naw')
            unsaved.add(continue_work, nope)
            bot.edit_message_text(
                    '<b>Внимание! У вас есть несохраненные данные!</b>\n\nИмя: ' + product[0] + '\nОписание: ' + product[
                        1] + '\nЦена: ' + str(product[2]) + '\nКатегория: ' + product[3] + '\nВидимость: ' + product[
                        4] + '\nАкция: ' + product[6] + '\nФото: ' + product[5] + '\nПродолжить работу?', call.message.chat.id, call.message.message_id,
                    parse_mode='HTML', reply_markup=unsaved)
        else:
            sql = 'UPDATE users SET step = %s WHERE user_id = %s'
            val = ('naming', call.message.chat.id)
            cursor.execute(sql, val)
            db.commit()
            markup_cancel = types.InlineKeyboardMarkup()
            cancel = types.InlineKeyboardButton('Отмена\u274c', callback_data='cancel_name')
            markup_cancel.add(cancel)
            bot.edit_message_text('Введите имя для нового продукта:', call.message.chat.id, call.message.message_id, reply_markup=markup_cancel)
    if call.data == 'cancel_name':
        markup_action = types.InlineKeyboardMarkup(1)
        add = types.InlineKeyboardButton('Добавить продукт\u2795', callback_data='add_prod')
        among_cates = types.InlineKeyboardButton('Редактировать категорию', callback_data='among_cates')
        back = types.InlineKeyboardButton('Назад \u2b05', callback_data='backadnu')
        markup_action.add(add, among_cates, back)
        bot.edit_message_text('<b>Продукты</b>\nВыберите действие:', call.message.chat.id, call.message.message_id,
                              parse_mode='HTML', reply_markup=markup_action)
    if call.data == 'edit_name':
        sql = 'UPDATE users SET step = %s WHERE user_id = %s'
        val = ('edit_name', call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        markup_cancel = types.InlineKeyboardMarkup()
        cancel = types.InlineKeyboardButton('Отмена\u274c', callback_data='cancel_all')
        markup_cancel.add(cancel)
        bot.edit_message_text('Введите новое имя для продукта:', call.message.chat.id, call.message.message_id, reply_markup=markup_cancel)
    if call.data == 'cancel_all':
        sql = 'SELECT product, desc_, price, type, seen, link, disc FROM productlist WHERE user_id = %s AND process = %s'
        val = (call.message.chat.id, 'making')
        cursor.execute(sql, val)
        product = cursor.fetchone()
        adding_prod = types.InlineKeyboardMarkup(2)
        preview = types.InlineKeyboardButton('Предпросмотр', callback_data='preview_prod')
        edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='edit_name')
        edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='edit_photo')
        edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='edit_desc')
        edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='edit_price')
        edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='edit_cate')
        show = types.InlineKeyboardButton('Показать\u2705', callback_data='show' + product[0])
        hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='hide' + product[0])
        ready = types.InlineKeyboardButton('Сохранить\ud83d\udcbe', callback_data='ready_prod')
        delete = types.InlineKeyboardButton('Удалить продукт\u274c', callback_data='delete_prod')
        disc_no = types.InlineKeyboardButton('Акция \u2611\ufe0f', callback_data='disc_no')
        disc = types.InlineKeyboardButton('Акция \u25fb\ufe0f', callback_data='disc')
        adding_prod.add(ready, delete)
        if product[4] == '<b>ДА</b>':
            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, hide)
        else:
            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show)
        if product[6] == '<b>ДА</b>':
            adding_prod.add(disc_no)
        else:
            adding_prod.add(disc)
        adding_prod.add(preview)
        bot.edit_message_text(chat_id=call.message.chat.id,
                                  text='<b>Информация о продукте</b>\n\nИмя: ' + product[
                                      0] + '\nОписание: \n' + product[
                                           1] + '\nЦена: ' + str(product[2]) + '\nКатегория: ' + product[
                                           3] + '\nВидимость: ' +
                                       product[4] + '\nФото: ' + product[5], parse_mode='HTML', message_id=call.message.message_id,
                                  reply_markup=adding_prod)
    if call.data == 'continue_work':
            sql = 'SELECT product, desc_, price, type, seen, link, disc FROM productlist WHERE user_id = %s AND process = %s'
            val = (call.message.chat.id, 'making')
            cursor.execute(sql, val)
            product = cursor.fetchone()
            adding_prod = types.InlineKeyboardMarkup(2)
            preview = types.InlineKeyboardButton('Предпросмотр', callback_data='preview_prod')
            edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='edit_name')
            edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='edit_photo')
            edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='edit_desc')
            edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='edit_price')
            edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='edit_cate')
            show = types.InlineKeyboardButton('Показать\u2705', callback_data='show' + product[0])
            hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='hide' + product[0])
            ready = types.InlineKeyboardButton('Сохранить\ud83d\udcbe', callback_data='ready_prod')
            delete = types.InlineKeyboardButton('Удалить продукт\u274c', callback_data='delete_prod')
            adding_prod.add(ready, delete)
            disc_no = types.InlineKeyboardButton('Акция \u2611\ufe0f', callback_data='disc_no')
            disc = types.InlineKeyboardButton('Акция \u25fb\ufe0f', callback_data='disc')
            if product[4] == '<b>ДА</b>':
                adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, hide)
            else:
                adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show)
            if product[6] == '<b>ДА</b>':
                adding_prod.add(disc_no)
            else:
                adding_prod.add(disc)
            adding_prod.add(preview)
            bot.edit_message_text(chat_id=call.message.chat.id,
                                      text='<b>Информация о продукте</b>\n\nИмя: ' + product[
                                          0] + '\nОписание: \n' + product[
                                               1] + '\nЦена: ' + str(product[2]) + '\nКатегория: ' + product[
                                               3] + '\nВидимость: ' +
                                           product[4] + '\nФото: ' + product[5], parse_mode='HTML',
                                      message_id=call.message.message_id,
                                      reply_markup=adding_prod)
    if call.data == 'naw':
        sql = 'DELETE FROM productlist WHERE user_id = %s AND process = %s'
        val = (call.message.chat.id, 'making')
        cursor.execute(sql, val)
        sql = 'UPDATE users SET step = %s WHERE user_id = %s'
        val = ('naming', call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        markup_cancel = types.InlineKeyboardMarkup()
        cancel = types.InlineKeyboardButton('Отмена\u274c', callback_data='cancel_name')
        markup_cancel.add(cancel)
        bot.edit_message_text('Введите имя для нового продукта:', call.message.chat.id, call.message.message_id,
                              reply_markup=markup_cancel)
    if call.data == 'edit_desc':
        sql = 'UPDATE users SET step = %s WHERE user_id = %s'
        val = ('edit_desc', call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        markup_cancel = types.InlineKeyboardMarkup()
        cancel = types.InlineKeyboardButton('Отмена\u274c', callback_data='cancel_all')
        markup_cancel.add(cancel)
        bot.edit_message_text('Введите новое описание для продукта:', call.message.chat.id, call.message.message_id, reply_markup=markup_cancel)
    if call.data == 'edit_price':
        sql = 'UPDATE users SET step = %s WHERE user_id = %s'
        val = ('edit_price', call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        markup_cancel = types.InlineKeyboardMarkup()
        cancel = types.InlineKeyboardButton('Отмена\u274c', callback_data='cancel_all')
        markup_cancel.add(cancel)
        bot.edit_message_text('Введите цену для продукта:', call.message.chat.id, call.message.message_id,
                              reply_markup=markup_cancel)
    if call.data == 'edit_cate':
        food_type = types.InlineKeyboardMarkup(2)
        sql = 'SELECT name FROM categorylist'
        cursor.execute(sql)
        category = cursor.fetchall()
        for i in range(len(category)):
            button = types.InlineKeyboardButton(category[i][0], callback_data='catess' + str(i))
            food_type.add(button)
        cancel = types.InlineKeyboardButton('Отмена\u274c', callback_data='cancel_all')
        food_type.add(cancel)
        bot.edit_message_text(text='Выберите категорию в которой будет находиться продукт', chat_id=call.message.chat.id,
                              message_id=call.message.message_id, reply_markup=food_type)
    if call.data.startswith('catess'):
        sql = 'SELECT name FROM categorylist'
        cursor.execute(sql)
        category = cursor.fetchall()
        sql = 'UPDATE productlist SET type = %s WHERE user_id = %s AND process = %s'
        val = (category[int(call.data[6:])][0], call.message.chat.id, 'making')
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT product, desc_, price, type, seen, link, disc FROM productlist WHERE user_id = %s AND process = %s'
        val = (call.message.chat.id, 'making')
        cursor.execute(sql, val)
        product = cursor.fetchone()
        adding_prod = types.InlineKeyboardMarkup(2)
        preview = types.InlineKeyboardButton('Предпросмотр', callback_data='preview_prod')
        edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='edit_name')
        edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='edit_photo')
        edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='edit_desc')
        edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='edit_price')
        edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='edit_cate')
        show = types.InlineKeyboardButton('Показать\u2705', callback_data='show' + product[0])
        hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='hide' + product[0])
        ready = types.InlineKeyboardButton('Сохранить\ud83d\udcbe', callback_data='ready_prod')
        delete = types.InlineKeyboardButton('Удалить продукт\u274c', callback_data='delete_prod')
        disc_no = types.InlineKeyboardButton('Акция \u2611\ufe0f', callback_data='disc_no')
        disc = types.InlineKeyboardButton('Акция \u25fb\ufe0f', callback_data='disc')
        adding_prod.add(ready, delete)
        if product[4] == '<b>ДА</b>':
            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, hide)
        else:
            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show)
        if product[6] == '<b>ДА</b>':
            adding_prod.add(disc_no)
        else:
            adding_prod.add(disc)
        adding_prod.add(preview)
        bot.edit_message_text(chat_id=call.message.chat.id,
                                text='Категория продукта отредактирована!\n<b>Информация о продукте</b>\n\nИмя: ' + product[
                                    0] + '\nОписание: \n' +
                                     product[
                                         1] + '\nЦена: ' + str(product[2]) + '\nКатегория: ' + product[
                                         3] + '\nВидимость: ' +product[4] + '\nФото: ' + product[5], parse_mode='HTML', message_id=call.message.message_id, reply_markup=adding_prod)
    if call.data == 'edit_photo':
        sql = 'UPDATE users SET step = %s WHERE user_id = %s'
        val = ('edit_photo', call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        markup_cancel = types.InlineKeyboardMarkup()
        cancel = types.InlineKeyboardButton('Отмена\u274c', callback_data='cancel_all')
        markup_cancel.add(cancel)
        bot.edit_message_text('Введите ссылку, ведущую к фотографии продукта:', call.message.chat.id, call.message.message_id,
                              reply_markup=markup_cancel)
    if call.data.startswith('show'):
        sql = 'UPDATE productlist SET seen = %s WHERE user_id = %s AND process = %s'
        val = ('<b>ДА</b>', call.message.chat.id, 'making')
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT product, desc_, price, type, seen, link, disc FROM productlist WHERE user_id = %s AND process = %s'
        val = (call.message.chat.id, 'making')
        cursor.execute(sql, val)
        product = cursor.fetchone()
        adding_prod = types.InlineKeyboardMarkup(2)
        preview = types.InlineKeyboardButton('Предпросмотр', callback_data='preview_prod')
        edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='edit_name')
        edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='edit_photo')
        edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='edit_desc')
        edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='edit_price')
        edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='edit_cate')
        show = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='hide' + product[0])
        ready = types.InlineKeyboardButton('Сохранить\ud83d\udcbe', callback_data='ready_prod')
        delete = types.InlineKeyboardButton('Удалить продукт\u274c', callback_data='delete_prod')
        disc_no = types.InlineKeyboardButton('Акция \u2611\ufe0f', callback_data='disc_no')
        disc = types.InlineKeyboardButton('Акция \u25fb\ufe0f', callback_data='disc')
        adding_prod.add(ready, delete)
        adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show)
        if product[6] == '<b>ДА</b>':
            adding_prod.add(disc_no)
        else:
            adding_prod.add(disc)
        adding_prod.add(preview)
        bot.edit_message_text(chat_id=call.message.chat.id,
                                  text='<b>Информация о продукте</b>\n\nИмя: ' + product[
                                      0] + '\nОписание: \n' + product[
                                           1] + '\nЦена: ' + str(product[2]) + '\nКатегория: ' + product[
                                           3] + '\nВидимость: ' +
                                       product[4] + '\nФото: ' + product[5], parse_mode='HTML',
                                  message_id=call.message.message_id,
                                  reply_markup=adding_prod)
    if call.data.startswith('hide'):
        sql = 'UPDATE productlist SET seen = %s WHERE user_id = %s AND process = %s'
        val = ('<b>НЕТ</b>', call.message.chat.id, 'making')
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT product, desc_, price, type, seen, link, disc FROM productlist WHERE user_id = %s AND process = %s'
        val = (call.message.chat.id, 'making')
        cursor.execute(sql, val)
        product = cursor.fetchone()
        adding_prod = types.InlineKeyboardMarkup(2)
        preview = types.InlineKeyboardButton('Предпросмотр', callback_data='preview_prod')
        edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='edit_name')
        edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='edit_photo')
        edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='edit_desc')
        edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='edit_price')
        edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='edit_cate')
        show = types.InlineKeyboardButton('Показать\u2705', callback_data='show' + product[0])
        ready = types.InlineKeyboardButton('Сохранить\ud83d\udcbe', callback_data='ready_prod')
        delete = types.InlineKeyboardButton('Удалить продукт\u274c', callback_data='delete_prod')
        disc_no = types.InlineKeyboardButton('Акция \u2611\ufe0f', callback_data='disc_no')
        disc = types.InlineKeyboardButton('Акция \u25fb\ufe0f', callback_data='disc')
        adding_prod.add(ready, delete)
        adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show)
        if product[6] == '<b>ДА</b>':
            adding_prod.add(disc_no)
        else:
            adding_prod.add(disc)
        adding_prod.add(preview)
        bot.edit_message_text(chat_id=call.message.chat.id,
                                  text='<b>Информация о продукте</b>\n\nИмя: ' + product[
                                      0] + '\nОписание: \n' + product[
                                           1] + '\nЦена: ' + str(product[2]) + '\nКатегория: ' + product[
                                           3] + '\nВидимость: ' +
                                       product[4] + '\nФото: ' + product[5], parse_mode='HTML',
                                  message_id=call.message.message_id,
                                  reply_markup=adding_prod)
    if call.data == 'delete_prod':
        delete = types.InlineKeyboardMarkup()
        yes = types.InlineKeyboardButton('Да\u2705', callback_data='yes_delete_prod')
        nope_don = types.InlineKeyboardButton('Нет\u274c', callback_data='no_dont_prod')
        delete.add(yes, nope_don)
        bot.edit_message_text('Вы уверены, что хотите удалить данный продукт?', call.message.chat.id, call.message.message_id, reply_markup=delete)
    if call.data == 'yes_delete_prod':
        sql = 'DELETE FROM productlist WHERE user_id = %s AND process = %s'
        val = (call.message.chat.id, 'making')
        cursor.execute(sql, val)
        db.commit()
        markup_action = types.InlineKeyboardMarkup(1)
        add = types.InlineKeyboardButton('Добавить продукт\u2795', callback_data='add_prod')
        among_cates = types.InlineKeyboardButton('Редактировать продукт', callback_data='among_cates')
        back = types.InlineKeyboardButton('Назад \u2b05', callback_data='backadnu')
        markup_action.add(add, among_cates, back)
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text='Продукт удален!\n<b>Продукты</b>\nВыберите действие:',
                              reply_markup=markup_action, parse_mode='HTML')
    if call.data == 'no_dont_prod':
        sql = 'SELECT product, desc_, price, type, seen, link, disc FROM productlist WHERE user_id = %s AND process = %s'
        val = (call.message.chat.id, 'making')
        cursor.execute(sql, val)
        product = cursor.fetchone()
        adding_prod = types.InlineKeyboardMarkup(2)
        preview = types.InlineKeyboardButton('Предпросмотр', callback_data='preview_prod')
        edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='edit_name')
        edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='edit_photo')
        edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='edit_desc')
        edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='edit_price')
        edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='edit_cate')
        show = types.InlineKeyboardButton('Показать\u2705', callback_data='show' + product[0])
        hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='hide' + product[0])
        ready = types.InlineKeyboardButton('Сохранить\ud83d\udcbe', callback_data='ready_prod')
        delete = types.InlineKeyboardButton('Удалить продукт\u274c', callback_data='delete_prod')
        disc_no = types.InlineKeyboardButton('Акция \u2611\ufe0f', callback_data='disc_no')
        disc = types.InlineKeyboardButton('Акция \u25fb\ufe0f', callback_data='disc')
        adding_prod.add(ready, delete)
        if product[4] == '<b>ДА</b>':
            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, hide)
        else:
            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show)
        if product[6] == '<b>ДА</b>':
            adding_prod.add(disc_no)
        else:
            adding_prod.add(disc)
        adding_prod.add(preview)
        bot.edit_message_text(chat_id=call.message.chat.id,
                                  text='<b>Информация о продукте</b>\n\nИмя: ' + product[
                                      0] + '\nОписание: \n' + product[
                                           1] + '\nЦена: ' + str(product[2]) + '\nКатегория: ' + product[
                                           3] + '\nВидимость: ' +
                                       product[4] + '\nФото: ' + product[5], parse_mode='HTML',
                                  message_id=call.message.message_id,
                                  reply_markup=adding_prod)
    if call.data == 'ready_prod':
        sql = 'SELECT type FROM productlist WHERE user_id = %s AND process = %s'
        val = (call.message.chat.id, 'making')
        cursor.execute(sql, val)
        product = cursor.fetchone()
        if product[0] == 'не указано':
            bot.answer_callback_query(call.id, 'Вы не указали категорию продукта!', True)
            return
        sql = 'UPDATE productlist SET process = %s WHERE user_id = %s AND process = %s'
        val = ('saved', call.message.chat.id, 'making')
        cursor.execute(sql, val)
        db.commit()
        markup_action = types.InlineKeyboardMarkup(1)
        add = types.InlineKeyboardButton('Добавить\u2795', callback_data='add_prod')
        among_cates = types.InlineKeyboardButton('Поиск среди категорий', callback_data='among_cates')
        back = types.InlineKeyboardButton('Назад \u2b05', callback_data='backadnu')
        markup_action.add(add, among_cates, back)
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text='Продукт успешно сохранен!\n<b>Продукты</b>\nВыберите действие:',
                         reply_markup=markup_action, parse_mode='HTML')

    if call.data == 'preview_prod':
        sql = 'SELECT product, desc_, price, type, seen, link FROM productlist WHERE user_id = %s AND process = %s'
        val = (call.message.chat.id, 'making')
        cursor.execute(sql, val)
        product = cursor.fetchone()
        markup = types.InlineKeyboardMarkup()
        back = types.InlineKeyboardButton('Назад \u2b05', callback_data='cancel_all')
        markup.add(back)
        if product[5] != 'не указано':
            bot.edit_message_text('<a href="' + product[5] + '"> </a><b>' + product[0] + '</b> - ' + str(product[2]) + 'тг.\n\n' + product[1], call.message.chat.id, call.message.message_id, parse_mode='HTML', reply_markup=markup)
        else:
            bot.edit_message_text(product[0] + ' - ' + str(product[2]) + 'тг.\n\n' + product[1] + '\nОТСУТСТВУЕТ ФОТОГРАФИЯ', call.message.chat.id, call.message.message_id, parse_mode='HTML', reply_markup=markup)

    if call.data == 'among_cates':
        food_type = types.InlineKeyboardMarkup(2)
        sql = 'SELECT name FROM categorylist'
        cursor.execute(sql)
        category = cursor.fetchall()
        if cursor.rowcount > 0:
            for i in range(len(category)):
                button = types.InlineKeyboardButton(category[i][0], callback_data='_choose' + str(i))
                food_type.add(button)
            cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='prod')
            food_type.add(cancel)
            bot.edit_message_text('Выберите категорию:', call.message.chat.id, call.message.message_id, reply_markup=food_type)
        else:
            mk = types.InlineKeyboardMarkup()
            cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='prod')
            mk.add(cancel)
            bot.edit_message_text('Категорий не найдено. Попробуйте создать парочку в разделе <b>Категории</b> нажав на енопку <b>"Добавить категорию"</b>!', call.message.chat.id, call.message.message_id, parse_mode='HTML', reply_markup=mk)
    if call.data.startswith('_choose'):
        sql = 'UPDATE calldata SET start_msg = %s WHERE user_id = %s'
        val = (call.message.message_id, call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        sql = 'DELETE FROM messages WHERE user_id = %s'
        val = (call.message.chat.id, )
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT name FROM categorylist'
        cursor.execute(sql)
        category = cursor.fetchall()
        ind = int(call.data[7:])
        sql = 'SELECT product, price, desc_, link, seen FROM productlist WHERE type = %s AND process NOT IN (%s, %s)'
        val = (category[ind][0], 'making', 'editing')
        cursor.execute(sql, val)
        product = cursor.fetchall()
        if cursor.rowcount == 0:
            food_type = types.InlineKeyboardMarkup(2)
            cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
            food_type.add(cancel)
            bot.edit_message_text('В этой категории ничего нет.', call.message.chat.id, call.message.message_id, reply_markup=food_type)
        else:
            sql = 'UPDATE calldata SET message = %s WHERE user_id = %s'
            val = (None, call.message.chat.id)
            cursor.execute(sql, val)
            db.commit()
            sql = 'UPDATE calldata SET calldata = %s WHERE user_id = %s'
            val = (None, call.message.chat.id)
            cursor.execute(sql, val)
            db.commit()
            sql = 'UPDATE calldata SET calldata2 = %s WHERE user_id = %s'
            val = (ind, call.message.chat.id)
            cursor.execute(sql, val)
            db.commit()
            cursor.execute('DELETE FROM messages WHERE user_id = %s', (call.message.chat.id, ))
            db.commit()
            bot.edit_message_text(text='<b>Категория - ' + category[ind][0] + '</b>', chat_id=call.message.chat.id,
                                  message_id=call.message.message_id, parse_mode='HTML')
            for i in range(len(product)):
                food_type = types.InlineKeyboardMarkup(2)
                button = types.InlineKeyboardButton('Редактировать\ud83d\udd27', callback_data='edit_prod' + str(i))
                food_type.add(button)
                sql = 'INSERT INTO messages (user_id, message) VALUES (%s, %s)'
                val = (call.message.chat.id, call.message.message_id + 1 + i)
                cursor.execute(sql, val)
                db.commit()
                if i == len(product) - 1:
                    cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
                    food_type.add(cancel)
                if product[i][3] != 'не указано':
                    bot.send_message(
                        text='<a href="' + product[i][3] + '"> </a>' + product[i][0] + ' - ' + str(product[i][1]) + 'тг.\n\n' + product[i][2] + '\nВидно: ' + product[i][4],
                        chat_id=call.message.chat.id, parse_mode='HTML', reply_markup=food_type)
                else:
                    bot.send_message(
                        text=product[i][0] + ' - ' + str(product[i][1]) + 'тг.\n\n' + product[i][2] + '\nОТСУТСТВУЕТ ФОТОГРАФИЯ\nВидно: ' + product[i][4],
                        chat_id=call.message.chat.id, parse_mode='HTML', reply_markup=food_type)

    if call.data.startswith('edit_prod'):
        sql = 'SELECT message FROM messages WHERE user_id = %s'
        val = (call.message.chat.id, )
        cursor.execute(sql, val)
        messages = cursor.fetchall()
        call_check1 = 'SELECT calldata FROM calldata WHERE user_id = %s'
        call_check2 = (call.message.chat.id,)
        cursor.execute(call_check1, call_check2)
        call_fetch = cursor.fetchone()
        calldata = call_fetch[0]
        calldata1 = 'UPDATE calldata SET calldata = %s WHERE user_id = %s'
        calldata2 = (int((call.data[9:])), call.message.chat.id)
        cursor.execute(calldata1, calldata2)
        db.commit()
        call_check1 = 'SELECT calldata FROM calldata WHERE user_id = %s'
        call_check2 = (call.message.chat.id,)
        cursor.execute(call_check1, call_check2)
        call_fetch = cursor.fetchone()
        calldataa = call_fetch[0]
        call_check1 = 'SELECT calldata2 FROM calldata WHERE user_id = %s'
        call_check2 = (call.message.chat.id,)
        cursor.execute(call_check1, call_check2)
        call_fetch = cursor.fetchone()
        category_call = call_fetch[0]
        sql = 'SELECT name FROM categorylist'
        cursor.execute(sql)
        category = cursor.fetchall()
        sql = 'SELECT product FROM productlist WHERE type = %s'
        val = (category[category_call][0], )
        cursor.execute(sql, val)
        cate_prod = cursor.fetchall()
        sql = 'SELECT message FROM calldata WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        message = cursor.fetchone()
        sql = 'UPDATE calldata SET message = %s WHERE user_id = %s'
        val = (call.message.message_id, call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        print(calldata, calldataa)
        sql = 'UPDATE productlist SET process = %s, user_id = %s WHERE product = %s'
        val = ('editing', call.message.chat.id, cate_prod[calldataa][0])
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT product, desc_, price, type, seen, link FROM productlist WHERE product = %s'
        val = (cate_prod[calldataa][0],)
        cursor.execute(sql, val)
        product = cursor.fetchone()
        adding_prod = types.InlineKeyboardMarkup(2)
        preview = types.InlineKeyboardButton('Предпросмотр', callback_data='epreview_prod' + product[0])
        edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='eedit_name' + product[0])
        edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='eedit_photo' + product[0])
        edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='eedit_desc' + product[0])
        edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='eedit_price' + product[0])
        edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='eedit_cate' + product[0])
        show = types.InlineKeyboardButton('Показать\u2705', callback_data='eshow' + product[0])
        hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='ehide' + product[0])
        delete = types.InlineKeyboardButton('Удалить продукт\u274c', callback_data='edelete_prod' + product[0])
        adding_prod.add(delete, preview)
        if product[4] == '<b>ДА</b>':
            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, hide)
        else:
            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show)
        if message[0] != None:
            print(calldata)
            sql = 'SELECT product, desc_, price, type, seen, link FROM productlist WHERE product = %s'
            val = (cate_prod[calldata][0],)
            cursor.execute(sql, val)
            productt = cursor.fetchone()
            sql = 'UPDATE productlist SET process = %s WHERE process = %s AND product = %s AND user_id = %s'
            val = ('saved', 'editing', productt[0], call.message.chat.id)
            cursor.execute(sql, val)
            db.commit()
            order_markup = types.InlineKeyboardMarkup(1)
            button = types.InlineKeyboardButton('Редактировать\ud83d\udd27', callback_data='edit_prod' + str(calldata))
            order_markup.add(button)
            if calldataa == len(cate_prod) - 1:
                cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
                adding_prod.row(cancel)
            if calldata == len(cate_prod) - 1:
                cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
                order_markup.row(cancel)
            if product[3] != 'не указано':
                bot.edit_message_text(
                    text='<a href="' + productt[5] + '"> </a>' + productt[0] + ' - ' + str(
                        productt[2]) + 'тг.\n\n' + productt[1] + '\nВидно: ' + productt[4],
                    chat_id=call.message.chat.id, parse_mode='HTML', message_id=message[0], reply_markup=order_markup)
            else:
                bot.edit_message_text(
                    text='<a href="' + productt[5] + '"> </a>' + productt[0] + ' - ' + str(
                        productt[2]) + 'тг.\n\n' + productt[1] + '\nВидно: ' + productt[4],
                    chat_id=call.message.chat.id, parse_mode='HTML', message_id=message[0], reply_markup=order_markup)
        else:
            if calldataa == len(cate_prod) - 1:
                cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
                adding_prod.row(cancel)
        bot.edit_message_text(chat_id=call.message.chat.id,
                              text='<b>Информация о продукте</b>\n\nИмя: ' + product[
                                  0] + '\nОписание: \n' + product[
                                       1] + '\nЦена: ' + str(product[2]) + '\nКатегория: ' + product[
                                       3] + '\nВидимость: ' +
                                   product[4] + '\nФото: ' + product[5], parse_mode='HTML',
                              message_id=call.message.message_id,
                              reply_markup=adding_prod)
    if call.data.startswith('eedit_name'):
        sql = 'UPDATE users SET step = %s WHERE user_id = %s'
        val = ('eedit_name', call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT product, desc_, price, type, seen, link FROM productlist WHERE process = %s AND user_id = %s'
        val = ('editing', call.message.chat.id)
        cursor.execute(sql, val)
        product = cursor.fetchone()
        sql = 'SELECT message FROM messages WHERE user_id = %s'
        val = (call.message.chat.id, )
        cursor.execute(sql, val)
        messages = cursor.fetchall()
        food_type = types.InlineKeyboardMarkup(2)
        sql = 'SELECT start_msg FROM calldata WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        start_msg = cursor.fetchone()
        bot.delete_message(call.message.chat.id, start_msg[0])
        for i in range(len(messages)):
            bot.delete_message(call.message.chat.id, messages[i][0])
        cancel = types.InlineKeyboardButton('Отмена\u274c', callback_data='stop_edit' + product[0])
        food_type.add(cancel)
        bot.send_message(text='Введите новое имя к продукту ' + product[0] + ':', chat_id=call.message.chat.id, reply_markup=food_type)
        sql = 'DELETE FROM messages WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        db.commit()
    if call.data == 'among_cates1':
        sql = 'UPDATE users SET step = %s WHERE user_id = %s'
        val = (None, call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT message FROM messages WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        messages = cursor.fetchall()
        sql = 'UPDATE productlist SET process = %s WHERE user_id = %s AND process = %s'
        val = ('saved', call.message.chat.id, 'editing')
        cursor.execute(sql, val)
        db.commit()
        if messages == []:
            sql = 'SELECT name FROM categorylist'
            cursor.execute(sql)
            category = cursor.fetchall()
            food_type = types.InlineKeyboardMarkup(2)
            for i in range(len(category)):
                button = types.InlineKeyboardButton(category[i][0], callback_data='_choose' + str(i))
                food_type.add(button)
            cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='backadnu')
            food_type.add(cancel)
            bot.edit_message_text('Выберите категорию:', call.message.chat.id, call.message.message_id, reply_markup=food_type)
        else:
            sql = 'SELECT start_msg FROM calldata WHERE user_id = %s'
            val = (call.message.chat.id, )
            cursor.execute(sql, val)
            start_msg = cursor.fetchone()
            bot.delete_message(call.message.chat.id, start_msg[0])
            for i in range(len(messages)):
                bot.delete_message(call.message.chat.id, messages[i][0])
            sql = 'SELECT name FROM categorylist'
            cursor.execute(sql)
            category = cursor.fetchall()
            food_type = types.InlineKeyboardMarkup(2)
            for i in range(len(category)):
                button = types.InlineKeyboardButton(category[i][0], callback_data='_choose' + str(i))
                food_type.add(button)
            cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='backadnu')
            food_type.add(cancel)
            bot.send_message(text='Выберите категорию:', chat_id=call.message.chat.id, reply_markup=food_type)
            sql = 'DELETE FROM messages WHERE user_id = %s'
            val = (call.message.chat.id,)
            cursor.execute(sql, val)
            db.commit()
    if call.data.startswith('stop_edit'):
        sql = 'UPDATE users SET step = %s WHERE user_id = %s'
        val = (None, call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        sql = 'UPDATE calldata SET start_msg = %s WHERE user_id = %s'
        val = (call.message.message_id, call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        sql = 'UPDATE calldata SET start_msg = %s WHERE user_id = %s'
        val = (call.message.message_id, call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT type FROM productlist WHERE product = %s'
        val = (call.data[9:], )
        cursor.execute(sql, val)
        type = cursor.fetchone()
        sql = 'SELECT product, price, desc_, link, seen, type FROM productlist WHERE type = %s AND process NOT IN (%s)'
        val = (type[0], 'making')
        cursor.execute(sql, val)
        product = cursor.fetchall()
        bot.edit_message_text(text='Категория - ' + type[0], chat_id=call.message.chat.id,
                              message_id=call.message.message_id)
        for i in range(len(product)):
            food_type = types.InlineKeyboardMarkup(2)
            button = types.InlineKeyboardButton('Редактировать\ud83d\udd27', callback_data='edit_prod' + str(i))
            food_type.add(button)
            sql = 'INSERT INTO messages (user_id, message) VALUES (%s, %s)'
            val = (call.message.chat.id, call.message.message_id + 1 + i)
            cursor.execute(sql, val)
            db.commit()
            if product[i][0] == call.data[9:]:
                sql = 'UPDATE calldata SET message = %s WHERE user_id = %s'
                val = (call.message.message_id + 1 + i, call.message.chat.id)
                cursor.execute(sql, val)
                adding_prod = types.InlineKeyboardMarkup(2)
                preview = types.InlineKeyboardButton('Предпросмотр', callback_data='epreview_prod' + product[i][0])
                edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='eedit_name' + product[i][0])
                edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='eedit_photo' + product[i][0])
                edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='eedit_desc' + product[i][0])
                edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='eedit_price' + product[i][0])
                edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='eedit_cate' + product[i][0])
                show = types.InlineKeyboardButton('Показать\u2705', callback_data='eshow' + product[i][0])
                hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='ehide' + product[i][0])
                delete = types.InlineKeyboardButton('Удалить продукт\u274c', callback_data='edelete_prod' + product[i][0])
                adding_prod.add(delete, preview)
                if product[i][4] == '<b>ДА</b>':
                    adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, hide)
                else:
                    adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show)
                if i == len(product) - 1:
                    cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
                    adding_prod.row(cancel)
                bot.send_message(chat_id=call.message.chat.id,
                                      text='<b>Информация о продукте</b>\n\nИмя: ' + product[
                                          i][0] + '\nОписание: \n' + product[
                                               i][2] + '\nЦена: ' + str(product[i][1]) + '\nКатегория: ' + product[
                                               i][5] + '\nВидимость: ' +
                                           product[i][4] + '\nФото: ' + product[i][3], parse_mode='HTML', reply_markup=adding_prod)
                continue
            if i == len(product) - 1:
                cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
                food_type.add(cancel)
            if product[i][3] != 'не указано':
                bot.send_message(
                    text='<a href="' + product[i][3] + '"> </a>' + product[i][0] + ' - ' + str(
                        product[i][1]) + 'тг.\n' + product[i][2] + '\nВидно: ' + product[i][4],
                    chat_id=call.message.chat.id, parse_mode='HTML', reply_markup=food_type)
            else:
                bot.send_message(
                    text=product[i][0] + ' - ' + str(product[i][1]) + 'тг.\n' + product[i][
                        2] + '\nОТСУТСТВУЕТ ФОТОГРАФИЯ\nВидно: ' + product[i][4],
                    chat_id=call.message.chat.id, parse_mode='HTML', reply_markup=food_type)

    if call.data.startswith('eedit_photo'):
        sql = 'SELECT start_msg FROM calldata WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        start_msg = cursor.fetchone()
        sql = 'UPDATE users SET step = %s WHERE user_id = %s'
        val = ('eedit_photo', call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT product, desc_, price, type, seen, link FROM productlist WHERE process = %s AND user_id = %s'
        val = ('editing', call.message.chat.id)
        cursor.execute(sql, val)
        product = cursor.fetchone()
        sql = 'SELECT message FROM messages WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        messages = cursor.fetchall()
        food_type = types.InlineKeyboardMarkup(2)
        bot.delete_message(call.message.chat.id, start_msg[0])
        for i in range(len(messages)):
            bot.delete_message(call.message.chat.id, messages[i][0])
        cancel = types.InlineKeyboardButton('Отмена\u274c', callback_data='stop_edit' + product[0])
        food_type.add(cancel)
        bot.send_message(text='Вставьте ссылку ведущую к фото к продукту' + product[0] + ':', chat_id=call.message.chat.id,
                         reply_markup=food_type)
        sql = 'DELETE FROM messages WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        db.commit()
    if call.data.startswith('eedit_desc'):
        sql = 'SELECT start_msg FROM calldata WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        start_msg = cursor.fetchone()
        sql = 'UPDATE users SET step = %s WHERE user_id = %s'
        val = ('eedit_desc', call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT product, desc_, price, type, seen, link FROM productlist WHERE process = %s AND user_id = %s'
        val = ('editing', call.message.chat.id)
        cursor.execute(sql, val)
        product = cursor.fetchone()
        sql = 'SELECT message FROM messages WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        messages = cursor.fetchall()
        food_type = types.InlineKeyboardMarkup(2)
        bot.delete_message(call.message.chat.id, start_msg[0])
        for i in range(len(messages)):
            bot.delete_message(call.message.chat.id, messages[i][0])
        cancel = types.InlineKeyboardButton('Отмена\u274c', callback_data='stop_edit' + product[0])
        food_type.add(cancel)
        bot.send_message(text='Введите новое описание к продукту ' + product[0] + ':', chat_id=call.message.chat.id,
                         reply_markup=food_type)
        sql = 'DELETE FROM messages WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        db.commit()

    if call.data.startswith('eedit_price'):
        sql = 'SELECT start_msg FROM calldata WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        start_msg = cursor.fetchone()
        sql = 'UPDATE users SET step = %s WHERE user_id = %s'
        val = ('eedit_price', call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT product, desc_, price, type, seen, link FROM productlist WHERE process = %s AND user_id = %s'
        val = ('editing', call.message.chat.id)
        cursor.execute(sql, val)
        product = cursor.fetchone()
        sql = 'SELECT message FROM messages WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        messages = cursor.fetchall()
        food_type = types.InlineKeyboardMarkup(2)
        bot.delete_message(call.message.chat.id, start_msg[0])
        for i in range(len(messages)):
            bot.delete_message(call.message.chat.id, messages[i][0])
        cancel = types.InlineKeyboardButton('Отмена\u274c', callback_data='stop_edit' + product[0])
        food_type.add(cancel)
        bot.send_message(text='Введите новую цену к продукту ' + product[0] + ':', chat_id=call.message.chat.id,
                         reply_markup=food_type)
        sql = 'DELETE FROM messages WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        db.commit()
    if call.data.startswith('eedit_cate'):
        sql = 'SELECT start_msg FROM calldata WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        start_msg = cursor.fetchone()
        sql = 'SELECT product, desc_, price, type, seen, link FROM productlist WHERE product = %s'
        val = (call.data[10:], )
        cursor.execute(sql, val)
        product = cursor.fetchone()
        sql = 'SELECT message FROM messages WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        messages = cursor.fetchall()
        food_type = types.InlineKeyboardMarkup(2)
        bot.delete_message(call.message.chat.id, start_msg[0])
        for i in range(len(messages)):
            bot.delete_message(call.message.chat.id, messages[i][0])
        food_type = types.InlineKeyboardMarkup(2)
        sql = 'SELECT name FROM categorylist'
        cursor.execute(sql)
        category = cursor.fetchall()
        for i in range(len(category)):
            button = types.InlineKeyboardButton(category[i][0], callback_data='cate_edit' + str(i))
            food_type.add(button)
        cancel = types.InlineKeyboardButton('Отмена\u274c', callback_data='stop_edit' + product[0])
        food_type.add(cancel)
        bot.send_message(text='Выберите категорию в которой будет находиться продукт',
                              chat_id=call.message.chat.id, reply_markup=food_type)
        sql = 'DELETE FROM messages WHERE user_id = %s'
        val = (call.message.chat.id,)
        cursor.execute(sql, val)
        db.commit()
    if call.data.startswith('cate_edit'):
        sql = 'UPDATE calldata SET start_msg = %s WHERE user_id = %s'
        val = (call.message.message_id, call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT name FROM categorylist'
        cursor.execute(sql)
        category = cursor.fetchall()
        sql = 'UPDATE calldata SET calldata2 = %s WHERE user_id = %s'
        val = (int(call.data[9:]), call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        bot.edit_message_text('Категория - ' + category[int(call.data[9:])][0], call.message.chat.id, call.message.message_id)
        sql = 'UPDATE productlist SET type = %s WHERE user_id = %s AND process = %s'
        val = (category[int(call.data[9:])][0], call.message.chat.id, 'editing')
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT product, price, desc_, type, seen, link FROM productlist WHERE user_id = %s AND process = %s'
        val = (call.message.chat.id, 'editing')
        cursor.execute(sql, val)
        productt = cursor.fetchone()
        sql = 'SELECT product, price, desc_, link, seen, type FROM productlist WHERE type = %s AND process NOT IN (%s)'
        val = (productt[3], 'making')
        cursor.execute(sql, val)
        product = cursor.fetchall()
        for i in range(len(product)):
            food_type = types.InlineKeyboardMarkup(2)
            button = types.InlineKeyboardButton('Редактировать\ud83d\udd27', callback_data='edit_prod' + str(i))
            food_type.add(button)
            sql = 'INSERT INTO messages (user_id, message) VALUES (%s, %s)'
            val = (call.message.chat.id, call.message.message_id + 1 + i)
            cursor.execute(sql, val)
            db.commit()
            if product[i][0] == productt[0]:
                sql = 'UPDATE calldata SET message = %s WHERE user_id = %s'
                val = (call.message.message_id + 1 + i, call.message.chat.id)
                cursor.execute(sql, val)
                db.commit()
                adding_prod = types.InlineKeyboardMarkup(2)
                preview = types.InlineKeyboardButton('Предпросмотр', callback_data='epreview_prod' + product[i][0])
                edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='eedit_name' + product[i][0])
                edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='eedit_photo' + product[i][0])
                edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='eedit_desc' + product[i][0])
                edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='eedit_price' + product[i][0])
                edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='eedit_cate' + product[i][0])
                show = types.InlineKeyboardButton('Показать\u2705', callback_data='eshow' + product[i][0])
                hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='ehide' + product[i][0])
                delete = types.InlineKeyboardButton('Удалить продукт\u274c',
                                                    callback_data='edelete_prod' + product[i][0])
                adding_prod.add(delete, preview)
                if product[i][4] == '<b>ДА</b>':
                    adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, hide)
                else:
                    adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show)
                if i == len(product) - 1:
                    cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
                    adding_prod.row(cancel)
                bot.send_message(chat_id=call.message.chat.id,
                                 text='Категория продукта отредактирована!\n<b>Информация о продукте</b>\n\nИмя: ' +
                                      product[
                                          i][0] + '\nОписание: \n' + product[
                                          i][2] + '\nЦена: ' + str(product[i][1]) + '\nКатегория: ' + product[
                                          i][5] + '\nВидимость: ' +
                                      product[i][4] + '\nФото: ' + product[i][3], parse_mode='HTML',
                                 reply_markup=adding_prod)
                continue
            if i == len(product) - 1:
                cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
                food_type.add(cancel)
            if product[i][3] != 'не указано':
                bot.send_message(
                    text='<a href="' + product[i][3] + '"> </a>' + product[i][0] + ' - ' + str(
                        product[i][1]) + 'тг.\n' + product[i][2] + '\nВидно: ' + product[i][4],
                    chat_id=call.message.chat.id, parse_mode='HTML', reply_markup=food_type)
            else:
                bot.send_message(
                    text=product[i][0] + ' - ' + str(product[i][1]) + 'тг.\n' + product[i][
                        2] + '\nОТСУТСТВУЕТ ФОТОГРАФИЯ\nВидно: ' + product[i][4],
                    chat_id=call.message.chat.id, parse_mode='HTML', reply_markup=food_type)
    if call.data.startswith('eshow'):
        sql = 'UPDATE productlist SET seen = %s WHERE user_id = %s AND process = %s'
        val = ('<b>ДА</b>', call.message.chat.id, 'editing')
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT product, price, desc_, seen, type, link FROM productlist WHERE user_id = %s AND process = %s'
        val = (call.message.chat.id, 'editing')
        cursor.execute(sql, val)
        product = cursor.fetchone()
        sql = 'SELECT message FROM messages WHERE user_id = %s'
        val = (call.message.chat.id, )
        cursor.execute(sql, val)
        messages = cursor.fetchall()
        adding_prod = types.InlineKeyboardMarkup(2)
        preview = types.InlineKeyboardButton('Предпросмотр', callback_data='epreview_prod' + product[0])
        edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='eedit_name' + product[0])
        edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='eedit_photo' + product[0])
        edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='eedit_desc' + product[0])
        edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='eedit_price' + product[0])
        edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='eedit_cate' + product[0])
        show = types.InlineKeyboardButton('Показать\u2705', callback_data='eshow' + product[0])
        hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='ehide' + product[0])
        delete = types.InlineKeyboardButton('Удалить продукт\u274c',
                                            callback_data='edelete_prod' + product[0])
        adding_prod.add(delete, preview)
        if product[3] == '<b>ДА</b>':
            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, hide)
        else:
            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show)
        if call.message.message_id == messages[len(messages) - 1][0]:
            cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
            adding_prod.row(cancel)
        bot.edit_message_text(chat_id=call.message.chat.id,
                         text='Продукт показан!\n<b>Информация о продукте</b>\n\nИмя: ' +
                              product
                              [0] + '\nОписание: \n' + product[2] + '\nЦена: ' + str(product[1]) + '\nКатегория: ' +
                              product[4] + '\nВидимость: ' +
                              product[3] + '\nФото: ' + product[5], message_id=call.message.message_id, parse_mode='HTML',
                         reply_markup=adding_prod)
    if call.data.startswith('ehide'):
        sql = 'UPDATE productlist SET seen = %s WHERE user_id = %s AND process = %s'
        val = ('<b>НЕТ</b>', call.message.chat.id, 'editing')
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT product, price, desc_, seen, type, link FROM productlist WHERE user_id = %s AND process = %s'
        val = (call.message.chat.id, 'editing')
        cursor.execute(sql, val)
        product = cursor.fetchone()
        sql = 'SELECT message FROM messages WHERE user_id = %s'
        val = (call.message.chat.id, )
        cursor.execute(sql, val)
        messages = cursor.fetchall()
        adding_prod = types.InlineKeyboardMarkup(2)
        preview = types.InlineKeyboardButton('Предпросмотр', callback_data='epreview_prod' + product[0])
        edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='eedit_name' + product[0])
        edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='eedit_photo' + product[0])
        edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='eedit_desc' + product[0])
        edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='eedit_price' + product[0])
        edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='eedit_cate' + product[0])
        show = types.InlineKeyboardButton('Показать\u2705', callback_data='eshow' + product[0])
        hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='ehide' + product[0])
        delete = types.InlineKeyboardButton('Удалить продукт\u274c',
                                            callback_data='edelete_prod' + product[0])
        adding_prod.add(delete, preview)
        if product[3] == '<b>ДА</b>':
            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, hide)
        else:
            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show)
        if call.message.message_id == messages[len(messages) - 1][0]:
            cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
            adding_prod.row(cancel)
        bot.edit_message_text(chat_id=call.message.chat.id,
                         text='Продукт скрыт!\n<b>Информация о продукте</b>\n\nИмя: ' +
                              product
                                  [0] + '\nОписание: \n' + product[2] + '\nЦена: ' + str(product[1]) + '\nКатегория: ' +
                              product[4] + '\nВидимость: ' +
                              product[3] + '\nФото: ' + product[5], message_id=call.message.message_id, parse_mode='HTML',
                         reply_markup=adding_prod)
    if call.data.startswith('edelete_prod'):
        sql = 'SELECT product, price, desc_, link, seen, type FROM productlist WHERE product = %s'
        val = (call.data[12:],)
        cursor.execute(sql, val)
        product = cursor.fetchone()
        sql = 'SELECT product FROM productlist WHERE type = %s'
        val = (product[5],)
        cursor.execute(sql, val)
        typee = cursor.fetchall()
        string = call.data[12:]
        tup = (string,)
        delete = types.InlineKeyboardMarkup()
        yes = types.InlineKeyboardButton('Да\u2705', callback_data='eyes_delete_prod' + call.data[12:])
        nope_don = types.InlineKeyboardButton('Нет\u274c', callback_data='eno_dont_prod' + call.data[12:])
        delete.add(yes, nope_don)
        if typee.index(tup) == len(typee) - 1:
            cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
            delete.row(cancel)
        bot.edit_message_text('Вы уверены, что хотите удалить продукт ' + call.data[12:] + '?', call.message.chat.id, call.message.message_id, reply_markup=delete)
    if call.data.startswith('eyes_delete_prod'):
        sql = 'UPDATE calldata SET message = %s WHERE user_id = %s'
        val = (None, call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        sql = 'DELETE FROM messages WHERE user_id = %s AND message = %s'
        val = (call.message.chat.id, call.message.message_id)
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT type FROM productlist WHERE product = %s'
        val = (call.data[16:],)
        cursor.execute(sql, val)
        typef = cursor.fetchone()
        sql = 'SELECT product FROM productlist WHERE type = %s'
        val = (typef[0], )
        cursor.execute(sql, val)
        product = cursor.fetchall()
        sql = 'DELETE FROM productlist WHERE product = %s'
        val = (call.data[16:],)
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT product FROM productlist WHERE type = %s'
        val = (typef[0],)
        cursor.execute(sql, val)
        sql = 'SELECT * FROM productlist WHERE type = %s'
        val = (typef[0], )
        cursor.execute(sql, val)
        productt = cursor.fetchall()
        if cursor.rowcount == 0:
            sql = 'SELECT start_msg FROM calldata WHERE user_id = %s'
            val = (call.message.chat.id, )
            cursor.execute(sql, val)
            message = cursor.fetchone()
            bot.delete_message(call.message.chat.id, message[0])
            bot.delete_message(call.message.chat.id, call.message.message_id)
            sql = 'DELETE FROM messages WHERE user_id = %s'
            val = (call.message.chat.id,)
            cursor.execute(sql, val)
            db.commit()
            food_type = types.InlineKeyboardMarkup(2)
            cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
            food_type.add(cancel)
            bot.send_message(text='В этой категории ничего нет.', chat_id=call.message.chat.id, reply_markup=food_type)
        else:
            bot.delete_message(call.message.chat.id, call.message.message_id)
            sql = 'SELECT message FROM messages WHERE user_id = %s'
            val = (call.message.chat.id,)
            cursor.execute(sql, val)
            messages = cursor.fetchall()
            for i in range(len(messages)):
                try:
                    food_type = types.InlineKeyboardMarkup(2)
                    button = types.InlineKeyboardButton('Редактировать\ud83d\udd27', callback_data='edit_prod' + str(i))
                    food_type.add(button)
                    if i == len(messages) - 1:
                        cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
                        food_type.add(cancel)
                    bot.edit_message_reply_markup(call.message.chat.id, messages[i][0], reply_markup=food_type)
                except:
                    if i == len(messages) - 1:
                        cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
                        food_type.add(cancel)
                        bot.edit_message_reply_markup(call.message.chat.id, messages[i][0], reply_markup=food_type)
                    else:
                        pass
            sql = 'UPDATE calldata SET message = %s WHERE user_id = %s'
            val = (None, call.message.chat.id)
            cursor.execute(sql, val)
            db.commit()
    if call.data.startswith('eno_dont_prod'):
        sql = 'SELECT product, price, desc_, link, seen, type FROM productlist WHERE product = %s'
        val = (call.data[13:], )
        cursor.execute(sql, val)
        product = cursor.fetchone()
        sql = 'SELECT product FROM productlist WHERE type = %s'
        val = (product[5], )
        cursor.execute(sql, val)
        typee = cursor.fetchall()
        string = call.data[13:]
        tup = (string, )
        adding_prod = types.InlineKeyboardMarkup(2)
        preview = types.InlineKeyboardButton('Предпросмотр', callback_data='epreview_prod' + product[0])
        edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='eedit_name' + product[0])
        edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='eedit_photo' + product[0])
        edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='eedit_desc' + product[0])
        edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='eedit_price' + product[0])
        edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='eedit_cate' + product[0])
        show = types.InlineKeyboardButton('Показать\u2705', callback_data='eshow' + product[0])
        hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='ehide' + product[0])
        delete = types.InlineKeyboardButton('Удалить продукт\u274c', callback_data='edelete_prod' + product[0])
        adding_prod.add(delete, preview)
        if product[4] == '<b>ДА</b>':
            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, hide)
        else:
            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show)
        if typee.index(tup) == len(typee) - 1:
            cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
            adding_prod.row(cancel)
        bot.edit_message_text(chat_id=call.message.chat.id,
                         text='<b>Информация о продукте</b>\n\nИмя: ' + product
                                [0] + '\nОписание: \n' + product
                                  [2] + '\nЦена: ' + str(product[1]) + '\nКатегория: ' + product
                                  [5] + '\nВидимость: ' +
                              product[4] + '\nФото: ' + product[3],
                              message_id=call.message.message_id, parse_mode='HTML', reply_markup=adding_prod)
    if call.data.startswith('epreview_prod'):
        sql = 'SELECT product, price, desc_, link, type FROM productlist WHERE product = %s'
        val = (call.data[13:], )
        cursor.execute(sql, val)
        product = cursor.fetchone()
        sql = 'SELECT product FROM productlist WHERE type = %s'
        val = (product[4],)
        cursor.execute(sql, val)
        typee = cursor.fetchall()
        string = call.data[13:]
        tup = (string,)
        back_mk = types.InlineKeyboardMarkup()
        back = types.InlineKeyboardButton('Вернуться к продукту', callback_data='eback_to' + product[0])
        back_mk.add(back)
        if typee.index(tup) == len(typee) - 1:
            cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
            back_mk.row(cancel)
        if product[3] != 'не указано':
            bot.edit_message_text('<a href="' + product[3] + '"> </a>' + product[0] + ' - ' + str(product[1]) + '\n' + product[2], call.message.chat.id, call.message.message_id, parse_mode='HTML', reply_markup=back_mk)
        else:
            bot.edit_message_text(
                product[0] + ' - ' + str(product[1]) + '\n' + product[2] + '\nОТСУТСВУЕТ ФОТОГРАФИЯ',
                call.message.chat.id, call.message.message_id, parse_mode='HTML', reply_markup=back_mk)
    if call.data.startswith('eback_to'):
        sql = 'SELECT product, price, desc_, link, seen, type FROM productlist WHERE product = %s'
        val = (call.data[8:],)
        cursor.execute(sql, val)
        product = cursor.fetchone()
        sql = 'SELECT product FROM productlist WHERE type = %s'
        val = (product[5],)
        cursor.execute(sql, val)
        typee = cursor.fetchall()
        string = call.data[8:]
        tup = (string,)
        adding_prod = types.InlineKeyboardMarkup(2)
        preview = types.InlineKeyboardButton('Предпросмотр', callback_data='epreview_prod' + product[0])
        edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='eedit_name' + product[0])
        edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='eedit_photo' + product[0])
        edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='eedit_desc' + product[0])
        edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='eedit_price' + product[0])
        edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='eedit_cate' + product[0])
        show = types.InlineKeyboardButton('Показать\u2705', callback_data='eshow' + product[0])
        hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='ehide' + product[0])
        delete = types.InlineKeyboardButton('Удалить продукт\u274c', callback_data='edelete_prod' + product[0])
        adding_prod.add(delete, preview)
        if product[4] == '<b>ДА</b>':
            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, hide)
        else:
            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show)
        if typee.index(tup) == len(typee) - 1:
            cancel = types.InlineKeyboardButton('Назад \u2b05', callback_data='among_cates1')
            adding_prod.row(cancel)
        bot.edit_message_text(chat_id=call.message.chat.id,
                              text='<b>Информация о продукте</b>\n\nИмя: ' + product
                              [0] + '\nОписание: \n' + product
                                   [2] + '\nЦена: ' + str(product[1]) + '\nКатегория: ' + product
                                   [5] + '\nВидимость: ' +
                                   product[4] + '\nФото: ' + product[3],
                              message_id=call.message.message_id, parse_mode='HTML', reply_markup=adding_prod)
    if call.data == 'cate':
        mk = types.InlineKeyboardMarkup(1)
        add_cate = types.InlineKeyboardButton('Добавить\u2795', callback_data='add_cate')
        cate_list = types.InlineKeyboardButton('Список категорий', callback_data='cate_list')
        back = types.InlineKeyboardButton('Назад \u2b05', callback_data='backadnu')
        mk.add(add_cate, cate_list, back)
        bot.edit_message_text('<b>Категории</b>\nВыберите действие:', call.message.chat.id, call.message.message_id, parse_mode='HTML', reply_markup=mk)
    if call.data == 'add_cate':
        sql = 'UPDATE calldata SET start_msg = %s WHERE user_id = %s'
        val = (call.message.message_id, call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        markup_cancel = types.InlineKeyboardMarkup()
        cancel = types.InlineKeyboardButton('Отмена\u274c', callback_data='cate')
        markup_cancel.add(cancel)
        sql = 'UPDATE users SET step = %s WHERE user_id = %s'
        val = ('cate_naming', call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        bot.edit_message_text('Введите имя новой категории:', call.message.chat.id, call.message.message_id, reply_markup=markup_cancel)
    if call.data.startswith('yes_hide_cate'):
        sql = 'SELECT name FROM categorylist WHERE name = %s'
        val = (call.data[13:], )
        cursor.execute(sql, val)
        category = cursor.fetchone()
        mk = types.InlineKeyboardMarkup(1)
        add_cate = types.InlineKeyboardButton('Добавить категорию\u2795', callback_data='add_cate')
        cate_list = types.InlineKeyboardButton('Редактировать категорию', callback_data='cate_list')
        back = types.InlineKeyboardButton('Назад \u2b05', callback_data='backadnu')
        mk.add(add_cate, cate_list, back)
        bot.edit_message_text('<b>Категории</b>\nКатегория ' + category[0] + ' была создана и скрыта от пользователей!\nВыберите действие:', call.message.chat.id, call.message.message_id,
                              parse_mode='HTML', reply_markup=mk)
    if call.data.startswith('no_donthide_cate'):
        sql = 'UPDATE categorylist SET seen = %s WHERE name = %s'
        val = ('<b>ДА</b>', call.data[16:])
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT name FROM categorylist WHERE name = %s'
        val = (call.data[16:],)
        cursor.execute(sql, val)
        category = cursor.fetchone()
        mk = types.InlineKeyboardMarkup(1)
        add_cate = types.InlineKeyboardButton('Добавить категорию\u2795', callback_data='add_cate')
        cate_list = types.InlineKeyboardButton('Редактировать категорию', callback_data='cate_list')
        back = types.InlineKeyboardButton('Назад \u2b05', callback_data='backadnu')
        mk.add(add_cate, cate_list, back)
        bot.edit_message_text('<b>Категории</b>\nКатегория ' + category[
            0] + ' была создана и видна пользователям!\nВыберите действие:', call.message.chat.id,
                              call.message.message_id,
                              parse_mode='HTML', reply_markup=mk)
    if call.data == 'cate_list':
        sql = 'SELECT name, seen FROM categorylist'
        cursor.execute(sql)
        category = cursor.fetchall()
        cate_list = types.InlineKeyboardMarkup(4)
        if cursor.rowcount > 0:
            for i in range(len(category)):
                button = types.InlineKeyboardButton(category[i][0], callback_data=str(i) + '//')
                name = types.InlineKeyboardButton('Ред.Имя', callback_data='go_cate' + category[i][0])
                show = types.InlineKeyboardButton('Показать\u2705', callback_data='gshow_cate' + category[i][0])
                hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='ghide_cate' + category[i][0])
                delete = types.InlineKeyboardButton('Удалить\ud83d\uddd1', callback_data='gdeletecate' + category[i][0])
                if category[i][1] == '<b>НЕТ</b>':
                    cate_list.add(button, name, show, delete)
                else:
                    cate_list.add(button, name, hide, delete)
            back = types.InlineKeyboardButton('Назад\u2b05', callback_data='cate')
            cate_list.row(back)
            bot.edit_message_text('<b>Список всех категорий</b>',
                                  call.message.chat.id, call.message.message_id,
                                  reply_markup=cate_list, parse_mode='HTML')
        else:
            back = types.InlineKeyboardButton('Назад\u2b05', callback_data='cate')
            cate_list.row(back)
            bot.edit_message_text('Категорий не найдено. Попробуйте создать парочку нажав на кнопку Добавить\u2795!', call.message.chat.id, call.message.message_id, reply_markup=cate_list, parse_mode='HTML')
    if call.data.startswith('gshow_cate'):
        sql = 'UPDATE categorylist SET seen = %s WHERE name = %s'
        val = ('<b>ДА</b>', call.data[10:])
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT name, seen FROM categorylist'
        cursor.execute(sql)
        category = cursor.fetchall()
        cate_list = types.InlineKeyboardMarkup(4)
        for i in range(len(category)):
            button = types.InlineKeyboardButton(category[i][0], callback_data=str(i) + '//')
            name = types.InlineKeyboardButton('Ред.Имя', callback_data='go_cate' + category[i][0])
            show = types.InlineKeyboardButton('Показать\u2705', callback_data='gshow_cate' + category[i][0])
            hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='ghide_cate' + category[i][0])
            delete = types.InlineKeyboardButton('Удалить\ud83d\uddd1', callback_data='gdeletecate' + category[i][0])
            if category[i][1] == '<b>НЕТ</b>':
                cate_list.add(button, name, show, delete)
            else:
                cate_list.add(button, name, hide, delete)
        back = types.InlineKeyboardButton('Назад\u2b05', callback_data='cate')
        cate_list.row(back)
        bot.edit_message_text('Категория ' + call.data[10:] + ' показана!\n<b>Список всех категорий</b>',
                              call.message.chat.id, call.message.message_id,
                              reply_markup=cate_list, parse_mode='HTML')
    if call.data.startswith('ghide_cate'):
        sql = 'UPDATE categorylist SET seen = %s WHERE name = %s'
        val = ('<b>НЕТ</b>', call.data[10:])
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT name, seen FROM categorylist'
        cursor.execute(sql)
        category = cursor.fetchall()
        cate_list = types.InlineKeyboardMarkup(4)
        for i in range(len(category)):
            button = types.InlineKeyboardButton(category[i][0], callback_data=str(i) + '//')
            name = types.InlineKeyboardButton('Ред.Имя', callback_data='go_cate' + category[i][0])
            show = types.InlineKeyboardButton('Показать\u2705', callback_data='gshow_cate' + category[i][0])
            hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='ghide_cate' + category[i][0])
            delete = types.InlineKeyboardButton('Удалить\ud83d\uddd1', callback_data='gdeletecate' + category[i][0])
            if category[i][1] == '<b>НЕТ</b>':
                cate_list.add(button, name, show, delete)
            else:
                cate_list.add(button, name, hide, delete)
        back = types.InlineKeyboardButton('Назад\u2b05', callback_data='cate')
        cate_list.row(back)
        bot.edit_message_text('Категория ' + call.data[10:] + ' скрыта!\n<b>Список всех категорий</b>', call.message.chat.id, call.message.message_id,
                              reply_markup=cate_list, parse_mode='HTML')
    if call.data.startswith('gdeletecate'):
        yes_or_no = types.InlineKeyboardMarkup()
        yes = types.InlineKeyboardButton('Да\u2705', callback_data='del_cate' + call.data[11:])
        no = types.InlineKeyboardButton('Нет\u274c', callback_data='cate_list')
        yes_or_no.add(yes, no)
        bot.edit_message_text('Вы уверены, что хотите удалить категорию ' + call.data[11:] + '?\n<b>ВНИМАНИЕ</b>: Вместе с категорией удалятся все находящиеся внутри продукты.', call.message.chat.id, call.message.message_id, parse_mode='HTML', reply_markup=yes_or_no)


    if call.data.startswith('del_cate'):
        sql = 'DELETE FROM categorylist WHERE name = %s'
        val = (call.data[8:], )
        cursor.execute(sql, val)
        db.commit()
        sql = 'DELETE FROM productlist WHERE type = %s'
        val = (call.data[8:], )
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT name, seen FROM categorylist'
        cursor.execute(sql)
        category = cursor.fetchall()
        cate_list = types.InlineKeyboardMarkup(4)
        if cursor.rowcount > 0:
            for i in range(len(category)):
                button = types.InlineKeyboardButton(category[i][0], callback_data=str(i) + '//')
                name = types.InlineKeyboardButton('Ред.Имя', callback_data='go_cate' + category[i][0])
                show = types.InlineKeyboardButton('Показать\u2705', callback_data='gshow_cate' + category[i][0])
                hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='ghide_cate' + category[i][0])
                delete = types.InlineKeyboardButton('Удалить\ud83d\uddd1',
                                                    callback_data='gdeletecate' + category[i][0])
                if category[i][1] == '<b>НЕТ</b>':
                    cate_list.add(button, name, show, delete)
                else:
                    cate_list.add(button, name, hide, delete)
            back = types.InlineKeyboardButton('Назад\u2b05', callback_data='cate')
            cate_list.row(back)
            bot.edit_message_text('Категория ' + call.data[8:] + ' удалена!\n<b>Список всех категорий</b>', call.message.chat.id, call.message.message_id,
                                  reply_markup=cate_list, parse_mode='HTML')
        else:
            back = types.InlineKeyboardButton('Назад\u2b05', callback_data='cate')
            cate_list.row(back)
            bot.edit_message_text('Категорий не найдено. Попробуйте создать парочку нажав на кнопку "Добавить категорию\u2795"!', call.message.chat.id, call.message.message_id,
                                  reply_markup=cate_list, parse_mode='HTML')
    if call.data.startswith('go_cate'):
        sql = 'UPDATE calldata SET start_msg = %s WHERE user_id = %s'
        val = (call.message.message_id, call.message.chat.id)
        cursor.execute(sql, val)
        sql = 'UPDATE users SET step = %s WHERE user_id = %s'
        val = ('edit_name_cate' + call.data[7:], call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        mk = types.InlineKeyboardMarkup()
        cancel = types.InlineKeyboardButton('Отмена\u274c', callback_data='cate_list')
        mk.add(cancel)
        bot.edit_message_text('Введите новое название категории:', call.message.chat.id, call.message.message_id, reply_markup=mk)
    if call.data == 'disc':
        sql = 'UPDATE productlist SET disc = %s WHERE process = %s AND user_id = %s'
        val = ('<b>ДА</b>', 'making', call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT product, desc_, price, type, seen, link FROM productlist WHERE user_id = %s AND process = %s'
        val = (call.message.chat.id, 'making')
        cursor.execute(sql, val)
        product = cursor.fetchone()
        adding_prod = types.InlineKeyboardMarkup(2)
        preview = types.InlineKeyboardButton('Предпросмотр', callback_data='preview_prod')
        edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='edit_name')
        edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='edit_photo')
        edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='edit_desc')
        edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='edit_price')
        edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='edit_cate')
        show = types.InlineKeyboardButton('Показать\u2705', callback_data='show' + product[0])
        hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='hide' + product[0])
        ready = types.InlineKeyboardButton('Сохранить\ud83d\udcbe', callback_data='ready_prod')
        delete = types.InlineKeyboardButton('Удалить продукт\u274c', callback_data='delete_prod')
        disc_no = types.InlineKeyboardButton('Акция \u2611\ufe0f', callback_data='disc_no')
        adding_prod.add(ready, delete)
        if product[4] == '<b>ДА</b>':
            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, hide, disc_no)
        else:
            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show, disc_no)
        adding_prod.add(preview)
        bot.edit_message_text(chat_id=call.message.chat.id,
                              text='<b>Информация о продукте</b>\n\nИмя: ' + product[
                                  0] + '\nОписание: \n' + product[
                                       1] + '\nЦена: ' + str(product[2]) + '\nКатегория: ' + product[3] + '\nВидимость: ' +
                                   product[4] + '\nФото: ' + product[5], parse_mode='HTML',
                              message_id=call.message.message_id,
                              reply_markup=adding_prod)
    if call.data == 'disc_no':
        sql = 'UPDATE productlist SET disc = %s WHERE process = %s AND user_id = %s'
        val = ('<b>НЕТ</b>', 'making', call.message.chat.id)
        cursor.execute(sql, val)
        db.commit()
        sql = 'SELECT product, desc_, price, type, seen, link FROM productlist WHERE user_id = %s AND process = %s'
        val = (call.message.chat.id, 'making')
        cursor.execute(sql, val)
        product = cursor.fetchone()
        adding_prod = types.InlineKeyboardMarkup(2)
        preview = types.InlineKeyboardButton('Предпросмотр', callback_data='preview_prod')
        edit_name = types.InlineKeyboardButton('Ред. имя', callback_data='edit_name')
        edit_photo = types.InlineKeyboardButton('Ред. фото', callback_data='edit_photo')
        edit_desc = types.InlineKeyboardButton('Ред. описание', callback_data='edit_desc')
        edit_price = types.InlineKeyboardButton('Ред. цену', callback_data='edit_price')
        edit_cate = types.InlineKeyboardButton('Ред. категорию', callback_data='edit_cate')
        show = types.InlineKeyboardButton('Показать\u2705', callback_data='show' + product[0])
        hide = types.InlineKeyboardButton('Скрыть\u26d4', callback_data='hide' + product[0])
        ready = types.InlineKeyboardButton('Сохранить\ud83d\udcbe', callback_data='ready_prod')
        delete = types.InlineKeyboardButton('Удалить продукт\u274c', callback_data='delete_prod')
        disc = types.InlineKeyboardButton('Акция \u25fb\ufe0f', callback_data='disc')
        adding_prod.add(ready, delete)
        if product[4] == '<b>ДА</b>':
            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, hide, disc)
        else:
            adding_prod.add(edit_name, edit_photo, edit_desc, edit_cate, edit_price, show, disc)
        adding_prod.add(preview)
        bot.edit_message_text(chat_id=call.message.chat.id,
                              text='<b>Информация о продукте</b>\n\nИмя: ' + product[
                                  0] + '\nОписание: \n' + product[
                                       1] + '\nЦена: ' + str(product[2]) + '\nКатегория: ' + product[3] + '\nВидимость: ' +
                                   product[4] + '\nФото: ' + product[5], parse_mode='HTML',
                              message_id=call.message.message_id,
                              reply_markup=adding_prod)


while True:
    try:

        bot.polling(none_stop=True)

    except Exception as e:
        logging.error(e)
        time.sleep(5)